"""
checkers/excel.py
Reusable check functions for Excel (.xlsx) assignments.
All functions take filepath as first argument so base_marker can call them uniformly.
"""

import openpyxl
from openpyxl.chart import PieChart, ProjectedPieChart, DoughnutChart


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize(value):
    """Strip spaces and uppercase a string value."""
    if value is None:
        return ""
    return str(value).replace(" ", "").upper()


def _load(filepath, data_only=False):
    return openpyxl.load_workbook(filepath, data_only=data_only)


def _ws(filepath, sheet=None, data_only=False):
    """Return a worksheet. Uses active sheet if sheet name not provided."""
    wb = _load(filepath, data_only=data_only)
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in workbook.")
        return wb[sheet]
    return wb.active


# ── Cell value checks ─────────────────────────────────────────────────────────

def cell_value(filepath, cell, expected, sheet=None, exact=False):
    """
    Check if a cell contains an expected value.
    By default does a normalized (stripped, uppercase) comparison.
    Set exact=True for case-sensitive exact match.
    """
    ws = _ws(filepath, sheet)
    actual = ws[cell].value
    if exact:
        return str(actual) == str(expected)
    return normalize(actual) == normalize(expected)


def cell_contains(filepath, cell, substring, sheet=None):
    """Check if a cell value contains a substring (case-insensitive)."""
    ws = _ws(filepath, sheet)
    actual = normalize(ws[cell].value)
    return normalize(substring) in actual


def cell_starts_with(filepath, cell, prefix, sheet=None):
    """Check if a cell value starts with a given prefix."""
    ws = _ws(filepath, sheet)
    val = ws[cell].value
    if val is None:
        return False
    return str(val).startswith(prefix)


def cell_is_numeric(filepath, cell, sheet=None):
    """Check if a cell contains a numeric value."""
    ws = _ws(filepath, sheet)
    return isinstance(ws[cell].value, (int, float))


# ── Formula checks ────────────────────────────────────────────────────────────

def formula_contains(filepath, cell, expected, sheet=None):
    """
    Check if a cell formula contains an expected string.
    Uses data_only=False to read raw formulas.
    Example: formula_contains(fp, 'M2', '=SUM')
    """
    ws = _ws(filepath, sheet, data_only=False)
    val = normalize(ws[cell].value)
    return normalize(expected) in val


def formula_exact(filepath, cell, expected, sheet=None):
    """Check if a cell formula exactly matches expected (normalized)."""
    ws = _ws(filepath, sheet, data_only=False)
    val = normalize(ws[cell].value)
    return val == normalize(expected)


# ── Format checks ─────────────────────────────────────────────────────────────

def cell_is_bold(filepath, cell, sheet=None):
    """Check if a cell is bold."""
    ws = _ws(filepath, sheet)
    return ws[cell].font.bold is True


def cell_font_size(filepath, cell, expected_size, sheet=None):
    """Check if a cell has a specific font size."""
    ws = _ws(filepath, sheet)
    return ws[cell].font.size == expected_size


def cell_font_name(filepath, cell, expected_font, sheet=None):
    """Check if a cell uses a specific font name (case-insensitive)."""
    ws = _ws(filepath, sheet)
    actual = (ws[cell].font.name or "").upper()
    return actual == expected_font.upper()


def cell_number_format(filepath, cell, expected_format, sheet=None):
    """Check if a cell has a specific number format string."""
    ws = _ws(filepath, sheet)
    nf = (ws[cell].number_format or "").strip()
    return normalize(nf) == normalize(expected_format)


def cell_number_format_contains(filepath, cell, keyword, sheet=None):
    """Check if a cell number format contains a keyword."""
    ws = _ws(filepath, sheet)
    nf = (ws[cell].number_format or "").upper()
    return keyword.upper() in nf


def range_number_format(filepath, start_row, end_row, col, keywords, sheet=None):
    """
    Check if all cells in a column range have a number format containing
    at least one of the given keywords.
    Example: check C2:C45 are currency formatted.
    """
    ws = _ws(filepath, sheet)
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col}{row}"]
        nf = (cell.number_format or "").upper().replace(" ", "")
        if not any(k.upper() in nf for k in keywords):
            return False
    return True


def cell_has_background(filepath, cell, sheet=None, exclude=("FFFFFF", "000000", "")):
    """
    Check if a cell has a non-white/non-blank background fill.
    Matches your existing check_a1_a6_background logic.
    """
    ws = _ws(filepath, sheet)
    c = ws[cell]
    fill = c.fill
    if fill and fill.fill_type not in (None, "none"):
        fg = fill.fgColor
        if fg:
            if fg.type == "theme":
                return True
            color = fg.rgb.upper() if fg.type == "rgb" else ""
            if len(color) == 8:
                color = color[2:]
            return color not in exclude
    return False


def cell_alignment(filepath, cell, horizontal=None, vertical=None, sheet=None):
    """Check cell alignment. Pass horizontal='center' and/or vertical='center'."""
    ws = _ws(filepath, sheet)
    align = ws[cell].alignment
    if horizontal and (align.horizontal or "").lower() != horizontal.lower():
        return False
    if vertical and (align.vertical or "").lower() != vertical.lower():
        return False
    return True


# ── Merge checks ──────────────────────────────────────────────────────────────

def cells_merged(filepath, merge_range, sheet=None):
    """
    Check if a specific range is merged.
    Example: cells_merged(fp, 'A1:A6')
    """
    ws = _ws(filepath, sheet)
    target = normalize(merge_range)
    for merged in ws.merged_cells.ranges:
        if normalize(str(merged)) == target:
            return True
    return False


# ── Sheet checks ──────────────────────────────────────────────────────────────

def sheet_exists(filepath, sheet_name):
    """Check if a sheet with the given name exists (case-insensitive)."""
    wb = _load(filepath)
    return sheet_name.lower() in [s.lower() for s in wb.sheetnames]


def sheet_count(filepath, expected_count):
    """Check if the workbook has an exact number of sheets."""
    wb = _load(filepath)
    return len(wb.sheetnames) == expected_count


# ── Chart checks ──────────────────────────────────────────────────────────────

def chart_exists(filepath, sheet=None):
    """Check if at least one chart exists on the sheet."""
    ws = _ws(filepath, sheet)
    return len(ws._charts) > 0


def chart_is_pie(filepath, sheet=None):
    """Check if any chart on the sheet is a pie/doughnut chart."""
    ws = _ws(filepath, sheet)
    for chart in ws._charts:
        if isinstance(chart, (PieChart, ProjectedPieChart, DoughnutChart)):
            return True
    return False


def chart_uses_range(filepath, expected_range, sheet=None):
    """
    Check if any chart on the sheet references a specific data range.
    Example: chart_uses_range(fp, 'A2:E45')
    """
    ws = _ws(filepath, sheet)
    target = normalize(expected_range)
    for chart in ws._charts:
        for series in chart.series:
            for attr in ["val", "cat"]:
                try:
                    ref = normalize(str(getattr(series, attr).numRef.ref))
                    if ref == target:
                        return True
                except Exception:
                    pass
    return False


# ── Page setup checks ─────────────────────────────────────────────────────────

def page_is_landscape(filepath, sheet=None):
    """Check if the sheet page orientation is landscape."""
    ws = _ws(filepath, sheet)
    return ws.page_setup.orientation == "landscape"


def print_area_is(filepath, expected_range, sheet=None):
    """Check if the print area matches the expected range (normalized)."""
    ws = _ws(filepath, sheet)
    pa = ws.print_area or ""
    return normalize(pa) == normalize(expected_range)


def repeat_columns_is(filepath, expected, sheet=None):
    """
    Check if print title columns match expected value.
    Example: repeat_columns_is(fp, '$A:$A')
    """
    ws = _ws(filepath, sheet)
    titles = ws.print_title_cols or ""
    return normalize(titles) == normalize(expected)


def repeat_rows_is(filepath, expected, sheet=None):
    """
    Check if print title rows match expected value.
    Example: repeat_rows_is(fp, '$1:$1')
    """
    ws = _ws(filepath, sheet)
    titles = ws.print_title_rows or ""
    return normalize(titles) == normalize(expected)
