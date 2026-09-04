import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class CheckResult:
    status: str
    awarded: int | float | str
    reason: str


def normalize_formula(value: str) -> str:
    return re.sub(r"\s+", "", value or "").upper()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).upper()


def edit_distance(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(
                min(
                    prev[j] + 1,
                    curr[j - 1] + 1,
                    prev[j - 1] + (0 if ca == cb else 1),
                )
            )
        prev = curr
    return prev[-1]


def a1_anchor_to_tuple(anchor: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", anchor)
    if not match:
        return (-1, -1)
    letters, row = match.groups()
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    return (int(row), col)


def split_function_args(expr: str) -> List[str]:
    args: List[str] = []
    current: List[str] = []
    depth = 0
    for ch in expr:
        if ch == "," and depth == 0:
            args.append("".join(current).strip())
            current = []
            continue
        if ch == "(":
            depth += 1
        elif ch == ")" and depth > 0:
            depth -= 1
        current.append(ch)
    if current:
        args.append("".join(current).strip())
    return args


def parse_vlookup(formula: str) -> Optional[List[str]]:
    if not formula.startswith("=VLOOKUP(") or not formula.endswith(")"):
        return None
    inner = formula[len("=VLOOKUP("):-1]
    args = split_function_args(inner)
    return args if len(args) >= 3 else None


def normalize_ref(ref: str) -> str:
    return (ref or "").replace("$", "").upper()


def parse_range_ref(ref: str) -> Optional[tuple[tuple[int, int], tuple[int, int]]]:
    ref = normalize_ref(ref)
    if ":" not in ref:
        return None
    start, end = ref.split(":", 1)
    a = a1_anchor_to_tuple(start)
    b = a1_anchor_to_tuple(end)
    if -1 in a or -1 in b:
        return None
    return a, b


class Q3Workbook:
    def __init__(self, path: Path):
        self.path = path
        self.exists = path.exists()
        self.errors: List[str] = []

        self.sheetnames: List[str] = []
        self.main_sheet_title = ""
        self.sort_sheet_title = ""
        self.sort_sheet_is_last = False
        self.sort_sheet_tab_blue = False

        self.has_image = False
        self.image_anchor_a1 = False
        self.main_title = ""
        self.main_title_font = ""
        self.main_title_alignment = ""
        self.main_title_merged_d1_i1 = False
        self.row1_height: Optional[float] = None
        self.row2_height: Optional[float] = None
        self.widths_a_to_i: List[Optional[float]] = []
        self.header_details_visible = False
        self.f2_text = ""
        self.column_i_currency = False

        self.i5_formula = ""
        self.l5_formula = ""
        self.l6_formula = ""

        self.sorted_ok = False

        if self.exists:
            try:
                self._load()
            except Exception as exc:  # pragma: no cover
                self.errors.append(f"Q3 parse error: {exc}")
        else:
            self.errors.append("Q3 file not found")

    def _load(self) -> None:
        wb = load_workbook(self.path, data_only=False)
        self.sheetnames = wb.sheetnames
        for ws in wb.worksheets:
            if ws["D1"].value == "DIGITAL MARKETING":
                self.main_sheet_title = ws.title
            elif self._looks_like_sort_sheet(ws):
                self.sort_sheet_title = ws.title

        if not self.main_sheet_title and wb.worksheets:
            self.main_sheet_title = wb.worksheets[0].title
        if not self.sort_sheet_title and len(wb.worksheets) > 1:
            self.sort_sheet_title = wb.worksheets[-1].title

        self.sort_sheet_is_last = bool(self.sort_sheet_title) and self.sheetnames[-1] == self.sort_sheet_title

        if self.main_sheet_title:
            self._load_main_sheet(wb[self.main_sheet_title])
        if self.sort_sheet_title:
            self._load_sort_sheet(wb[self.sort_sheet_title])
        self._load_image_info()

    def _looks_like_sort_sheet(self, ws) -> bool:
        row1 = [ws.cell(1, c).value for c in range(1, 9)]
        expected = [
            "Platform",
            "Campaign_type",
            "Industry",
            "Country",
            "Times Shown",
            "Clicks",
            "Cost Per Purchase",
            "Retur on Ads",
        ]
        return row1 == expected

    def _load_main_sheet(self, ws) -> None:
        self.main_title = str(ws["D1"].value or "")
        self.main_title_font = str(ws["D1"].font.name or "")
        self.main_title_alignment = str(ws["D1"].alignment.horizontal or "")
        self.main_title_merged_d1_i1 = any(str(r) == "D1:I1" for r in ws.merged_cells.ranges)
        self.row1_height = ws.row_dimensions[1].height
        self.row2_height = ws.row_dimensions[2].height
        self.widths_a_to_i = [ws.column_dimensions[c].width for c in "ABCDEFGHI"]
        self.header_details_visible = all(
            str(ws[cell].value or "").strip()
            for cell in ("A2", "B2", "C2", "D2", "E2", "F2", "G2", "H2", "I2")
        )
        self.f2_text = str(ws["F2"].value or "")
        self.column_i_currency = self._number_format_is_currency(ws["I3"].number_format)
        self.i5_formula = str(ws["I5"].value or "")
        self.l5_formula = str(ws["L5"].value or "")
        self.l6_formula = str(ws["L6"].value or "")

    def _load_sort_sheet(self, ws) -> None:
        color = ws.sheet_properties.tabColor
        self.sort_sheet_tab_blue = False
        rgb = ""
        if color is not None:
            if getattr(color, "type", "") == "rgb" and color.rgb:
                rgb = str(color.rgb).upper()
                if len(rgb) == 8:
                    rgb = rgb[-6:]
                try:
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    self.sort_sheet_tab_blue = b >= max(r, g) and b >= 120
                except Exception:
                    self.sort_sheet_tab_blue = rgb.endswith("0070C0")
            elif getattr(color, "type", "") == "theme":
                # Office theme colour 4 is blue, with any tint/shade variation.
                self.sort_sheet_tab_blue = getattr(color, "theme", None) == 4
            elif getattr(color, "indexed", None) is not None:
                self.sort_sheet_tab_blue = int(color.indexed) in {4, 12, 41, 49}

        rows = []
        for row in range(2, ws.max_row + 1):
            industry = ws[f"C{row}"].value
            cpp = ws[f"G{row}"].value
            if industry is None and cpp is None:
                continue
            try:
                cpp_num = float(cpp)
            except Exception:
                self.sorted_ok = False
                return
            rows.append((str(industry or ""), cpp_num))
        self.sorted_ok = rows == sorted(rows, key=lambda item: (item[0].lower(), item[1]))

    def _load_image_info(self) -> None:
        with zipfile.ZipFile(self.path) as zf:
            media = [n for n in zf.namelist() if n.startswith("xl/media/")]
            self.has_image = bool(media)

        wb = load_workbook(self.path)
        if not self.main_sheet_title:
            return
        ws = wb[self.main_sheet_title]
        for img in getattr(ws, "_images", []):
            anchor = getattr(img, "anchor", None)
            from_marker = getattr(anchor, "_from", None)
            if from_marker is not None:
                self.image_anchor_a1 = from_marker.col == 0 and from_marker.row == 0
                if self.image_anchor_a1:
                    break

    def _number_format_is_currency(self, fmt: str) -> bool:
        fmt = str(fmt or "")
        return "R" in fmt and "0.00" in fmt


def evaluate_q3_check(doc: Q3Workbook, check: Dict) -> CheckResult:
    if not doc.exists:
        return CheckResult("manual", "", "Q3 file missing")
    if doc.errors:
        return CheckResult("manual", "", "; ".join(doc.errors))

    desc = check["description"]
    mark = check["mark"]
    i5 = normalize_formula(doc.i5_formula)
    l5 = normalize_formula(doc.l5_formula)
    l6 = normalize_formula(doc.l6_formula)
    title_norm = normalize_text(doc.main_title)
    expected_title = "DIGITAL MARKETING"

    def title_matches() -> bool:
        if title_norm == expected_title:
            return True
        if title_norm.replace(" ", "") == expected_title.replace(" ", ""):
            return True
        return edit_distance(title_norm, expected_title) <= 1

    def vlookup_arg_count(formula: str) -> int:
        args = parse_vlookup(formula)
        return len(args) if args else 0

    def vlookup_table_array_matches_data(formula: str) -> bool:
        args = parse_vlookup(formula)
        if not args:
            return False
        parsed = parse_range_ref(args[1])
        if not parsed:
            return False
        (start_row, start_col), (end_row, end_col) = parsed
        return start_col == 1 and end_col == 4 and start_row in {2, 3} and end_row in {83, 84}

    def vlookup_column_index(formula: str) -> Optional[int]:
        args = parse_vlookup(formula)
        if not args:
            return None
        try:
            return int(normalize_ref(args[2]))
        except Exception:
            return None

    def pass_fail(ok: bool, ok_reason: str, fail_reason: str) -> CheckResult:
        return CheckResult("pass" if ok else "fail", mark if ok else 0, ok_reason if ok else fail_reason)

    mapping = {
        "image": lambda: pass_fail(
            doc.has_image,
            "Workbook contains an image",
            "Workbook image not detected",
        ),
        "image 3Ads.png inserted in cell A1": lambda: pass_fail(
            doc.has_image and doc.image_anchor_a1,
            "Detected image anchored at A1",
            "Image anchored at A1 not detected",
        ),
        "merge and centre": lambda: pass_fail(
            doc.main_title_merged_d1_i1 and doc.main_title_alignment == "center",
            "Detected merged centered title block",
            "Merged centered title block not detected",
        ),
        "title ‘DIGITAL MARKETING’ inserted": lambda: pass_fail(
            title_matches(),
            "Main title matches DIGITAL MARKETING",
            f"Main title is {doc.main_title!r}",
        ),
        "in the font Algerian": lambda: pass_fail(
            doc.main_title_font.lower() == "algerian",
            "Main title font is Algerian",
            f"Main title font is {doc.main_title_font!r}",
        ),
        "merged and centred across cell range D1:I1": lambda: pass_fail(
            doc.main_title_merged_d1_i1 and doc.main_title_alignment == "center",
            "Detected D1:I1 merged and centered",
            "D1:I1 merged centered range not detected",
        ),
        "column widths and row heights": lambda: pass_fail(
            any((width or 0) != 13.0 for width in doc.widths_a_to_i) and (doc.row1_height or 0) >= 24 and (doc.row2_height or 0) >= 30,
            "Column widths and row heights were adjusted",
            f"Widths={doc.widths_a_to_i!r}, row1={doc.row1_height!r}, row2={doc.row2_height!r}",
        ),
        "the column widths and row heights of rows 1 and 2 adjusted": lambda: pass_fail(
            (doc.row1_height or 0) > 24 and (doc.row2_height or 0) >= 30,
            "Rows 1 and 2 heights were adjusted",
            f"Row heights are row1={doc.row1_height!r}, row2={doc.row2_height!r}",
        ),
        "all details are visible": lambda: pass_fail(
            doc.header_details_visible,
            "Header details are visible in row 2",
            "One or more row 2 headings are blank or unreadable",
        ),
        "Edit the text": lambda: pass_fail(
            bool(doc.f2_text),
            "Edited text present in F2",
            "Edited text not detected in F2",
        ),
        "text in cell F2 displays the same": lambda: pass_fail(
            doc.f2_text == "Times Shown",
            "F2 displays Times Shown",
            f"F2 displays {doc.f2_text!r}",
        ),
        "number format": lambda: pass_fail(
            doc.column_i_currency,
            "Currency format detected in column I",
            "Currency format not detected in column I",
        ),
        "number format currency assigned to column I": lambda: pass_fail(
            doc.column_i_currency,
            "Currency format detected in column I",
            "Currency format not detected in column I",
        ),
        "formula in cell I5": lambda: pass_fail(
            bool(i5),
            "Formula present in I5",
            "Formula not detected in I5",
        ),
        "=G5*H5": lambda: pass_fail(
            i5 == "=G5*H5",
            "I5 formula matches =G5*H5",
            f"I5 formula is {doc.i5_formula!r}",
        ),
        "correct function/operation: multiplication": lambda: pass_fail(
            "*" in i5 and "G5" in i5 and "H5" in i5,
            "I5 uses multiplication with G5 and H5",
            f"I5 formula is {doc.i5_formula!r}",
        ),
        "correct cell/cell range:  H5": lambda: pass_fail(
            "H5" in i5,
            "I5 references H5",
            f"I5 formula is {doc.i5_formula!r}",
        ),
        "correct cell/cell range: G5": lambda: pass_fail(
            "G5" in i5,
            "I5 references G5",
            f"I5 formula is {doc.i5_formula!r}",
        ),
        "function in cell L5": lambda: pass_fail(
            bool(l5),
            "Formula present in L5",
            "Formula not detected in L5",
        ),
        "=ROUNDUP(AVERAGE(H3:H83),1)": lambda: pass_fail(
            l5 == "=ROUNDUP(AVERAGE(H3:H83),1)",
            "L5 formula matches expected formula",
            f"L5 formula is {doc.l5_formula!r}",
        ),
        "correct function/operation:  ROUNDUP": lambda: pass_fail(
            "ROUNDUP(" in l5,
            "L5 uses ROUNDUP",
            f"L5 formula is {doc.l5_formula!r}",
        ),
        "correct function/operation: AVERAGE": lambda: pass_fail(
            "AVERAGE(" in l5,
            "L5 uses AVERAGE",
            f"L5 formula is {doc.l5_formula!r}",
        ),
        "correct cell/cell range: H3:H83": lambda: pass_fail(
            "H3:H83" in l5,
            "L5 references H3:H83",
            f"L5 formula is {doc.l5_formula!r}",
        ),
        "correct criteria: 1": lambda: pass_fail(
            l5.endswith(",1)") or l5.endswith(",1))"),
            "L5 uses criteria 1",
            f"L5 formula is {doc.l5_formula!r}",
        ),
        "error in cell L6": lambda: pass_fail(
            bool(l6),
            "Formula present in L6",
            "Formula not detected in L6",
        ),
        "=VLOOKUP(A8,A2:D83,4,FALSE)": lambda: pass_fail(
            l6 == "=VLOOKUP(A8,A2:D83,4,FALSE)",
            "L6 formula matches expected formula",
            f"L6 formula is {doc.l6_formula!r}",
        ),
        "correct table array: A2:D83": lambda: pass_fail(
            vlookup_table_array_matches_data(l6),
            "L6 table array matches the learner data block",
            f"L6 formula is {doc.l6_formula!r}",
        ),
        "correct column index number: 4": lambda: pass_fail(
            vlookup_column_index(l6) == 4 and vlookup_table_array_matches_data(l6),
            "L6 uses column index 4 against the learner data block",
            f"L6 formula is {doc.l6_formula!r}",
        ),
        "correct range_lookup: FALSE": lambda: pass_fail(
            l6.endswith("FALSE)") or ",FALSE)" in l6 or vlookup_arg_count(l6) == 3,
            "L6 uses FALSE range lookup or omits the optional fourth argument",
            f"L6 formula is {doc.l6_formula!r}",
        ),
        "rename the sheet": lambda: pass_fail(
            bool(doc.sort_sheet_title),
            "Sort sheet detected",
            "Sort sheet not detected",
        ),
        "the sheet renamed to ‘Sorting’": lambda: pass_fail(
            doc.sort_sheet_title == "Sorting",
            "Sheet renamed to Sorting",
            f"Sort sheet title is {doc.sort_sheet_title!r}",
        ),
        "moved to end": lambda: pass_fail(
            doc.sort_sheet_is_last,
            "Sort sheet is moved to the end",
            f"Sheet order is {doc.sheetnames!r}",
        ),
        "sheet tab colour": lambda: pass_fail(
            doc.sort_sheet_tab_blue,
            "Sort sheet tab colour is blue",
            "Sort sheet tab colour is not blue",
        ),
        "sheet tab colour blue": lambda: pass_fail(
            doc.sort_sheet_tab_blue,
            "Sort sheet tab colour is blue",
            "Sort sheet tab colour is not blue",
        ),
        "Sorting": lambda: pass_fail(
            bool(doc.sort_sheet_title),
            "Sorting sheet data detected",
            "Sorting sheet data not detected",
        ),
        "first alphabetically according to the industry": lambda: pass_fail(
            doc.sorted_ok,
            "Sorting sheet is sorted by industry then cost per purchase ascending",
            "Sorting sheet is not correctly sorted by industry and cost per purchase",
        ),
        "Secondly  according to the cost per purchase": lambda: pass_fail(
            doc.sorted_ok,
            "Sorting sheet is sorted by industry then cost per purchase ascending",
            "Sorting sheet is not correctly sorted by industry and cost per purchase",
        ),
        "Cost per purchace sorted ascending": lambda: pass_fail(
            doc.sorted_ok,
            "Sorting sheet is sorted by industry then cost per purchase ascending",
            "Sorting sheet is not correctly sorted by industry and cost per purchase",
        ),
    }

    if desc in mapping:
        return mapping[desc]()
    return CheckResult("manual", "", f"Q3 actual checker not implemented for {desc}")
