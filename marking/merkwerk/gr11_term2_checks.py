from __future__ import annotations

import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from zipfile import ZipFile

import openpyxl
from openpyxl.styles import Font, PatternFill
import win32com.client as win32


ROOT = Path(__file__).resolve().parent
PRACTICAL_DIR = ROOT / "Grade 11 Exam" / "Practical"
EXAM_DATA_DIR = ROOT / "Grade 11 Exam" / "Exam_Data"
MEMO_PATH = PRACTICAL_DIR / "Memo gr11.xlsx"
MARKSHEETS_DIR = ROOT / "Marksheets"
NOT_CHECKABLE_PATH = ROOT / "Not checkable.md"
REVIEW_PATH = ROOT / "review.md"
QUESTION_SHEETS = {"Question 1", "Question 2", "Question 3", "Question 4", "Question 5", "Question 6", "Question 7"}
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
MARK_FONT_COLOR = "FF0000"
REVIEW_ROWS = {}


@dataclass
class CheckResult:
    value: int | None
    note: str = ""


@dataclass
class LearnerContext:
    folder: Path
    files: dict[str, Path] = field(default_factory=dict)


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().lower()


def compact_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(text))


def compact_style_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def normalize_formula(formula: str) -> str:
    if formula is None:
        return ""
    formula = str(formula).strip()
    if formula.startswith("="):
        formula = formula[1:]
    if formula.startswith("+"):
        formula = formula[1:]
    return re.sub(r"\s+", "", formula).upper()


def extract_refs(formula: str) -> list[str]:
    return re.findall(r"\$?[A-Z]{1,3}\$?\d+", formula.upper())


def as_float(value) -> float | None:
    try:
        return float(value)
    except Exception:
        return None


def is_currency_or_accounting_format(number_format: str) -> bool:
    fmt = str(number_format or "").upper()
    return any(token in fmt for token in ["[$R", "R", "$", "ACCOUNTING"]) and "GENERAL" not in fmt


def field_prop_value(field, prop_name: str, default=None):
    try:
        return field.Properties(prop_name).Value
    except Exception:
        return default


def identify_docx(path: Path) -> str | None:
    try:
        xml = docx_xml_parts(path).get("word/document.xml", "")
        sample = normalize_text(" ".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml))[:4000])
        if (
            ("full name" in sample and "email address" in sample and "country" in sample)
            or (xml.count("<w:ffData>") >= 6 and xml.count("<w:checkBox>") >= 2 and xml.count("<w:textInput") >= 3)
        ):
            return "q7_word"
        if "global tourism stands among" in sample or "tourism evolution" in sample:
            return "q1_word"
        if "transport troubles around the world" in sample or "funny experiences in global travels" in sample:
            return "q2_word"
        return None
    except Exception:
        return None


def identify_xlsx(path: Path) -> str | None:
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        sheet_names = {str(name).strip().lower() for name in wb.sheetnames}
        if "famous places" in sheet_names:
            return "q3_excel"
        if "fatal accidents" in sheet_names:
            return "q4_excel"
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def identify_html(path: Path) -> str | None:
    text = path.read_text(encoding="utf-8", errors="ignore")
    norm = normalize_text(text)
    if "global tourism: an overview" in norm and "why people travel" in norm:
        return "q6_html"
    return None


def identify_accdb(path: Path) -> str | None:
    engine = win32.Dispatch("DAO.DBEngine.120")
    db = None
    try:
        db = engine.OpenDatabase(str(path))
        names = {db.TableDefs(i).Name.lower() for i in range(db.TableDefs.Count)}
        if "tblplaces" in names and "tbl5_1" in names:
            return "q5_access"
        return None
    finally:
        if db is not None:
            db.Close()


def discover_files(folder: Path) -> dict[str, Path]:
    found: dict[str, Path] = {}
    for path in folder.iterdir():
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        try:
            if suffix == ".xlsx":
                kind = identify_xlsx(path)
            elif suffix == ".docx":
                kind = identify_docx(path)
            elif suffix in {".htm", ".html"}:
                kind = identify_html(path)
            elif suffix in {".accdb", ".mdb"}:
                kind = identify_accdb(path)
            else:
                kind = None
        except Exception:
            kind = None
        if kind and kind not in found:
            found[kind] = path
    return found


def q3_checks(workbook_path: Path, excel_app=None) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    own_excel = excel_app is None
    excel = excel_app or win32.DispatchEx("Excel.Application")
    if own_excel:
        excel.Visible = False
        excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(workbook_path), ReadOnly=True)
        ws = wb.Worksheets("Famous Places")
        arrivals = wb.Worksheets(2)

        def put(row: int, value: bool):
            results[("Question 3", row)] = CheckResult(1 if value else 0)

        put(4, is_currency_or_accounting_format(str(ws.Range("E3").NumberFormat)))

        numfmt_f2 = str(ws.Range("F2").EntireColumn.NumberFormat)
        put(6, "0.00" in numfmt_f2)

        ws.Activate()
        put(8, bool(excel.ActiveWindow.FreezePanes and excel.ActiveWindow.SplitRow == 1))

        formula_e6 = normalize_formula(ws.Range("E6").Formula)
        refs_e6 = {ref.replace("$", "") for ref in extract_refs(formula_e6)}
        put(12, "/" in formula_e6 or "DIVIDE" in formula_e6)
        put(13, "F6" in refs_e6)
        put(14, "C6" in refs_e6)

        cf_ok = False
        cf_range_ok = False
        green_ok = False
        orange_ok = False
        red_ok = False
        try:
            fc = ws.Range("C2:C31").FormatConditions
            for idx in range(1, fc.Count + 1):
                condition = fc.Item(idx)
                if int(condition.Type) == 6:
                    cf_ok = True
                    applies_to = str(condition.AppliesTo.Address).replace("$", "")
                    cf_range_ok = "C2:C31" in applies_to
                    try:
                        criteria = condition.IconCriteria
                        if criteria.Count >= 3:
                            red = criteria.Item(1)
                            orange = criteria.Item(2)
                            green = criteria.Item(3)
                            red_icon = int(red.Icon)
                            orange_icon = int(orange.Icon)
                            green_icon = int(green.Icon)
                            orange_threshold = as_float(orange.Value)
                            green_threshold = as_float(green.Value)
                            orange_operator = int(orange.Operator)
                            green_operator = int(green.Operator)

                            # Excel icon sets define lower bounds for each icon.
                            # Correct setup: red below 5, orange from 5 to 10, green above 10.
                            red_ok = red_ok or (red_icon == 3 and orange_threshold == 5 and orange_operator == 7)
                            orange_ok = orange_ok or (
                                orange_icon == 2
                                and orange_threshold == 5
                                and green_threshold == 10
                                and orange_operator == 7
                            )
                            green_ok = green_ok or (green_icon == 1 and green_threshold == 10 and green_operator in {5, 7})
                    except Exception:
                        pass
        except Exception:
            pass
        put(17, cf_ok)
        put(18, green_ok)
        put(19, orange_ok)
        put(20, red_ok)
        put(21, cf_range_ok)

        formula_c33 = normalize_formula(ws.Range("C33").Formula)
        refs_c33 = {ref.replace("$", "") for ref in extract_refs(formula_c33)}
        uses_sum = "SUM(" in formula_c33
        uses_plus = "+" in formula_c33
        put(24, (uses_sum or uses_plus) and not (uses_sum and uses_plus))
        put(25, {"C2", "C11", "C27"} <= refs_c33)
        put(26, abs((ws.Range("C33").Value or 0) - 22) < 0.0001)

        formula_c34 = normalize_formula(ws.Range("C34").Formula)
        refs_c34 = {ref.replace("$", "") for ref in extract_refs(formula_c34)}
        put(29, "*" in formula_c34)
        put(30, "C9" in refs_c34)
        put(31, "1000000" in formula_c34)

        formula_c35 = normalize_formula(ws.Range("C35").Formula)
        put(32, "SMALL(E2:E31,30)" == formula_c35)

        formula_c36 = normalize_formula(ws.Range("C36").Formula)
        put(36, "MAX(E2:E31)" == formula_c36)

        formula_c37 = normalize_formula(ws.Range("C37").Formula)
        refs_c37 = {ref.replace("$", "") for ref in extract_refs(formula_c37)}
        put(42, formula_c37.startswith("COUNTIF("))
        put(43, "D2:D31" in formula_c37)
        put(44, '"YES"' in formula_c37 or "'YES'" in formula_c37)

        put(47, normalize_text(arrivals.Name) != "arrivals")
        try:
            put(48, arrivals.Tab.Color == 65535)
        except Exception:
            put(48, False)

        year_col = 1
        for col in range(1, arrivals.UsedRange.Columns.Count + 1):
            header = normalize_text(str(arrivals.Cells(1, col).Value or ""))
            if header == "year":
                year_col = col
                break
        years = [arrivals.Cells(r, year_col).Value for r in range(2, arrivals.UsedRange.Rows.Count + 1)]
        ordered = [y for y in years if isinstance(y, (int, float))]
        year_sorted_desc = len(ordered) > 1 and ordered == sorted(ordered, reverse=True)
        put(50, year_sorted_desc)
        put(51, year_sorted_desc)
    finally:
        if wb is not None:
            wb.Close(False)
        if own_excel:
            excel.Quit()
    return results


def q4_checks(workbook_path: Path, excel_app=None) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    own_excel = excel_app is None
    excel = excel_app or win32.DispatchEx("Excel.Application")
    if own_excel:
        excel.Visible = False
        excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(workbook_path), ReadOnly=True)
        sheet1 = wb.Worksheets("Sheet1")
        fatal = wb.Worksheets("Fatal Accidents")

        def put(row: int, value: bool):
            results[("Question 4", row)] = CheckResult(1 if value else 0)

        formula = normalize_formula(sheet1.Range("D20").Formula)
        put(6, formula.startswith('IF(I7>50000000,"HIGH","LOW")'))
        put(7, '"HIGH"' in formula)
        put(8, '"LOW"' in formula)

        chart_count = fatal.ChartObjects().Count
        if chart_count:
            chart = fatal.ChartObjects(1).Chart
            chart_type_ok = chart.SeriesCollection().Count >= 2
            put(11, chart_type_ok)
            series_name_parts = []
            for i in range(1, chart.SeriesCollection().Count + 1):
                try:
                    series_name_parts.append(str(chart.SeriesCollection(i).Name))
                except Exception:
                    continue
            series_names = " ".join(series_name_parts).lower()
            put(12, "hijacking" in series_names and "fatal" in series_names)
            labels_ok = False
            for i in range(1, chart.SeriesCollection().Count + 1):
                try:
                    labels_ok = labels_ok or bool(chart.SeriesCollection(i).HasDataLabels)
                except Exception:
                    continue
            put(13, labels_ok)
            put(14, bool(chart.HasTitle))
            axes_ok = False
            try:
                axes_ok = bool(chart.Axes(1).HasTitle and chart.Axes(2).HasTitle)
            except Exception:
                axes_ok = False
            put(15, axes_ok)
            legend_top = False
            if chart.HasLegend:
                legend = chart.Legend
                plot_area = chart.PlotArea
                legend_bottom = float(legend.Top) + float(legend.Height)
                plot_top = min(float(plot_area.Top), float(plot_area.InsideTop))
                chart_top = float(chart.ChartArea.Top)
                chart_height = float(chart.ChartArea.Height)
                legend_top = bool(
                    legend.Position == -4160
                    or legend_bottom <= plot_top + 8
                    or float(legend.Top) <= chart_top + (chart_height * 0.25)
                )
            put(16, legend_top)
        else:
            for row in range(11, 17):
                put(row, False)
    finally:
        if wb is not None:
            wb.Close(False)
        if own_excel:
            excel.Quit()
    return results


def dao_prop(obj, name: str, default=None):
    try:
        return getattr(obj, name)
    except Exception:
        pass
    try:
        return obj.Properties(name).Value
    except Exception:
        return default


def dao_field(table_def, *names: str):
    wanted = {compact_text(name) for name in names}
    for i in range(table_def.Fields.Count):
        field = table_def.Fields(i)
        if compact_text(field.Name) in wanted:
            return field
    return None


def access_sql_text(sql: str) -> str:
    return compact_text(sql.replace("[", " ").replace("]", " "))


def q5_form_checks(access_path: Path, results: dict[tuple[str, int], CheckResult], app=None) -> None:
    access_path = access_path.resolve()
    owns_app = app is None
    if owns_app:
        app = win32.DispatchEx("Access.Application")
    try:
        app.OpenCurrentDatabase(str(access_path))
        form_names = [app.CurrentProject.AllForms.Item(i).Name for i in range(app.CurrentProject.AllForms.Count)]
        best_form = None
        best_score = -1
        best_checks = {
            "title_ok": False,
            "image_ok": False,
            "footer_sources": [],
            "form_sources": [],
        }

        for candidate_name in form_names:
            try:
                app.DoCmd.OpenForm(candidate_name, 1)  # acDesign
                form = app.Forms(candidate_name)
                title_ok = False
                title_present = False
                image_ok = False
                footer_sources: list[str] = []
                form_sources: list[str] = []
                for i in range(form.Controls.Count):
                    control = form.Controls(i)
                    control_type = dao_prop(control, "ControlType")
                    section = dao_prop(control, "Section")
                    caption = normalize_text(str(dao_prop(control, "Caption", "")))
                    control_source = normalize_text(str(dao_prop(control, "ControlSource", "")))
                    if control_source:
                        form_sources.append(control_source)
                    if control_type == 100 and "form 5_2" in caption and "air passengers" in caption:
                        title_present = True
                        align = dao_prop(control, "TextAlign")
                        underline = bool(dao_prop(control, "FontUnderline", False))
                        title_ok = underline and align in {2, 3}
                    picture = normalize_text(str(dao_prop(control, "Picture", "")))
                    if control_type == 103 and "5aeroplane" in picture:
                        image_ok = True
                    if section == 2 and control_source:
                        footer_sources.append(control_source)

                footer_expr = compact_text(" ".join(footer_sources))
                form_expr = compact_text(" ".join(form_sources))
                form_content_present = title_present or (
                    "countryname" in form_expr
                    and "countrycode" in form_expr
                    and "airdepartures" in form_expr
                    and "aveticketprice" in form_expr
                )
                score = sum(
                    [
                        form_content_present,
                        title_ok,
                        image_ok,
                        "airdepartures" in footer_expr,
                        "*" in " ".join(footer_sources) or "multiply" in footer_expr,
                        "aveticketprice" in footer_expr,
                    ]
                )
                if compact_text(candidate_name) == "frm52":
                    score += 1
                if score > best_score:
                    best_score = score
                    best_form = candidate_name
                    best_checks = {
                        "form_content_present": form_content_present,
                        "title_ok": title_ok,
                        "image_ok": image_ok,
                        "footer_sources": footer_sources,
                        "form_sources": form_sources,
                    }
            except Exception:
                pass
            finally:
                try:
                    app.DoCmd.Close(2, candidate_name, 2)  # acForm, acSaveNo
                except Exception:
                    pass

        results[("Question 5", 23)] = CheckResult(1 if best_form and best_checks.get("form_content_present") else 0)
        if not best_form or best_score <= 0:
            for row in [25, 26, 27, 28, 29]:
                results[("Question 5", row)] = CheckResult(0)
            return

        title_ok = best_checks["title_ok"]
        image_ok = best_checks["image_ok"]
        footer_sources = best_checks["footer_sources"]
        form_sources = best_checks["form_sources"]
        footer_expr = compact_text(" ".join(footer_sources))
        form_expr = compact_text(" ".join(form_sources))
        results[("Question 5", 25)] = CheckResult(1 if title_ok else 0)
        results[("Question 5", 26)] = CheckResult(1 if image_ok else 0)
        results[("Question 5", 27)] = CheckResult(1 if "airdepartures" in footer_expr or "airdepartures" in form_expr else 0)
        results[("Question 5", 28)] = CheckResult(1 if "*" in " ".join(footer_sources + form_sources) or "multiply" in footer_expr or "multiply" in form_expr else 0)
        results[("Question 5", 29)] = CheckResult(1 if "aveticketprice" in footer_expr or "aveticketprice" in form_expr else 0)
    except Exception:
        for row in [23, 25, 26, 27, 28, 29]:
            results.setdefault(("Question 5", row), CheckResult(0))
    finally:
        try:
            app.CloseCurrentDatabase()
        except Exception:
            pass
        if owns_app:
            app.Quit()


def q5_checks(access_path: Path, access_app=None) -> dict[tuple[str, int], CheckResult]:
    access_path = access_path.resolve()
    results: dict[tuple[str, int], CheckResult] = {}

    def put(row: int, value: bool):
        results[("Question 5", row)] = CheckResult(1 if value else 0)

    engine = win32.Dispatch("DAO.DBEngine.120")
    db = None
    try:
        db = engine.OpenDatabase(str(access_path))
        tables = {compact_text(db.TableDefs(i).Name): db.TableDefs(i) for i in range(db.TableDefs.Count) if not db.TableDefs(i).Name.startswith("MSys")}
        places = tables.get("tblplaces")
        if places is not None:
            code = dao_field(places, "Code")
            heritage = dao_field(places, "Heritage?", "Heritage")
            region = dao_field(places, "Region")
            year_built = dao_field(places, "Year_Built", "Year Built")

            put(4, code is not None and dao_prop(code, "Size") == 6)
            put(7, heritage is not None and compact_text(heritage.Name) == "heritage")
            put(8, heritage is not None and dao_prop(heritage, "Type") == 1)
            primary_code = False
            for i in range(places.Indexes.Count):
                index = places.Indexes(i)
                if bool(dao_prop(index, "Primary", False)):
                    fields = {compact_text(index.Fields(j).Name) for j in range(index.Fields.Count)}
                    primary_code = "code" in fields
                    break
            put(9, primary_code)
            put(11, region is not None and bool(dao_prop(region, "Required", False)))
            default_value = compact_text(str(dao_prop(year_built, "DefaultValue", ""))) if year_built is not None else ""
            put(13, default_value == "2026")
            mask = str(dao_prop(code, "InputMask", "")) if code is not None else ""
            compact_mask = mask.replace("\\", "").replace(" ", "")
            put(17, ">" in compact_mask)
            put(18, "LL" in compact_mask.upper())
            put(19, "-" in compact_mask)
            put(20, "000" in compact_mask)
        else:
            for row in [4, 7, 8, 9, 11, 13, 17, 18, 19, 20]:
                put(row, False)

        query_sql = [db.QueryDefs(i).SQL for i in range(db.QueryDefs.Count) if not db.QueryDefs(i).Name.startswith("~")]
        compact_sql = [(sql, access_sql_text(sql), normalize_formula(sql)) for sql in query_sql]

        q53 = [item for item in compact_sql if "tbl51" in item[1] and "aveticketprice" in item[1]]
        put(33, any((">=50000" in item[2] or ">50000" in item[2] or "50000" in item[1]) for item in q53))
        put(34, any("50000" in item[1] for item in q53))
        put(35, any(all(field in item[1] for field in ["countryname", "countrycode", "aveticketprice"]) for item in q53))

        q54 = [item for item in compact_sql if "tbl51" in item[1] and "airarrivals" in item[1]]
        put(38, any("continent" in item[1] and "europe" in item[1] for item in q54))
        put(39, any("26000000" in item[1] for item in q54))
        put(40, any("40000000" in item[1] for item in q54))
        put(41, any(all(field in item[1] for field in ["countryname", "continent", "airdepartures", "airarrivals"]) for item in q54))

        q55 = [item for item in compact_sql if "tblplaces" in item[1] or "heritage" in item[1] or "monument" in item[1]]
        put(46, any("heritage" in item[1] and ("yes" in item[1] or "true" in item[1] or "-1" in item[1]) for item in q55))
        put(47, any("type" in item[1] and "monument" in item[1] for item in q55))
        put(48, any(all(field in item[1] for field in ["placename", "type", "heritage"]) for item in q55))
    except Exception:
        for row in [4, 7, 8, 9, 11, 13, 17, 18, 19, 20, 33, 34, 35, 38, 39, 40, 41, 46, 47, 48]:
            put(row, False)
    finally:
        if db is not None:
            db.Close()

    q5_form_checks(access_path, results, access_app)
    return results


def docx_xml_parts(docx_path: Path) -> dict[str, str]:
    parts: dict[str, str] = {}
    with ZipFile(docx_path) as archive:
        for name in archive.namelist():
            if (name.startswith("word/") or name.startswith("docProps/")) and name.endswith(".xml"):
                parts[name] = archive.read(name).decode("utf-8", errors="ignore")
    return parts


def paragraph_xml_by_text(document_xml: str, expected_text: str) -> str:
    expected = compact_text(expected_text)
    for para in re.findall(r"<w:p[\s\S]*?</w:p>", document_xml):
        text = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", para))
        if compact_text(text) == expected:
            return para
    return ""


def xml_text(xml: str) -> str:
    return "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml))


def run_has_text_and_prop(document_xml: str, expected: str, prop_pattern: str) -> bool:
    expected_compact = compact_text(expected)
    for run in re.findall(r"<w:r[\s\S]*?</w:r>", document_xml):
        if expected_compact in compact_text(xml_text(run)) and re.search(prop_pattern, run):
            return True
    return False


def all_word_instances_have_prop(document_xml: str, expected: str, prop_pattern: str) -> bool:
    chars: list[tuple[str, bool]] = []
    for run in re.findall(r"<w:r[\s\S]*?</w:r>", document_xml):
        has_prop = re.search(prop_pattern, run) is not None
        for char in xml_text(run):
            chars.append((char, has_prop))
    full_text = "".join(char for char, _ in chars).lower()
    expected_lower = expected.lower()
    matches = list(re.finditer(rf"\b{re.escape(expected_lower)}\b", full_text))
    return bool(matches) and all(all(chars[i][1] for i in range(match.start(), match.end())) for match in matches)


def meaningful_alt_text(value: str) -> bool:
    cleaned = normalize_text(value)
    default_titles = {
        "",
        "0",
        "picture",
        "picture 1",
        "image",
        "image 1",
        "graphic",
        "graphic 1",
    }
    return cleaned not in default_titles


def q1_checks(docx_path: Path) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    xml_parts = docx_xml_parts(docx_path)
    document_xml = xml_parts.get("word/document.xml", "")
    styles_xml = xml_parts.get("word/styles.xml", "")
    core_xml = xml_parts.get("docProps/core.xml", "")
    all_xml_text = normalize_text(" ".join(xml_text(xml) for xml in xml_parts.values()))
    footer_xml = "\n".join(xml for name, xml in xml_parts.items() if name.startswith("word/footer"))
    title_xml = paragraph_xml_by_text(document_xml, "Global Tourism")

    def put(row: int, value: bool):
        results[("Question 1", row)] = CheckResult(1 if value else 0)

    put(4, "<w:pStyle w:val=\"Title\"" in title_xml)
    title_has_shadow = "shadow" in title_xml.lower()
    put(5, title_has_shadow)
    put(6, title_has_shadow)

    heading_paras = [p for p in re.findall(r"<w:p[\s\S]*?</w:p>", document_xml) if "<w:pStyle w:val=\"Heading1\"" in p]
    heading_style_match = re.search(r'<w:style[^>]+w:styleId="Heading1"[\s\S]*?</w:style>', styles_xml)
    heading_style_xml = heading_style_match.group(0) if heading_style_match else ""
    put(8, "Comic Sans" in heading_style_xml or any("Comic Sans" in p for p in heading_paras))
    put(9, "<w:u " in heading_style_xml or any("<w:u " in p for p in heading_paras))

    author_match = re.search(r"<(?:[A-Za-z0-9_]+:)?creator[^>]*>(.*?)</(?:[A-Za-z0-9_]+:)?creator>", core_xml)
    author = normalize_text(author_match.group(1)) if author_match else ""
    author_control_changed = author not in {"", "admin", "cat examiner"} and "cat examiner" not in all_xml_text
    put(10, author_control_changed)

    body_text = normalize_text(xml_text(document_xml))
    guests_count = len(re.findall(r"\bguests\b", body_text))
    guests_found = guests_count > 0
    visitors_found = re.search(r"\bvisitors\b", body_text) is not None
    put(14, guests_count == 3)
    put(15, guests_found and not visitors_found)
    put(16, guests_count == 3 and all_word_instances_have_prop(document_xml, "guests", r"<w:b\b"))
    put(17, run_has_text_and_prop(document_xml, "guests", r"<w:u\b"))

    drawings = re.findall(r"<w:drawing[\s\S]*?</w:drawing>", document_xml)
    target_image = drawings[0] if drawings else ""
    extent_match = re.search(r"<wp:extent[^>]+cx=\"(\d+)\"[^>]+cy=\"(\d+)\"", target_image)
    cx = int(extent_match.group(1)) if extent_match else 0
    cy = int(extent_match.group(2)) if extent_match else 0
    put(19, bool(target_image) and "grayscl" not in target_image.lower() and "duotone" not in target_image.lower())
    put(20, cy == 2160000)
    put(21, cx == 2880000)
    put(22, "reflection" in target_image.lower())
    doc_pr_match = re.search(r"<wp:docPr[^>]*>", target_image)
    alt_values = []
    if doc_pr_match:
        alt_values = re.findall(r"descr=\"([^\"]*)\"", doc_pr_match.group(0))
    put(23, any(meaningful_alt_text(value) for value in alt_values))

    has_footer_page = "PAGE" in footer_xml or "fldCharType=\"begin\"" in footer_xml
    put(25, has_footer_page)
    put(26, "Page Numbers (Bottom of Page)" in footer_xml and "flowChartAlternateProcess" in footer_xml)
    put(27, "rightMargin" in footer_xml or 'w:jc w:val="right"' in footer_xml)
    different_first = "<w:titlePg" in document_xml
    put(28, different_first)
    first_footer_xml = xml_parts.get("word/footer2.xml", "")
    put(29, different_first and "PAGE" not in first_footer_xml)
    return results


def q2_checks(docx_path: Path) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    xml_parts = docx_xml_parts(docx_path)
    document_xml = xml_parts.get("word/document.xml", "")
    styles_xml = xml_parts.get("word/styles.xml", "")

    def put(row: int, value: bool):
        results[("Question 2", row)] = CheckResult(1 if value else 0)

    put(3, '<w:br w:type="page"' not in document_xml)

    drop_para = ""
    for para in re.findall(r"<w:p[\s\S]*?</w:p>", document_xml):
        if "w:dropCap" in para or "w:framePr" in para:
            drop_para = para
            break
    put(7, 'w:dropCap="margin"' in drop_para)
    put(8, "Arial Black" in drop_para)
    put(9, re.search(r'<w:color[^>]+w:val="(?:FF0000|C00000|FF0000)"', drop_para, flags=re.I) is not None or "red" in drop_para.lower())
    put(10, 'w:hSpace="284"' in drop_para or 'w:hSpace="283"' in drop_para)
    put(11, 'w:lines="8"' in drop_para)

    tables = re.findall(r"<w:tbl>[\s\S]*?</w:tbl>", document_xml)
    table_xml = tables[0] if tables else ""
    table_rows = re.findall(r"<w:tr[\s\S]*?</w:tr>", table_xml)
    header_row = table_rows[0] if table_rows else ""
    put(13, "<w:shd " in header_row)
    put(14, 'w:jc w:val="center"' in table_xml)
    put(15, "vMerge" in table_xml)
    merged_cells = re.findall(r"<w:tc>[\s\S]*?</w:tc>", table_xml)
    put(16, any("vMerge" in cell and "<w:drawing" in cell for cell in merged_cells))
    put(17, "vMerge" in table_xml or "gridSpan" in table_xml)

    toplabel_xml = ""
    toplabel_style_ids: set[str] = set()
    for style_xml in re.findall(r"<w:style[\s\S]*?</w:style>", styles_xml):
        style_id_match = re.search(r'w:styleId="([^"]+)"', style_xml)
        style_name_match = re.search(r'<w:name w:val="([^"]+)"', style_xml)
        style_id = style_id_match.group(1) if style_id_match else ""
        style_name = style_name_match.group(1) if style_name_match else ""
        variants = {compact_style_name(style_id), compact_style_name(style_name)}
        # Accept common spelling/spacing variants: TopLabel, Top Label, TopLevel, Top Level.
        if variants & {"toplabel", "toplevel"}:
            toplabel_xml = style_xml
            toplabel_style_ids.add(style_id)
            break

    style_paras = [
        para
        for para in re.findall(r"<w:p[\s\S]*?</w:p>", document_xml)
        if any(f'<w:pStyle w:val="{style_id}"' in para for style_id in toplabel_style_ids)
    ]
    bottom_border = "<w:pBdr" in toplabel_xml and "<w:bottom" in toplabel_xml
    bottom_border = bottom_border or any("<w:pBdr" in para and "<w:bottom" in para for para in style_paras)
    put(19, bool(toplabel_xml))
    put(20, '<w:basedOn w:val="Title"' in toplabel_xml)
    put(21, "shadow" in toplabel_xml.lower())
    put(22, bottom_border)
    return results


def q7_checks(docx_path: Path) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    document_xml = docx_xml_parts(docx_path).get("word/document.xml", "")
    fields = re.findall(r"<w:ffData>[\s\S]*?</w:ffData>", document_xml)

    def put(row: int, value: bool):
        results[("Question 7", row)] = CheckResult(1 if value else 0)

    has_tourism_graphic = bool(re.search(r"T\s*O\s*U\s*R\s*I\s*S\s*M|TOURISM", document_xml, flags=re.I))
    has_wordart_on_one_line = has_tourism_graphic and bool(re.search(r"<w:t[^>]*>\s*TOURISM\s*</w:t>", document_xml, flags=re.I))

    put(4, has_wordart_on_one_line)

    dropdowns = [field for field in fields if "<w:ddList>" in field]
    gender_dropdown = dropdowns[0] if dropdowns else ""
    entries = [
        normalize_text(match)
        for match in re.findall(r'<w:listEntry[^>]+w:val="([^"]*)"', gender_dropdown)
    ]
    text_fields = [field for field in fields if "<w:textInput" in field]
    full_name_field = text_fields[0] if text_fields else ""
    put(7, '<w:maxLength w:val="60"' in full_name_field)

    put(9, bool(gender_dropdown))
    put(10, bool(gender_dropdown))
    put(11, {"choose below", "male", "female"} <= set(entries))
    put(12, bool(entries) and entries[0] == "choose below")

    help_fields = [field for field in fields if "statusText" in field or "helpText" in field]
    help_text = " ".join(re.findall(r'<w:(?:statusText|helpText)[^>]+w:val="([^"]*)"', " ".join(help_fields)))
    put(15, "@" in help_text or "email" in normalize_text(help_text))

    put(18, any(re.search(r'<w:format[^>]+w:val="uppercase"', field, flags=re.I) for field in fields))

    date_fields = [field for field in fields if 'w:type w:val="date"' in field]
    date_field = date_fields[0] if date_fields else ""
    date_format_match = re.search(r'<w:format[^>]+w:val="([^"]*)"', date_field)
    date_format = normalize_text(date_format_match.group(1)) if date_format_match else ""
    put(20, bool(date_field))
    put(21, "dddd" in date_format and "mmmm" in date_format and "yyyy" in date_format)
    return results


class SimpleHtmlNestingParser(HTMLParser):
    VOID_TAGS = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.stack: list[str] = []
        self.errors = 0

    def handle_starttag(self, tag: str, attrs):
        if tag.lower() not in self.VOID_TAGS:
            self.stack.append(tag.lower())

    def handle_endtag(self, tag: str):
        tag = tag.lower()
        if tag in self.VOID_TAGS:
            return
        if not self.stack:
            self.errors += 1
            return
        if self.stack[-1] == tag:
            self.stack.pop()
            return
        if tag in self.stack:
            self.errors += 1
            while self.stack and self.stack[-1] != tag:
                self.stack.pop()
            if self.stack:
                self.stack.pop()
            return
        self.errors += 1

    @property
    def is_balanced(self) -> bool:
        ignored = {"html", "head", "body"}
        return self.errors == 0 and all(tag in ignored for tag in self.stack)


def html_attrs(tag_text: str) -> dict[str, str]:
    attrs: dict[str, str] = {}
    for name, quoted, bare in re.findall(r"([a-zA-Z_:][-a-zA-Z0-9_:.]*)\s*=\s*(?:[\"']([^\"']*)[\"']|([^\s>]+))", tag_text):
        attrs[name.lower()] = (quoted or bare).strip().strip("/")
    return attrs


def has_tagged_text(text: str, tag: str, expected_text: str) -> bool:
    expected = normalize_text(expected_text)
    pattern = rf"<{tag}\b[^>]*>(.*?)</{tag}>"
    for match in re.finditer(pattern, text, flags=re.I | re.S):
        inner = re.sub(r"<[^>]+>", "", match.group(1))
        if expected in normalize_text(inner):
            return True
    return False


def ordered_list_items(text: str) -> list[str]:
    items: list[str] = []
    for ordered_list in re.findall(r"<ol\b[^>]*>(.*?)</ol>", text, flags=re.I | re.S):
        for item in re.findall(r"<li\b[^>]*>(.*?)</li>", ordered_list, flags=re.I | re.S):
            items.append(normalize_text(re.sub(r"<[^>]+>", " ", item)))
    return items


def q6_checks(html_path: Path) -> dict[tuple[str, int], CheckResult]:
    results: dict[tuple[str, int], CheckResult] = {}
    text = html_path.read_text(encoding="utf-8", errors="ignore")
    lower = text.lower()

    def put(row: int, value: bool):
        results[("Question 6", row)] = CheckResult(1 if value else 0)

    title_match = re.search(r"<title\b[^>]*>(.*?)</title>", text, flags=re.I | re.S)
    title = normalize_text(re.sub(r"<[^>]+>", "", title_match.group(1))) if title_match else ""
    put(3, title == "global tourism")

    h1_match = re.search(r"<h1\b[^>]*>(.*?)</h1>", text, flags=re.I | re.S)
    h1_html = h1_match.group(1) if h1_match else ""
    h1_open = re.search(r"<h1\b[^>]*>", text, flags=re.I)
    h1_attrs_text = (h1_open.group(0) if h1_open else "") + " " + h1_html
    put(6, "verdana" in h1_attrs_text.lower())
    put(7, "brown" in h1_attrs_text.lower())

    put(10, has_tagged_text(text, "u", "Global tourism"))
    put(11, has_tagged_text(text, "b", "outside their usual environment"))
    put(12, has_tagged_text(text, "i", "leisure, business, education, religion, sport, or other purposes"))

    comments = re.findall(r"<!--(.*?)-->", text, flags=re.S)
    learner_comments = [
        comment
        for comment in comments
        if not re.search(r"\bquestion\s*6\.1\.\d+\b", normalize_text(comment), flags=re.I)
    ]
    target_phrases = {
        "thereasonswhypeopletravel",
        "thereasonwhypeopletravel",
        "reasonswhypeopletravel",
        "reasonwhypeopletravel",
        "whypeopletravel",
    }
    has_reasons_comment = any(
        any(phrase in compact_text(comment) for phrase in target_phrases)
        for comment in learner_comments
    )
    put(15, has_reasons_comment)
    put(16, has_reasons_comment)

    ol_items = ordered_list_items(text)
    put(19, bool(ol_items))
    put(20, any("conferences" in item for item in ol_items))
    put(21, any("seminars" in item for item in ol_items))

    hr_ok = False
    for hr in re.findall(r"<hr\b[^>]*>", text, flags=re.I):
        attrs = html_attrs(hr)
        width = attrs.get("width", "").replace(" ", "")
        color = attrs.get("color", "").lower()
        if width == "80%" and color == "green":
            hr_ok = True
            break
    put(22, hr_ok)

    put(24, bool(re.search(r"<br\b[^>]*\/?>|</br\s*>", text, flags=re.I)))

    smaller_heading = False
    for tag in ("h3", "h4", "h5", "h6"):
        if has_tagged_text(text, tag, "More Information"):
            smaller_heading = True
            break
    put(26, smaller_heading)

    parser = SimpleHtmlNestingParser()
    try:
        parser.feed(text)
        parser.close()
        put(29, parser.is_balanced)
    except Exception:
        put(29, False)

    return results


ALL_CHECKABLE_ROWS = {
    "Question 1": {4, 5, 6, 8, 9, 10, 14, 15, 16, 17, 19, 20, 21, 22, 23, 25, 26, 27, 28, 29},
    "Question 2": {3, 7, 8, 9, 10, 11, 13, 14, 15, 16, 17, 19, 20, 21, 22},
    "Question 3": {4, 6, 8, 12, 13, 14, 17, 18, 19, 20, 21, 24, 25, 26, 29, 30, 31, 32, 36, 42, 43, 44, 47, 48, 50, 51},
    "Question 4": {6, 7, 8, 11, 12, 13, 14, 15, 16},
    "Question 5": {4, 7, 8, 9, 11, 13, 17, 18, 19, 20, 23, 25, 26, 27, 28, 29, 33, 34, 35, 38, 39, 40, 41, 46, 47, 48},
    "Question 6": {3, 6, 7, 10, 11, 12, 15, 16, 19, 20, 21, 22, 24, 26, 29},
    "Question 7": {4, 7, 9, 10, 11, 12, 15, 18, 20, 21},
}


def all_rubric_rows() -> dict[str, list[int]]:
    wb = openpyxl.load_workbook(MEMO_PATH)
    rows: dict[str, list[int]] = {}
    for ws in wb.worksheets:
        if ws.title not in QUESTION_SHEETS:
            continue
        rows[ws.title] = []
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 2).value is not None or ws.cell(row, 1).value is not None:
                rows[ws.title].append(row)
    return rows


def write_marksheet(learner: LearnerContext, results: dict[tuple[str, int], CheckResult], rubric_rows: dict[str, list[int]]) -> None:
    MARKSHEETS_DIR.mkdir(exist_ok=True)
    out_path = MARKSHEETS_DIR / f"{learner.folder.name}.xlsx"
    shutil.copy2(MEMO_PATH, out_path)
    wb = openpyxl.load_workbook(out_path)

    def writable_cell(ws, row: int, column: int):
        cell = ws.cell(row, column)
        if cell.__class__.__name__ != "MergedCell":
            return cell
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col)
        return ws.cell(row, column)

    for sheet_name, rows in rubric_rows.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.column_dimensions["C"].width = ws.column_dimensions["D"].width
        markable_rows = ALL_CHECKABLE_ROWS.get(sheet_name, set()) | {row for sheet, row in REVIEW_ROWS if sheet == sheet_name}
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_col <= 3 <= merged_range.max_col:
                rows_in_range = [
                    row
                    for row in markable_rows
                    if merged_range.min_row <= row <= merged_range.max_row
                ]
                if len(rows_in_range) > 1:
                    ws.unmerge_cells(str(merged_range))
        for row in rows:
            key = (sheet_name, row)
            target = writable_cell(ws, row, 3)
            if key in results:
                target.value = results[key].value
            elif key in REVIEW_ROWS:
                target.value = None
            elif sheet_name in QUESTION_SHEETS and row in ALL_CHECKABLE_ROWS.get(sheet_name, set()):
                target.value = 0
        for row in range(1, ws.max_row + 1):
            source = ws.cell(row, 4)
            target = writable_cell(ws, row, 3)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.font = Font(
                name=target.font.name,
                sz=target.font.sz,
                b=target.font.b,
                i=target.font.i,
                vertAlign=target.font.vertAlign,
                underline=target.font.underline,
                strike=target.font.strike,
                color=MARK_FONT_COLOR,
                charset=target.font.charset,
                family=target.font.family,
                scheme=target.font.scheme,
                outline=target.font.outline,
                shadow=target.font.shadow,
                condense=target.font.condense,
                extend=target.font.extend,
            )
        for sheet_row in REVIEW_ROWS:
            if sheet_row[0] == sheet_name:
                writable_cell(ws, sheet_row[1], 3).fill = REVIEW_FILL
    wb.save(out_path)


def learner_folders() -> list[Path]:
    folders: list[Path] = []
    for grade_folder in [ROOT / group for group in ("F2", "F3", "RR2", "RR3")]:
        if not grade_folder.exists():
            continue
        folders.extend([p for p in grade_folder.iterdir() if p.is_dir()])
    return sorted(folders)


def main() -> None:
    rubric_rows = all_rubric_rows()
    if NOT_CHECKABLE_PATH.exists():
        NOT_CHECKABLE_PATH.unlink()
    access_app = win32.DispatchEx("Access.Application")
    try:
        for folder in learner_folders():
            learner = LearnerContext(folder=folder, files=discover_files(folder))
            results: dict[tuple[str, int], CheckResult] = {}

            if "q1_word" in learner.files:
                results.update(q1_checks(learner.files["q1_word"]))
            if "q2_word" in learner.files:
                results.update(q2_checks(learner.files["q2_word"]))
            if "q3_excel" in learner.files:
                results.update(q3_checks(learner.files["q3_excel"]))
            if "q4_excel" in learner.files:
                results.update(q4_checks(learner.files["q4_excel"]))
            if "q5_access" in learner.files:
                results.update(q5_checks(learner.files["q5_access"], access_app))
            if "q6_html" in learner.files:
                results.update(q6_checks(learner.files["q6_html"]))
            if "q7_word" in learner.files:
                results.update(q7_checks(learner.files["q7_word"]))

            write_marksheet(learner, results, rubric_rows)
        update_review_file()
    finally:
        try:
            access_app.Quit()
        except Exception:
            pass


def update_review_file() -> None:
    existing = REVIEW_PATH.read_text(encoding="utf-8") if REVIEW_PATH.exists() else "# Review\n"
    marker = "## Manual Review Rows"
    base = existing.split(marker)[0].rstrip()
    lines = [base, "", marker, ""]
    lines.append("These rows are highlighted yellow in each generated marksheet and should be resolved manually.")
    lines.append("")
    for sheet, row in sorted(REVIEW_ROWS):
        lines.append(f"- {sheet} row {row}: {REVIEW_ROWS[(sheet, row)]}")
    REVIEW_PATH.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
