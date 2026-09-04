import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
import win32com.client as win32

ROOT = Path(__file__).resolve().parent
Q7_DEFAULT_SOURCE = ROOT / "Gr12_TERM 2_Exam" / "Exam_Data" / "7Marketing Letter.docx"


@dataclass
class CheckResult:
    status: str
    awarded: int | float | str
    reason: str


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).lower()


class Q7Document:
    def __init__(self, learner_dir: Path):
        self.learner_dir = learner_dir
        self.letter_path = self._pick_first(
            [
                learner_dir / "7Marketing Letter.docx",
                learner_dir / "7Marketing_Letter.docx",
                learner_dir / "7Marketing Letters.docx",
            ]
        )
        self.client_list_path = learner_dir / "7Client List.xlsx"
        self.client_list_exists = self.client_list_path.exists()
        self.merged_path = None
        self.exists = bool(
            (self.letter_path and self.letter_path.exists())
            or self.client_list_exists
            or any(p.name != (self.letter_path.name if self.letter_path else "") for p in learner_dir.glob("*.docx"))
        )
        self.errors: List[str] = []

        self.letter_text = ""
        self.letter_xml = ""
        self.merged_text = ""
        self.merged_xml = ""

        self.has_date_field = False
        self.date_format_ok = False
        self.heading_text_ok = False
        self.heading2_style_ok = False
        self.letter_changed_from_default = False
        self.has_merge_fields = False
        self.has_client_name = False
        self.has_business_name = False
        self.has_email = False
        self.has_city = False
        self.has_merge_field_codes = False
        self.has_greeting = False
        self.has_cta = False
        self.references_business = False
        self.expected_selected_count = 0
        self.expected_selected_names: List[str] = []
        self.expected_selected_businesses: List[str] = []
        self.expected_cape_town_names: List[str] = []
        self.expected_pretoria_names: List[str] = []
        self.merged_exists = bool(self.merged_path and self.merged_path.exists())
        self.merged_content_present = False
        self.merged_letter_count = 0
        self.merged_has_only_target_cities = False
        self.merged_has_target_names = False
        self.merged_has_cape_town = False
        self.merged_has_pretoria = False
        self.merged_target_name_count = 0
        self.personalized_greeting_detected = False
        self.merged_has_multiple_letters = False
        self.letter_has_target_name = False
        self.com_letter_text = ""
        self.com_merged_text = ""
        self.letter_field_codes: List[str] = []
        self.merged_field_codes: List[str] = []
        self.heading2_style_ok_com = False
        self.has_greeting_com = False
        self.docx_cache: Dict[Path, tuple[str, str]] = {}

        if self.exists:
            try:
                self._load()
            except Exception as exc:  # pragma: no cover
                self.errors.append(f"Q7 parse error: {exc}")

    def _pick_first(self, paths: List[Path]) -> Optional[Path]:
        for path in paths:
            if path.exists():
                return path
        return paths[0] if paths else None

    def _read_docx(self, path: Path) -> tuple[str, str]:
        if path in self.docx_cache:
            return self.docx_cache[path]
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
        plain = re.sub(r"<[^>]+>", " ", xml)
        plain = re.sub(r"\s+", " ", plain)
        self.docx_cache[path] = (xml, plain)
        return xml, plain

    def _load(self) -> None:
        if self.letter_path and self.letter_path.exists():
            self.letter_xml, self.letter_text = self._read_docx(self.letter_path)
            self.letter_changed_from_default = self._letter_differs_from_default()

        self._load_client_list()
        self._find_merged_document()
        self._load_com_views()
        self._evaluate_letter()
        self._evaluate_merged()

    def _load_com_views(self) -> None:
        word = None
        try:
            word = win32.DispatchEx("Word.Application")
            word.Visible = False
            if self.letter_path and self.letter_path.exists():
                info = self._read_word_com(word, self.letter_path)
                self.com_letter_text = info["text"]
                self.letter_field_codes = info["field_codes"]
                self.heading2_style_ok_com = info["heading2"]
                self.has_greeting_com = info["greeting"]
            if self.merged_exists and self.merged_path:
                info = self._read_word_com(word, self.merged_path)
                self.com_merged_text = info["text"]
                self.merged_field_codes = info["field_codes"]
        except Exception:
            pass
        finally:
            if word is not None:
                word.Quit()

    def _read_word_com(self, word, path: Path) -> Dict:
        doc = None
        try:
            doc = word.Documents.Open(str(path.resolve()), False, True)
            text = doc.Range().Text.replace("\r", " ").replace("\x07", " ")
            field_codes = [field.Code.Text.replace("\r", " ").strip() for field in doc.Fields]
            heading2 = False
            greeting = False
            for para in doc.Paragraphs:
                ptext = para.Range.Text.replace("\r", " ").strip()
                try:
                    style_name = str(para.Range.Style)
                except Exception:
                    style_name = ""
                if "Digital Marketing Solutions for Your Business" in ptext and "Heading 2" in style_name:
                    heading2 = True
                if ptext.lower().startswith("dear"):
                    greeting = True
            return {"text": text, "field_codes": field_codes, "heading2": heading2, "greeting": greeting}
        finally:
            if doc is not None:
                doc.Close(False)

    def _load_client_list(self) -> None:
        if not self.client_list_exists:
            return
        wb = load_workbook(self.client_list_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in range(2, ws.max_row + 1):
            name = ws[f"A{row}"].value
            business = ws[f"B{row}"].value
            city = ws[f"D{row}"].value
            if city in {"Cape Town", "Pretoria"}:
                self.expected_selected_count += 1
                if name:
                    self.expected_selected_names.append(str(name))
                    if city == "Cape Town":
                        self.expected_cape_town_names.append(str(name))
                    if city == "Pretoria":
                        self.expected_pretoria_names.append(str(name))
                if business:
                    self.expected_selected_businesses.append(str(business))

    def _letter_differs_from_default(self) -> bool:
        if not self.letter_text:
            return False
        try:
            _, default_text = self._read_docx(Q7_DEFAULT_SOURCE)
        except Exception:
            return True
        return normalize_text(self.letter_text) != normalize_text(default_text)

    def _find_merged_document(self) -> None:
        candidates = []
        for path in sorted(self.learner_dir.glob("*.docx")):
            if self.letter_path and path.resolve() == self.letter_path.resolve():
                continue
            try:
                xml, text = self._read_docx(path)
            except Exception:
                continue
            heading_count = text.count("Digital Marketing Solutions for Your Business")
            dear_count = len(re.findall(r"\bDear\b", text, re.I))
            matched_names = sum(1 for name in self.expected_selected_names if name in text)
            business_hit = "BrightWave Marketing" in text
            score = heading_count * 5 + dear_count * 3 + matched_names * 8 + (4 if business_hit else 0)
            if score > 0:
                candidates.append((score, matched_names, heading_count, dear_count, path, xml, text))

        if not candidates:
            self.merged_path = self._pick_first(
                [
                    self.learner_dir / "7Merged_Letters.docx",
                    self.learner_dir / "7Merged Letters.docx",
                    self.learner_dir / "7Merged_Letter.docx",
                    self.learner_dir / "7Merged_ Letter.docx",
                    self.learner_dir / "7merged_letters.docx",
                ]
            )
            self.merged_exists = bool(self.merged_path and self.merged_path.exists())
            if self.merged_exists and self.merged_path:
                self.merged_xml, self.merged_text = self._read_docx(self.merged_path)
            return

        candidates.sort(key=lambda item: (item[0], item[1], item[2], item[3], item[4].name.lower()), reverse=True)
        _, _, _, _, best_path, best_xml, best_text = candidates[0]
        self.merged_path = best_path
        self.merged_exists = True
        self.merged_xml = best_xml
        self.merged_text = best_text

    def _evaluate_letter(self) -> None:
        combined_letter = f"{self.letter_text} {self.com_letter_text}"
        field_code_text = " || ".join(self.letter_field_codes).upper()
        self.has_date_field = " DATE " in combined_letter or "DATE \\" in combined_letter or "DATE  \\@" in combined_letter or "DATE" in field_code_text
        self.date_format_ok = 'dddd, dd,mm,yyyy' in combined_letter or 'dddd, dd,mm,yyyy' in self.letter_xml or 'DDDD, DD,MM,YYYY' in field_code_text
        self.heading_text_ok = "Digital Marketing Solutions for Your Business" in combined_letter
        self.heading2_style_ok = self.heading2_style_ok_com or bool(
            re.search(
                r'<w:p[^>]*>.*?<w:pStyle w:val="Heading2".*?Digital Marketing Solutions for Your Business',
                self.letter_xml,
                re.S,
            )
        ) or bool(
            re.search(
                r'Digital Marketing Solutions for Your Business.*?<w:pStyle w:val="Heading2"',
                self.letter_xml,
                re.S,
            )
        )
        merge_upper = combined_letter.upper()
        self.has_client_name = (
            "MERGEFIELD \"CLIENT_NAME\"" in merge_upper
            or "MERGEFIELD CLIENT_NAME" in field_code_text
            or "GREETINGLINE" in field_code_text
        )
        self.has_business_name = (
            "MERGEFIELD \"BUSINESS_NAME\"" in merge_upper
            or "MERGEFIELD BUSINESS_NAME" in field_code_text
        )
        self.has_email = (
            "MERGEFIELD \"EMAIL\"" in merge_upper
            or "MERGEFIELD EMAIL" in field_code_text
        )
        self.has_city = (
            "MERGEFIELD \"CITY\"" in merge_upper
            or "MERGEFIELD CITY" in field_code_text
        )
        self.has_merge_field_codes = any(token in field_code_text for token in ["MERGEFIELD CLIENT_NAME", "MERGEFIELD BUSINESS_NAME", "MERGEFIELD EMAIL", "MERGEFIELD CITY", "GREETINGLINE"])
        self.has_merge_fields = self.has_client_name and self.has_business_name and self.has_email and self.has_city
        greetingline_present = "GREETINGLINE" in merge_upper or "GREETINGLINE" in field_code_text
        real_mergefield_greeting = bool(re.search(r"\bdear\b", combined_letter, re.I)) and (
            "MERGEFIELD CLIENT_NAME" in field_code_text
            or "MERGEFIELD \"CLIENT_NAME\"" in merge_upper
            or "GREETINGLINE" in field_code_text
        )
        self.letter_has_target_name = any(name.upper() in merge_upper for name in self.expected_selected_names)
        actual_name_greeting = bool(re.search(r"\bdear\b", combined_letter, re.I)) and self.letter_has_target_name
        self.personalized_greeting_detected = greetingline_present or real_mergefield_greeting or actual_name_greeting
        self.has_greeting = self.personalized_greeting_detected
        self.has_cta = "contact us today to grow your online presence!" in normalize_text(combined_letter)
        merged_upper = f"{self.merged_text} {self.com_merged_text}".upper()
        merged_business_names = sum(1 for name in self.expected_selected_businesses if name.upper() in merged_upper)
        self.references_business = self.has_business_name or ("MERGEFIELD BUSINESS_NAME" in field_code_text) or merged_business_names >= 2

    def _evaluate_merged(self) -> None:
        if not self.merged_exists:
            return
        merged_combined = f"{self.merged_text} {self.com_merged_text}"
        self.merged_letter_count = merged_combined.count("Digital Marketing Solutions for Your Business")
        matched_names = [name for name in self.expected_selected_names if name in merged_combined]
        self.merged_target_name_count = len(matched_names)
        self.merged_has_cape_town = any(name in merged_combined for name in self.expected_cape_town_names)
        self.merged_has_pretoria = any(name in merged_combined for name in self.expected_pretoria_names)
        greeting_count = len(re.findall(r"\bDear\b", merged_combined, re.I))
        self.merged_has_multiple_letters = (
            self.merged_letter_count >= 2
            or greeting_count >= 2
            or self.merged_target_name_count >= 2
        )
        self.merged_content_present = self.merged_has_multiple_letters and self.merged_target_name_count >= 2
        self.merged_has_target_names = len(matched_names) >= min(self.expected_selected_count, 4) if self.expected_selected_count else False
        forbidden_cities = [city for city in ["Johannesburg", "Durban", "Port Elizabeth", "Bloemfontein"] if city in merged_combined]
        self.merged_has_only_target_cities = not forbidden_cities


def evaluate_q7_check(doc: Q7Document, check: Dict) -> CheckResult:
    if not doc.exists:
        return CheckResult("fail", 0 if check["mark"] else "", "Q7 evidence files missing")
    if doc.errors:
        return CheckResult("manual", "", "; ".join(doc.errors))

    desc = check["description"]
    mark = check["mark"]

    def pass_fail(ok: bool, ok_reason: str, fail_reason: str) -> CheckResult:
        return CheckResult("pass" if ok else "fail", mark if ok else 0, ok_reason if ok else fail_reason)

    mapping = {
        "date field": lambda: pass_fail(doc.has_date_field, "DATE field detected", "DATE field not detected"),
        "current date inserted": lambda: pass_fail(doc.has_date_field, "DATE field detected", "DATE field not detected"),
        "as an automatic field": lambda: pass_fail(doc.has_date_field, "Automatic DATE field detected", "Automatic DATE field not detected"),
        "displays in the format: ‘dddd, dd,mm,yyyy’": lambda: pass_fail(
            doc.date_format_ok,
            "DATE field format dddd, dd,mm,yyyy detected",
            "DATE field format dddd, dd,mm,yyyy not detected",
        ),
        "heading": lambda: pass_fail(doc.heading_text_ok, "Heading text detected", "Heading text not detected"),
        "heading: ‘Digital Marketing Solutions for Your Business’ inserted": lambda: pass_fail(
            doc.heading_text_ok and doc.letter_changed_from_default,
            "Heading text detected in an edited document",
            "Heading text not detected, or document still matches the default starter file",
        ),
        "heading: ‘Digital Marketing Solutions for Your Business’ Heading 2 style apply to it": lambda: pass_fail(
            doc.heading_text_ok and doc.heading2_style_ok and doc.letter_changed_from_default,
            "Heading text and Heading 2 style detected in an edited document",
            "Heading text and Heading 2 style not both detected, or document still matches the default starter file",
        ),
        "Heading 2 style apply to it": lambda: pass_fail(
            doc.heading2_style_ok and doc.letter_changed_from_default,
            "Heading 2 style detected on heading in an edited document",
            "Heading 2 style not detected on heading, or document still matches the default starter file",
        ),
        "mail merge": lambda: pass_fail(doc.has_merge_fields, "Mail merge fields detected", "Mail merge fields not detected"),
        "data source file 7Client List.xlsx used": lambda: pass_fail(
            doc.client_list_exists and (doc.has_merge_field_codes or doc.merged_content_present),
            "7Client List.xlsx appears linked to an actual mail merge",
            "7Client List.xlsx is not evidenced as linked to the mail merge",
        ),
        "personalised greeting included": lambda: pass_fail(
            doc.personalized_greeting_detected,
            "Personalised greeting merge detected",
            "Personalised greeting merge not detected",
        ),
        'personalised greeting included Expected "Dear \'<<?Client_Name>>(Merged field)\'" Exactly': lambda: pass_fail(
            doc.personalized_greeting_detected,
            "Personalised greeting merge detected",
            "Personalised greeting merge not detected",
        ),
        "Merged fields: Some merged fields pressent": lambda: pass_fail(
            doc.has_merge_field_codes,
            "At least one real merge field detected",
            "No real merge fields detected",
        ),
        "all Merge fields:  Client Name, Business Name, Email , City included": lambda: pass_fail(
            doc.has_merge_fields,
            "All required merge fields detected",
            "One or more required merge fields not detected",
        ),
        "fields Client Name, Business Name, Email , City included": lambda: pass_fail(
            doc.has_merge_fields,
            "Client Name, Business Name, Email and City fields detected",
            "One or more merge fields not detected",
        ),
        "Clients filtered from Cape Town ": lambda: pass_fail(
            doc.merged_has_cape_town,
            "Merged output includes Cape Town client(s)",
            "Merged output does not show Cape Town client(s)",
        ),
        "clients filtered OR Pretoria": lambda: pass_fail(
            doc.merged_has_pretoria,
            "Merged output includes Pretoria client(s)",
            "Merged output does not show Pretoria client(s)",
        ),
        "only clients from Cape Town  or Pretoria": lambda: pass_fail(
            doc.merged_content_present and doc.merged_has_only_target_cities and doc.merged_has_target_names,
            "Merged document appears limited to Cape Town and Pretoria clients",
            "Merged document does not appear limited to Cape Town and Pretoria clients",
        ),
        "included personalised message for each client": lambda: pass_fail(
            doc.merged_content_present and doc.merged_letter_count >= max(1, doc.expected_selected_count - 1),
            "Merged document contains letters for each selected client",
            f"Merged document count is {doc.merged_letter_count}, expected {doc.expected_selected_count}",
        ),
        "referencing their business": lambda: pass_fail(
            doc.references_business,
            "Business reference field detected",
            "Business reference field not detected",
        ),
        "a call-to-action is added": lambda: pass_fail(
            doc.has_cta,
            "Call-to-action detected",
            "Call-to-action not detected",
        ),
        "merge completed and individual letters for the identified clients issued": lambda: pass_fail(
            doc.merged_exists
            and doc.merged_content_present
            and doc.merged_has_multiple_letters
            and doc.merged_target_name_count >= 2,
            "Merged letters document detected with multiple identified client letters",
            f"Merged document evidence too weak: letters={doc.merged_letter_count}, matched_names={doc.merged_target_name_count}",
        ),
        "merged document present": lambda: pass_fail(
            doc.merged_exists and doc.merged_content_present and doc.merged_has_multiple_letters,
            "Merged document detected with at least two merged letters",
            f"Merged document evidence too weak: letters={doc.merged_letter_count}, matched_names={doc.merged_target_name_count}",
        ),
        "saved as 7Merged_Letters.docx": lambda: pass_fail(
            doc.merged_exists and doc.merged_path is not None and doc.merged_path.name.lower() == "7merged_letters.docx" and doc.merged_content_present,
            "Merged document saved as 7Merged_Letters.docx",
            f"Merged document name is {doc.merged_path.name if doc.merged_path else ''!r}",
        ),
    }

    if desc in mapping:
        return mapping[desc]()
    return CheckResult("manual", "", f"Q7 actual checker not implemented for {desc}")
