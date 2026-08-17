from datetime import datetime, timedelta

import pandas as pd
from flask import flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl.styles import Alignment, Font, PatternFill

from app.database import get_db, get_grades, get_groups, get_user_role, log_activity
from app.helper_attendance import (
    add_learner_note_entry,
    build_attendance_group_summary,
    get_attendance_data,
    get_all_term_days,
    get_last_21_days,
    get_term_dates,
    is_attendance_editable,
)
from app.helper_common import parse_module_names
from app.helper_theory import clone_bank_question_to_test, merge_bank_match_rows_into_test


def register_attendance_routes(app):
    def format_export_day_label(day):
        return datetime.strptime(day, "%Y-%m-%d").strftime("%m/%d")

    def style_attendance_sheet(worksheet):
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        present_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        absent_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        late_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center_alignment

        worksheet.freeze_panes = "F2"

        for row in worksheet.iter_rows(min_row=2):
            for index, cell in enumerate(row, start=1):
                cell.alignment = center_alignment if index >= 4 else Alignment(vertical="center")
                if index == 1:
                    worksheet.column_dimensions[cell.column_letter].width = 18
                elif index == 2:
                    worksheet.column_dimensions[cell.column_letter].width = 24
                elif index == 3:
                    worksheet.column_dimensions[cell.column_letter].width = 14
                elif index == 4:
                    worksheet.column_dimensions[cell.column_letter].width = 14
                else:
                    worksheet.column_dimensions[cell.column_letter].width = 12
                    if cell.value == "P":
                        cell.fill = present_fill
                    elif cell.value == "A":
                        cell.fill = absent_fill
                    elif cell.value == "L":
                        cell.fill = late_fill

    @app.route("/term_dates", methods=["GET", "POST"])
    def term_dates():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))
        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        if request.method == "POST":
            conn = get_db()
            cursor = conn.cursor()
            for term_num in range(1, 5):
                start = request.form.get(f"term{term_num}_start", "").strip()
                end = request.form.get(f"term{term_num}_end", "").strip()
                if start and end:
                    cursor.execute(
                        """
                        INSERT INTO term_dates (term, start_date, end_date)
                        VALUES (?, ?, ?)
                        ON CONFLICT(term) DO UPDATE SET start_date = excluded.start_date, end_date = excluded.end_date
                        """,
                        (term_num, start, end),
                    )
                else:
                    cursor.execute("DELETE FROM term_dates WHERE term = ?", (term_num,))
            conn.commit()
            conn.close()
            log_activity(username, "updated term dates")
            return redirect(url_for("attendance"))

        return redirect(url_for("attendance"))

    @app.route("/attendance")
    def attendance():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return "Access denied", 403

        selected_group = request.args.get("group")
        selected_grade = (request.args.get("grade") or "").strip()
        selected_teacher = (request.args.get("teacher") or "").strip() or None
        today_dt = datetime.now().date()
        today = today_dt.isoformat()
        year_start = today_dt.replace(month=1, day=1)
        term_days = sorted(day for day in get_all_term_days() if day.startswith(str(today_dt.year)))
        valid_week_starts = []
        if term_days:
            valid_week_starts = sorted(
                {
                    (
                        datetime.strptime(day, "%Y-%m-%d").date()
                        - timedelta(days=datetime.strptime(day, "%Y-%m-%d").date().weekday())
                    ).isoformat()
                    for day in term_days
                    if datetime.strptime(day, "%Y-%m-%d").date() <= today_dt
                }
            )
        week_start_param = (request.args.get("week_start") or "").strip()
        try:
            requested_week_start = datetime.strptime(week_start_param, "%Y-%m-%d").date() if week_start_param else today_dt
        except ValueError:
            requested_week_start = today_dt

        if requested_week_start.year != today_dt.year:
            requested_week_start = requested_week_start.replace(year=today_dt.year)
        requested_week_start = requested_week_start - timedelta(days=requested_week_start.weekday())

        if valid_week_starts:
            requested_week_start_str = requested_week_start.isoformat()
            current_index = 0
            for idx, week_value in enumerate(valid_week_starts):
                if week_value <= requested_week_start_str:
                    current_index = idx
                else:
                    break
            start_of_week = datetime.strptime(valid_week_starts[current_index], "%Y-%m-%d").date()
            prev_week_start_str = valid_week_starts[current_index - 1] if current_index > 0 else None
            next_week_start_str = valid_week_starts[current_index + 1] if current_index < len(valid_week_starts) - 1 else None
            has_prev_week = prev_week_start_str is not None
            has_next_week = next_week_start_str is not None
        else:
            start_of_week = requested_week_start
            if start_of_week < year_start:
                start_of_week = year_start
            if start_of_week > today_dt:
                start_of_week = today_dt - timedelta(days=today_dt.weekday())
                if start_of_week < year_start:
                    start_of_week = year_start

            prev_week_start = start_of_week - timedelta(days=7)
            next_week_start = start_of_week + timedelta(days=7)
            has_prev_week = prev_week_start >= year_start
            has_next_week = next_week_start <= today_dt
            prev_week_start_str = prev_week_start.isoformat() if has_prev_week else None
            next_week_start_str = next_week_start.isoformat() if has_next_week else None

        end_of_week = min(start_of_week + timedelta(days=6), today_dt)
        start_date = start_of_week.isoformat()
        end_date = end_of_week.isoformat()

        conn = get_db()
        cursor = conn.cursor()
        teacher_options = []
        grade_options = []
        class_scopes = []

        if role == "teacher":
            grade_options = get_grades(username)
            groups = get_groups(username, grade=selected_grade)
            selected_teacher = username
        else:
            cursor.execute(
                """
                SELECT DISTINCT teacher_username
                FROM users
                WHERE role = 'student'
                  AND teacher_username IS NOT NULL
                  AND teacher_username != ''
                ORDER BY teacher_username
                """
            )
            teacher_options = [row[0] for row in cursor.fetchall()]
            if selected_teacher and selected_teacher not in teacher_options:
                selected_teacher = None
            grade_options = get_grades(teacher_username=selected_teacher) if selected_teacher else get_grades()
            if selected_teacher:
                cursor.execute(
                    """
                    SELECT DISTINCT group_name
                    FROM users
                    WHERE role = 'student'
                      AND teacher_username = ?
                      AND group_name IS NOT NULL
                      AND group_name != ''
                      AND (? = '' OR grade = ?)
                    ORDER BY group_name
                    """,
                    (selected_teacher, selected_grade, selected_grade),
                )
                groups = [row[0] for row in cursor.fetchall()]
            else:
                groups = []
                cursor.execute(
                    """
                    SELECT DISTINCT teacher_username, group_name
                    FROM users
                    WHERE role = 'student'
                      AND teacher_username IS NOT NULL
                      AND teacher_username != ''
                      AND group_name IS NOT NULL
                      AND group_name != ''
                      AND (? = '' OR grade = ?)
                    ORDER BY teacher_username, group_name
                    """,
                    (selected_grade, selected_grade),
                )
                class_scopes = [
                    {"teacher_username": row[0], "group": row[1], "label": f"{row[1]} ({row[0]})"}
                    for row in cursor.fetchall()
                ]

        if selected_group and selected_group not in groups:
            selected_group = None

        days = []
        data = []
        daily_present_counts = {}
        daily_absent_counts = {}
        attendance_summary = []

        if selected_group:
            days, data = get_attendance_data(
                selected_group,
                start_date,
                end_date,
                teacher_username=selected_teacher,
            )
            daily_present_counts = {day: sum(1 for row in data if row["days"].get(day)) for day in days}
            daily_absent_counts = {day: len(data) - daily_present_counts[day] for day in days}
        else:
            if role == "admin" and not selected_teacher:
                for scope in class_scopes:
                    scope_summary = build_attendance_group_summary(
                        cursor,
                        [scope["group"]],
                        teacher_username=scope["teacher_username"],
                        days=get_last_21_days(),
                    )
                    if not scope_summary:
                        continue
                    item = scope_summary[0]
                    attendance_summary.append(
                        {
                            **item,
                            "teacher_username": scope["teacher_username"],
                            "label": scope["label"],
                        }
                    )
            else:
                attendance_summary = build_attendance_group_summary(
                    cursor,
                    groups,
                    teacher_username=selected_teacher,
                    days=get_last_21_days(),
                )
                for item in attendance_summary:
                    item["teacher_username"] = selected_teacher
                    item["label"] = item["group"] if role == "teacher" else f"{item['group']} ({selected_teacher})"

        if selected_group:
            cursor.execute(
                """
                SELECT date, group_name, reason, created_by, created_at
                FROM excluded_dates
                WHERE group_name IS NULL OR group_name = ?
                ORDER BY date DESC
                """,
                (selected_group,),
            )
        else:
            cursor.execute(
                """
                SELECT date, group_name, reason, created_by, created_at
                FROM excluded_dates
                ORDER BY date DESC
                """
            )
        excluded_dates = cursor.fetchall()
        cursor.execute("SELECT modules FROM question_bank_questions WHERE COALESCE(modules, '') != ''")
        module_names = []
        for (modules_text,) in cursor.fetchall():
            module_names.extend(parse_module_names(modules_text))
        module_names = sorted({item for item in module_names if item}, key=str.lower)

        module_notes = []
        if selected_group:
            if selected_teacher:
                cursor.execute(
                    """
                    SELECT id, note_date, module_name, progress_text, note_text, module_finished, generated_test_id, created_at
                    FROM class_module_notes
                    WHERE group_name = ? AND teacher_username = ?
                    ORDER BY note_date DESC, created_at DESC, id DESC
                    LIMIT 12
                    """,
                    (selected_group, selected_teacher),
                )
            else:
                cursor.execute(
                    """
                    SELECT id, note_date, module_name, progress_text, note_text, module_finished, generated_test_id, created_at
                    FROM class_module_notes
                    WHERE group_name = ? AND teacher_username IS NULL
                    ORDER BY note_date DESC, created_at DESC, id DESC
                    LIMIT 12
                    """,
                    (selected_group,),
                )
            module_notes = cursor.fetchall()
        conn.close()

        terms = get_term_dates()

        return render_template(
            "attendance.html",
            groups=groups,
            grade_options=grade_options,
            selected_grade=selected_grade,
            teacher_options=teacher_options,
            selected_teacher=selected_teacher,
            selected_group=selected_group,
            days=days,
            data=data,
            daily_present_counts=daily_present_counts,
            daily_absent_counts=daily_absent_counts,
            today=datetime.now().strftime("%Y-%m-%d"),
            week_start=start_date,
            week_end=end_date,
            prev_week_start=prev_week_start_str,
            next_week_start=next_week_start_str,
            has_prev_week=has_prev_week,
            has_next_week=has_next_week,
            excluded_dates=excluded_dates,
            terms=terms,
            attendance_summary=attendance_summary,
            module_names=module_names,
            module_notes=module_notes,
        )

    @app.route("/export/attendance")
    def export_attendance():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        conn = get_db()
        group = request.args.get("group")
        role = get_user_role(username)
        days, data = get_attendance_data(group, teacher_username=username if role == "teacher" else None)

        rows = []
        for row in data:
            base = {
                "Username": row["username"],
                "Name": row["name"],
                "Group": row["group"],
                "Total Days Absent": row["absent_days"],
                "Total Days Late": row["late_days"],
            }
            for day in days:
                val = row["days"].get(day)
                if not val:
                    base[format_export_day_label(day)] = "A"
                else:
                    base[format_export_day_label(day)] = "L" if val.get("late") else "P"
            rows.append(base)

        df = pd.DataFrame(rows)
        file_path = "attendance_export.xlsx"
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Attendance", index=False)
            style_attendance_sheet(writer.sheets["Attendance"])
        log_activity(username, "exported attendance")
        conn.close()
        return send_file(file_path, as_attachment=True)

    @app.route("/export_attendance_form")
    def export_attendance_form():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))
        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return "Access denied", 403
        all_groups = get_groups(username) if role == "teacher" else get_groups()
        return render_template("export_attendance.html", all_groups=all_groups)

    @app.route("/export_attendance_multi", methods=["POST"])
    def export_attendance_multi():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        selected_groups = request.form.getlist("groups")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")

        if not selected_groups:
            return "No groups selected", 400

        if not start_date or not end_date:
            return "Date range is required", 400
        if start_date > end_date:
            return "Start date must be before end date", 400

        conn = get_db()
        file_path = "multi_group_attendance_export.xlsx"
        role = get_user_role(username)
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for group in selected_groups:
                days, data = get_attendance_data(
                    group,
                    start_date,
                    end_date,
                    teacher_username=username if role == "teacher" else None,
                )
                attendance_days = set(days)
                rows = []
                for row in data:
                    base = {
                        "Username": row["username"],
                        "Name": row["name"],
                        "Group": row["group"],
                        "Total Days Absent": row["absent_days"],
                        "Total Days Late": row["late_days"],
                    }
                    for day in days:
                        val = row["days"].get(day)
                        if not val:
                            base[format_export_day_label(day)] = "A"
                        else:
                            base[format_export_day_label(day)] = "L" if val.get("late") else "P"
                    rows.append(base)
                df = pd.DataFrame(rows)
                sheet_name = group.replace("/", "_").replace("\\", "_")[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                style_attendance_sheet(writer.sheets[sheet_name])

        log_activity(username, "exported attendance by group")
        conn.close()
        response = send_file(file_path, as_attachment=True)
        response.headers["HX-Redirect"] = url_for("teacher_dashboard")
        return response

    @app.route("/reset_attendance", methods=["POST"])
    def reset_attendance():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        group = request.form.get("group")
        date = request.form.get("date")
        selected_teacher = (request.form.get("teacher") or "").strip() or None

        if not group or not date:
            return "Missing data", 400

        conn = get_db()
        cursor = conn.cursor()
        if selected_teacher:
            cursor.execute(
                """
                SELECT username
                FROM users
                WHERE group_name = ? AND role = 'student' AND teacher_username = ?
                """,
                (group, selected_teacher),
            )
        else:
            cursor.execute(
                """
                SELECT username FROM users WHERE group_name = ? AND role = 'student'
                """,
                (group,),
            )
        users = [u[0] for u in cursor.fetchall()]

        if users:
            placeholders = ",".join(["?"] * len(users))
            cursor.execute(
                f"""
                DELETE FROM login_history
                WHERE username IN ({placeholders}) AND date = ?
                """,
                (*users, date),
            )
            cursor.execute(
                f"""
                DELETE FROM attendance_override
                WHERE username IN ({placeholders}) AND date = ?
                """,
                (*users, date),
            )
            for user in users:
                add_learner_note_entry(cursor, user, f"Attendance reset for {date} in group {group}.", username)

        conn.commit()
        conn.close()
        return redirect(url_for("attendance", group=group, teacher=selected_teacher))

    @app.route("/mark_all_present", methods=["POST"])
    def mark_all_present():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        group = request.form.get("group")
        date = request.form.get("date")
        selected_teacher = (request.form.get("teacher") or "").strip() or None

        if not group or not date:
            return "Missing data", 400

        conn = get_db()
        cursor = conn.cursor()
        if selected_teacher:
            cursor.execute(
                """
                SELECT username
                FROM users
                WHERE group_name = ? AND role = 'student' AND teacher_username = ?
                """,
                (group, selected_teacher),
            )
        else:
            cursor.execute(
                """
                SELECT username FROM users WHERE group_name = ? AND role = 'student'
                """,
                (group,),
            )
        users = [u[0] for u in cursor.fetchall()]

        if users:
            for user in users:
                cursor.execute("SELECT status FROM attendance_override WHERE username = ? AND date = ?", (user, date))
                previous = cursor.fetchone()
                cursor.execute(
                    """
                    INSERT INTO attendance_override (username, date, status)
                    VALUES (?, ?, ?)
                    ON CONFLICT(username, date)
                    DO UPDATE SET status = excluded.status
                    """,
                    (user, date, "present"),
                )
                if previous is None or previous[0] != "present":
                    add_learner_note_entry(cursor, user, f"Marked present for {date} in group {group}.", username)

        conn.commit()
        log_activity(username, f"marked all in {group} present on {date}")
        conn.close()
        return redirect(url_for("attendance", group=group, teacher=selected_teacher))

    @app.route("/save_attendance", methods=["POST"])
    def save_attendance():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        group = request.form.get("group")
        selected_teacher = (request.form.get("teacher") or "").strip() or None

        conn = get_db()
        cursor = conn.cursor()

        for key, value in request.form.items():
            if not key.startswith("att_"):
                continue
            try:
                _, data = key.split("att_")
                user, day = data.split("|")
            except ValueError:
                continue

            value = value.strip().lower()
            if value == "":
                continue

            if value in {"x", "a", "absent"}:
                status = "absent"
            elif value in {"l", "late"}:
                status = "late"
            else:
                status = "present"

            cursor.execute("SELECT status FROM attendance_override WHERE username = ? AND date = ?", (user, day))
            previous = cursor.fetchone()
            cursor.execute(
                """
                INSERT INTO attendance_override (username, date, status)
                VALUES (?, ?, ?)
                ON CONFLICT(username, date)
                DO UPDATE SET status = excluded.status
                """,
                (user, day, status),
            )
            if previous is None or previous[0] != status:
                add_learner_note_entry(cursor, user, f"Attendance manually set to {status.upper()} for {day}.", username)

        conn.commit()
        log_activity(username, f"saved attendance for {group}")
        conn.close()
        return redirect(url_for("attendance", group=group, teacher=selected_teacher))

    @app.route("/attendance/module_note", methods=["POST"])
    def attendance_module_note():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return "Access denied", 403

        group = (request.form.get("group") or "").strip()
        selected_teacher = (request.form.get("teacher") or "").strip() or None
        note_date = (request.form.get("date") or "").strip()
        module_name = (request.form.get("module_name") or "").strip()
        progress_text = (request.form.get("progress_text") or "").strip()
        note_text = (request.form.get("note_text") or "").strip()
        module_finished = 1 if request.form.get("module_finished") else 0
        week_start = (request.form.get("week_start") or "").strip()

        redirect_kwargs = {"group": group}
        if selected_teacher:
            redirect_kwargs["teacher"] = selected_teacher
        if week_start:
            redirect_kwargs["week_start"] = week_start

        if not group or not note_date or not module_name:
            flash("Date, group, and module are required.", "error")
            return redirect(url_for("attendance", **redirect_kwargs))

        conn = get_db()
        cursor = conn.cursor()
        generated_test_id = None

        if module_finished:
            cursor.execute(
                """
                SELECT id
                FROM theory_tests
                WHERE LOWER(COALESCE(generated_module_name, '')) = LOWER(?)
                ORDER BY id DESC
                LIMIT 1
                """,
                (module_name,),
            )
            existing_test = cursor.fetchone()
            if existing_test:
                generated_test_id = existing_test[0]
            else:
                cursor.execute(
                    """
                    SELECT id, question_type, subject, question_text, COALESCE(modules, '')
                    FROM question_bank_questions
                    ORDER BY id
                    """
                )
                bank_rows = cursor.fetchall()
                matching_rows = [
                    row for row in bank_rows
                    if module_name.lower() in {item.lower() for item in parse_module_names(row[4] or "")}
                ]
                if not matching_rows:
                    conn.close()
                    flash(f"No question bank questions found for module '{module_name}'.", "error")
                    return redirect(url_for("attendance", **redirect_kwargs))

                subjects = [row[2] for row in matching_rows if row[2]]
                subject_label = subjects[0] if len(set(subjects)) == 1 else (subjects[0] if subjects else "Theory")
                cursor.execute(
                    """
                    INSERT INTO theory_tests
                        (title, subject, assign_date, time_limit, allow_multiple, max_attempts, show_answers, created_by, created_at, is_active, generated_module_name)
                    VALUES (?, ?, ?, 0, 0, 1, 1, ?, ?, 1, ?)
                    """,
                    (
                        f"{module_name} Class Test",
                        subject_label,
                        note_date,
                        username,
                        datetime.now().isoformat(),
                        module_name,
                    ),
                )
                generated_test_id = cursor.lastrowid

                order_index = 1
                non_match_rows = [row for row in matching_rows if row[1] != "match"]
                match_ids = [row[0] for row in matching_rows if row[1] == "match"]
                for bank_question_id, _question_type, _subject, _question_text, _modules in non_match_rows:
                    clone_bank_question_to_test(cursor, bank_question_id, generated_test_id, order_index)
                    order_index += 1
                if match_ids:
                    merge_bank_match_rows_into_test(cursor, match_ids, generated_test_id, order_index)

            cursor.execute(
                """
                INSERT OR IGNORE INTO theory_test_groups (test_id, group_name) VALUES (?, ?)
                """,
                (generated_test_id, group),
            )
            if selected_teacher:
                cursor.execute(
                    """
                    INSERT OR IGNORE INTO theory_test_teachers (test_id, teacher_username) VALUES (?, ?)
                    """,
                    (generated_test_id, selected_teacher),
                )
            cursor.execute(
                """
                UPDATE theory_tests
                SET is_active = 1
                WHERE id = ?
                """,
                (generated_test_id,),
            )

        cursor.execute(
            """
            INSERT INTO class_module_notes
                (note_date, group_name, teacher_username, module_name, progress_text, note_text, module_finished, generated_test_id, created_by, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                note_date,
                group,
                selected_teacher,
                module_name,
                progress_text,
                note_text,
                module_finished,
                generated_test_id,
                username,
                datetime.now().isoformat(),
            ),
        )
        conn.commit()
        conn.close()

        if generated_test_id:
            flash(f"Module note saved and class test assigned for {module_name}.", "success")
            log_activity(username, f"finished module {module_name} for {group} and generated theory test {generated_test_id}")
        else:
            flash(f"Module note saved for {module_name}.", "success")
            log_activity(username, f"saved module note for {module_name} in {group}")
        return redirect(url_for("attendance", **redirect_kwargs))

    @app.route("/attendance/update_cell", methods=["POST"])
    def update_attendance_cell():
        username = session.get("username")
        if not username:
            return jsonify({"ok": False, "error": "Not logged in"}), 401

        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return jsonify({"ok": False, "error": "Access denied"}), 403

        payload = request.get_json(silent=True) or {}
        learner_username = (payload.get("username") or "").strip()
        date = (payload.get("date") or "").strip()
        value = (payload.get("value") or "").strip().lower()
        group = (payload.get("group") or "").strip()
        selected_teacher = (payload.get("teacher") or "").strip() or None

        if not learner_username or not date or not group:
            return jsonify({"ok": False, "error": "Missing data"}), 400

        if role != "admin":
            try:
                if not is_attendance_editable(date, role=role):
                    return jsonify({"ok": False, "error": "Attendance is locked for this date"}), 400
            except Exception:
                pass

        if value in {"x", "a", "absent"}:
            status = "absent"
        elif value in {"l", "late"}:
            status = "late"
        elif value in {"p", "present"}:
            status = "present"
        else:
            return jsonify({"ok": False, "error": "Use P, L or X"}), 400

        conn = get_db()
        cursor = conn.cursor()
        if selected_teacher:
            cursor.execute(
                """
                SELECT 1
                FROM users
                WHERE username = ? AND group_name = ? AND role = 'student' AND teacher_username = ?
                """,
                (learner_username, group, selected_teacher),
            )
        else:
            cursor.execute(
                """
                SELECT 1
                FROM users
                WHERE username = ? AND group_name = ? AND role = 'student'
                """,
                (learner_username, group),
            )
        if not cursor.fetchone():
            conn.close()
            return jsonify({"ok": False, "error": "Learner not found in selected class"}), 404

        cursor.execute("SELECT status FROM attendance_override WHERE username = ? AND date = ?", (learner_username, date))
        previous = cursor.fetchone()
        previous_status = previous[0] if previous else None
        cursor.execute(
            """
            INSERT INTO attendance_override (username, date, status)
            VALUES (?, ?, ?)
            ON CONFLICT(username, date)
            DO UPDATE SET status = excluded.status
            """,
            (learner_username, date, status),
        )
        if previous_status != status:
            add_learner_note_entry(
                cursor,
                learner_username,
                f"Attendance manually set to {status.upper()} for {date}.",
                username,
            )
        conn.commit()
        conn.close()
        log_activity(username, f"updated attendance for {learner_username} on {date} to {status}")

        display = "L" if status == "late" else ("P" if status == "present" else "✖")
        bg = "#ffe0b2" if status == "late" else ("#c8f7c5" if status == "present" else "#f7c5c5")
        return jsonify(
            {
                "ok": True,
                "status": status,
                "display": display,
                "background": bg,
                "counts_as_present": status in {"present", "late"},
                "previous_counts_as_present": (previous_status or "") in {"present", "late"},
            }
        )

    @app.route("/exclude_date", methods=["POST"])
    def exclude_date():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        date = request.form.get("date")
        group = request.form.get("group")
        reason = request.form.get("reason", "")

        if not date:
            return "Missing date", 400

        if group == "":
            group = None

        conn = get_db()
        cursor = conn.cursor()

        timestamp = datetime.now().isoformat()
        if group is None:
            cursor.execute(
                """
                SELECT rowid
                FROM excluded_dates
                WHERE date = ? AND group_name IS NULL
                ORDER BY rowid
                LIMIT 1
                """,
                (date,),
            )
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    """
                    UPDATE excluded_dates
                    SET reason = ?, created_by = ?, created_at = ?
                    WHERE rowid = ?
                    """,
                    (reason, username, timestamp, existing[0]),
                )
                cursor.execute(
                    """
                    DELETE FROM excluded_dates
                    WHERE date = ? AND group_name IS NULL AND rowid != ?
                    """,
                    (date, existing[0]),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO excluded_dates (date, group_name, reason, created_by, created_at)
                    VALUES (?, NULL, ?, ?, ?)
                    """,
                    (date, reason, username, timestamp),
                )
        else:
            cursor.execute(
                """
                INSERT INTO excluded_dates (date, group_name, reason, created_by, created_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(date, group_name)
                DO UPDATE SET reason = excluded.reason, created_by = excluded.created_by, created_at = excluded.created_at
                """,
                (date, group, reason, username, timestamp),
            )

        conn.commit()
        log_activity(username, f"excluded date {date} for group {group or 'all groups'} ({reason})")
        conn.close()
        return redirect(
            url_for(
                "attendance",
                group=request.form.get("selected_group"),
                teacher=(request.form.get("selected_teacher") or "").strip() or None,
            )
        )

    @app.route("/include_date", methods=["POST"])
    def include_date():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        date = request.form.get("date")
        group = request.form.get("group")

        if not date:
            return "Missing date", 400

        if group == "":
            group = None

        conn = get_db()
        cursor = conn.cursor()
        if group:
            cursor.execute("DELETE FROM excluded_dates WHERE date = ? AND group_name = ?", (date, group))
        else:
            cursor.execute("DELETE FROM excluded_dates WHERE date = ? AND group_name IS NULL", (date,))

        conn.commit()
        log_activity(username, f"included date {date} for group {group or 'all groups'}")
        conn.close()
        return redirect(
            url_for(
                "attendance",
                group=request.form.get("selected_group"),
                teacher=(request.form.get("selected_teacher") or "").strip() or None,
            )
        )

    @app.route("/mark_all_absent", methods=["POST"])
    def mark_all_absent():
        username = session.get("username")
        if not username:
            return redirect(url_for("login"))

        if get_user_role(username) not in ["teacher", "admin"]:
            return "Access denied", 403

        group = request.form.get("group")
        date = request.form.get("date")
        selected_teacher = (request.form.get("teacher") or "").strip() or None

        if not group or not date:
            return "Missing data", 400

        conn = get_db()
        cursor = conn.cursor()
        if selected_teacher:
            cursor.execute(
                """
                SELECT username
                FROM users
                WHERE group_name = ? AND role = 'student' AND teacher_username = ?
                """,
                (group, selected_teacher),
            )
        else:
            cursor.execute(
                """
                SELECT username FROM users WHERE group_name = ? AND role = 'student'
                """,
                (group,),
            )
        users = [u[0] for u in cursor.fetchall()]

        if users:
            for user in users:
                cursor.execute("SELECT status FROM attendance_override WHERE username = ? AND date = ?", (user, date))
                previous = cursor.fetchone()
                cursor.execute(
                    """
                    INSERT INTO attendance_override (username, date, status)
                    VALUES (?, ?, ?)
                    ON CONFLICT(username, date)
                    DO UPDATE SET status = excluded.status
                    """,
                    (user, date, "absent"),
                )
                if previous is None or previous[0] != "absent":
                    add_learner_note_entry(cursor, user, f"Marked absent for {date} in group {group}.", username)

        conn.commit()
        log_activity(username, f"marked all in {group} absent on {date}")
        conn.close()
        return redirect(url_for("attendance", group=group, teacher=selected_teacher))

    @app.route("/api/excluded_dates", methods=["GET"])
    def api_excluded_dates():
        username = session.get("username")
        if not username:
            return {"error": "Unauthorized"}, 401

        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return {"error": "Access denied"}, 403

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT date, group_name, reason, created_by, created_at
            FROM excluded_dates
            ORDER BY date DESC
            """
        )
        excluded_dates = [
            {
                "date": row[0],
                "group_name": row[1],
                "reason": row[2],
                "created_by": row[3],
                "created_at": row[4],
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return {"excluded_dates": excluded_dates}

    @app.route("/api/attendance_data", methods=["GET"])
    def api_attendance_data():
        username = session.get("username")
        if not username:
            return {"error": "Unauthorized"}, 401

        role = get_user_role(username)
        if role not in ["teacher", "admin"]:
            return {"error": "Access denied"}, 403

        group = request.args.get("group")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        if not group:
            return {"error": "Group required"}, 400

        days, data = get_attendance_data(group, start_date, end_date, teacher_username=username if role == "teacher" else None)
        return {"days": days, "data": data, "success": True}
