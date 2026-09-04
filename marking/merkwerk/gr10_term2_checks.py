from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


NS = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
}

MANUAL_TYPES = {
    "docx_cover_page",
    "docx_cover_page_image",
    "docx_content_controls",
    "docx_textbox",
    "docx_textbox_style",
    "docx_textbox_effect",
    "external_autocorrect",
}


@dataclass
class CheckResult:
    learner: str
    class_group: str
    check_id: str
    file: str
    status: str
    mark_awarded: int
    mark_possible: int
    message: str


def norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def formula_without_absolute_markers(value: Any) -> str:
    return norm_formula(value).replace("$", "")


def canonical_formula(value: Any) -> str:
    formula = norm_formula(value)
    if formula.startswith("="):
        formula = formula[1:]
    while formula.startswith("(") and formula.endswith(")"):
        depth = 0
        balanced = True
        for index, char in enumerate(formula):
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0 and index != len(formula) - 1:
                    balanced = False
                    break
        if balanced:
            formula = formula[1:-1]
        else:
            break
    return formula


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def qn(name: str) -> str:
    return f"{{{NS['w']}}}{name}"


def looks_red(value: str) -> bool:
    text = (value or "").strip().upper()
    if text in {"RED", "FF0000", "C00000", "C0504D", "9C0006"}:
        return True
    if len(text) == 6:
        try:
            red = int(text[0:2], 16)
            green = int(text[2:4], 16)
            blue = int(text[4:6], 16)
        except ValueError:
            return False
        return red >= 120 and red > green and red > blue
    return False


def resolve_sheet_name(workbook: Any, requested: str | None) -> str | None:
    if not requested:
        return None
    for title in workbook.sheetnames:
        if title == requested:
            return title
    for title in workbook.sheetnames:
        if title.lower() == requested.lower():
            return title
    return None


class DocxReader:
    def __init__(self, path: Path):
        self.path = path
        self.ok = path.exists()
        self.error = ""
        self.document_root = None
        self.numbering_root = None
        self.core_root = None
        self.comments = []
        self.media_hashes = set()
        self.paragraphs = []
        self.tables = []
        if self.ok:
            self._load()

    def _load(self) -> None:
        try:
            with zipfile.ZipFile(self.path) as zf:
                self.document_root = ET.fromstring(zf.read("word/document.xml"))
                try:
                    self.numbering_root = ET.fromstring(zf.read("word/numbering.xml"))
                except KeyError:
                    self.numbering_root = None
                try:
                    self.core_root = ET.fromstring(zf.read("docProps/core.xml"))
                except KeyError:
                    self.core_root = None
                try:
                    comments_root = ET.fromstring(zf.read("word/comments.xml"))
                    self.comments = [
                        "".join(t.text or "" for t in c.findall(".//w:t", NS))
                        for c in comments_root.findall(".//w:comment", NS)
                    ]
                except KeyError:
                    self.comments = []
                for name in zf.namelist():
                    if name.startswith("word/media/"):
                        self.media_hashes.add(hashlib.sha256(zf.read(name)).hexdigest())
            self.paragraphs = self.document_root.findall(".//w:p", NS)
            self.tables = self.document_root.findall(".//w:tbl", NS)
        except Exception as exc:
            self.ok = False
            self.error = str(exc)

    def full_text(self) -> str:
        return "\n".join(self.paragraph_text(p) for p in self.paragraphs)

    def contains_image_matching_file(self, image_name: str) -> bool:
        image_path = self.path.parent / image_name
        if not image_path.exists():
            return False
        return hashlib.sha256(image_path.read_bytes()).hexdigest() in self.media_hashes

    def content_control_texts(self) -> list[str]:
        texts = []
        for control in self.document_root.findall(".//w:sdt", NS):
            texts.append(norm_text(" ".join(self.paragraph_text(p) for p in control.findall(".//w:p", NS))))
        return [text for text in texts if text]

    def has_cover_page_control(self) -> bool:
        for control in self.document_root.findall(".//w:sdt", NS):
            props = control.find("./w:sdtPr", NS)
            if props is None:
                continue
            gallery = props.find(".//w:docPartGallery", NS)
            if gallery is not None and gallery.attrib.get(qn("val"), "") == "Cover Pages":
                return True
        return False

    def content_control_aliases(self) -> list[str]:
        aliases = []
        for control in self.document_root.findall(".//w:sdt", NS):
            props = control.find("./w:sdtPr", NS)
            if props is None:
                continue
            alias = props.find("./w:alias", NS)
            if alias is not None:
                aliases.append(norm_text(alias.attrib.get(qn("val"), "")))
        return [alias for alias in aliases if alias]

    def textbox_nodes(self) -> list[ET.Element]:
        return self.document_root.findall(".//w:txbxContent", NS)

    def textbox_texts(self) -> list[str]:
        return [norm_text(" ".join(self.paragraph_text(p) for p in box.findall(".//w:p", NS))) for box in self.textbox_nodes()]

    def textbox_paragraph_styles(self) -> list[str]:
        styles = []
        for box in self.textbox_nodes():
            for paragraph in box.findall(".//w:p", NS):
                styles.append(self.paragraph_style(paragraph))
        return styles

    def has_textbox_shadow(self) -> bool:
        for tag in ("shadow", "outerShdw"):
            if any(local_name(node.tag) == tag for node in self.document_root.iter()):
                return True
        return False

    def textbox_anchors(self) -> list[ET.Element]:
        return [anchor for anchor in self.document_root.findall(".//wp:anchor", NS) if anchor.findall(".//w:txbxContent", NS)]

    def has_right_side_textbox(self) -> bool:
        for anchor in self.textbox_anchors():
            pos_h = anchor.find("./wp:positionH", NS)
            if pos_h is None:
                continue
            align = pos_h.find("./wp:align", NS)
            if align is not None and (align.text or "").lower() == "right":
                return True
            offset = pos_h.find("./wp:posOffset", NS)
            if offset is not None:
                try:
                    if int(offset.text or "0") > 0:
                        return True
                except ValueError:
                    continue
        return False

    def paragraph_text(self, paragraph: ET.Element) -> str:
        parts = []
        for node in paragraph.iter():
            if node.tag == qn("t"):
                parts.append(node.text or "")
            elif node.tag == qn("tab"):
                parts.append("\t")
            elif node.tag == qn("br"):
                parts.append("\n")
        return "".join(parts)

    def paragraph_style(self, paragraph: ET.Element) -> str:
        node = paragraph.find("./w:pPr/w:pStyle", NS)
        return node.attrib.get(qn("val"), "") if node is not None else ""

    def paragraph_runs(self, paragraph: ET.Element) -> list[dict[str, Any]]:
        runs = []
        for run in paragraph.findall("./w:r", NS):
            text = "".join(t.text or "" for t in run.findall(".//w:t", NS))
            props = run.find("./w:rPr", NS)
            runs.append({"text": text, "props": props})
        return runs

    def find_paragraph(self, anchor: dict[str, Any] | None) -> ET.Element | None:
        if not anchor:
            return None
        heading = anchor.get("heading")
        if heading:
            heading_index = None
            for index, paragraph in enumerate(self.paragraphs):
                if heading.lower() in self.paragraph_text(paragraph).lower():
                    heading_index = index
                    break
            if heading_index is None:
                return None
            starts = str(anchor.get("paragraph_starts_with", "")).lower()
            offset = anchor.get("paragraph_index_after_heading")
            seen = 0
            for paragraph in self.paragraphs[heading_index + 1 :]:
                text = norm_text(self.paragraph_text(paragraph))
                if not text:
                    continue
                seen += 1
                if offset and seen == offset:
                    return paragraph
                if starts and text.lower().startswith(starts):
                    return paragraph
            return None
        starts = str(anchor.get("paragraph_starts_with", "")).lower()
        ends = str(anchor.get("paragraph_ends_with", "")).lower()
        for paragraph in self.paragraphs:
            text = norm_text(self.paragraph_text(paragraph)).lower()
            if starts and not text.startswith(starts):
                continue
            if ends and ends not in text:
                continue
            return paragraph
        return None

    def author(self) -> str:
        if self.core_root is None:
            return ""
        creator = self.core_root.find(".//dc:creator", NS)
        return norm_text(creator.text if creator is not None else "")

    def table_after_heading(self, heading: str | None) -> ET.Element | None:
        if not heading:
            return self.tables[0] if self.tables else None
        body = self.document_root.find(".//w:body", NS)
        if body is None:
            return None
        seen_heading = False
        for child in list(body):
            if child.tag == qn("p") and heading.lower() in self.paragraph_text(child).lower():
                seen_heading = True
            elif seen_heading and child.tag == qn("tbl"):
                return child
        return None

    def table_after_paragraph_text(self, paragraph_text: str | None) -> ET.Element | None:
        if not paragraph_text:
            return None
        target = norm_text(paragraph_text).lower()
        body = self.document_root.find(".//w:body", NS)
        if body is None:
            return None
        seen_paragraph = False
        for child in list(body):
            if child.tag == qn("p"):
                text = norm_text(self.paragraph_text(child)).lower()
                if target in text:
                    seen_paragraph = True
            elif seen_paragraph and child.tag == qn("tbl"):
                return child
        return None

    def table_for_anchor(self, anchor: dict[str, Any]) -> ET.Element | None:
        table = self.table_after_paragraph_text(anchor.get("after_paragraph"))
        if table is not None:
            return table
        return self.table_after_heading(anchor.get("heading"))

    def paragraph_num_id(self, paragraph: ET.Element) -> str:
        node = paragraph.find("./w:pPr/w:numPr/w:numId", NS)
        return node.attrib.get(qn("val"), "") if node is not None else ""

    def paragraph_ilvl(self, paragraph: ET.Element) -> str:
        node = paragraph.find("./w:pPr/w:numPr/w:ilvl", NS)
        return node.attrib.get(qn("val"), "0") if node is not None else "0"

    def abstract_num_id(self, num_id: str) -> str:
        if self.numbering_root is None or not num_id:
            return ""
        num = self.numbering_root.find(f".//w:num[@w:numId='{num_id}']", NS)
        if num is None:
            return ""
        abstract = num.find("./w:abstractNumId", NS)
        return abstract.attrib.get(qn("val"), "") if abstract is not None else ""

    def uses_picture_bullet(self, paragraph: ET.Element) -> bool:
        if self.numbering_root is None:
            return False
        num_id = self.paragraph_num_id(paragraph)
        abstract_id = self.abstract_num_id(num_id)
        if not abstract_id:
            return False
        abstract = self.numbering_root.find(f".//w:abstractNum[@w:abstractNumId='{abstract_id}']", NS)
        if abstract is None:
            return False
        ilvl = self.paragraph_ilvl(paragraph)
        level = abstract.find(f"./w:lvl[@w:ilvl='{ilvl}']", NS)
        if level is None:
            return False
        return level.find(".//w:lvlPicBulletId", NS) is not None

    def heading2_numbering_uses_upper_letters(self) -> bool:
        found_heading2 = False
        if self.numbering_root is None:
            return False
        for paragraph in self.paragraphs:
            if "heading2" not in self.paragraph_style(paragraph).replace(" ", "").lower():
                continue
            if not norm_text(self.paragraph_text(paragraph)):
                continue
            found_heading2 = True
            num_id = self.paragraph_num_id(paragraph)
            abstract_id = self.abstract_num_id(num_id)
            if abstract_id:
                abstract = self.numbering_root.find(f".//w:abstractNum[@w:abstractNumId='{abstract_id}']", NS)
                if abstract is not None:
                    ilvl = self.paragraph_ilvl(paragraph)
                    level = abstract.find(f"./w:lvl[@w:ilvl='{ilvl}']", NS)
                    num_fmt = level.find("./w:numFmt", NS) if level is not None else None
                    if num_fmt is not None and num_fmt.attrib.get(qn("val")) == "upperLetter":
                        continue
            text = norm_text(self.paragraph_text(paragraph))
            if not re.match(r"^[A-Z][\.\)]\s+", text):
                return False
        return found_heading2

    def bullet_paragraphs_after_heading(self, heading: str) -> list[ET.Element]:
        heading_index = None
        for index, paragraph in enumerate(self.paragraphs):
            if heading.lower() in self.paragraph_text(paragraph).lower():
                heading_index = index
                break
        if heading_index is None:
            return []
        bullets = []
        started = False
        for paragraph in self.paragraphs[heading_index + 1 :]:
            text = norm_text(self.paragraph_text(paragraph))
            if not text:
                continue
            num_id = self.paragraph_num_id(paragraph)
            if num_id:
                bullets.append(paragraph)
                started = True
                continue
            if started:
                break
        return bullets

    def paragraphs_in_section(self, heading: str) -> list[ET.Element]:
        heading_index = None
        for index, paragraph in enumerate(self.paragraphs):
            if heading.lower() in self.paragraph_text(paragraph).lower():
                heading_index = index
                break
        if heading_index is None:
            return []

        section = []
        for paragraph in self.paragraphs[heading_index + 1 :]:
            style = self.paragraph_style(paragraph).replace(" ", "").lower()
            if style.startswith("heading") and norm_text(self.paragraph_text(paragraph)):
                break
            section.append(paragraph)
        return section

    def section_properties(self) -> ET.Element | None:
        props = self.document_root.findall(".//w:sectPr", NS)
        return props[-1] if props else None

    def table_outer_borders(self, table: ET.Element) -> dict[str, list[ET.Element]]:
        rows = table.findall("./w:tr", NS)
        result: dict[str, list[ET.Element]] = {"top": [], "left": [], "bottom": [], "right": []}
        if not rows:
            return result

        top_row = rows[0].findall("./w:tc", NS)
        bottom_row = rows[-1].findall("./w:tc", NS)
        for cell in top_row:
            tc_pr = cell.find("./w:tcPr", NS)
            borders = tc_pr.find("./w:tcBorders", NS) if tc_pr is not None else None
            if borders is not None:
                edge = borders.find("./w:top", NS)
                if edge is not None:
                    result["top"].append(edge)
        for cell in bottom_row:
            tc_pr = cell.find("./w:tcPr", NS)
            borders = tc_pr.find("./w:tcBorders", NS) if tc_pr is not None else None
            if borders is not None:
                edge = borders.find("./w:bottom", NS)
                if edge is not None:
                    result["bottom"].append(edge)

        for row in rows:
            cells = row.findall("./w:tc", NS)
            if not cells:
                continue
            for side, cell in (("left", cells[0]), ("right", cells[-1])):
                tc_pr = cell.find("./w:tcPr", NS)
                borders = tc_pr.find("./w:tcBorders", NS) if tc_pr is not None else None
                if borders is not None:
                    edge = borders.find(f"./w:{side}", NS)
                    if edge is not None:
                        result[side].append(edge)
        return result


def docx_check(reader: DocxReader, check: dict[str, Any]) -> tuple[str, str]:
    if not reader.ok:
        return "error", reader.error or "DOCX file missing or unreadable"

    if check.get("id") == "Q2.3":
        return "pass", "AutoCorrect/dictionary mark awarded to all learners."

    check_type = check["type"]
    prop = check["property"]
    expected = check.get("expected")
    anchor = check.get("anchor") or {}

    if check_type in MANUAL_TYPES and check_type not in {"docx_cover_page", "docx_content_controls", "docx_cover_page_image", "docx_textbox", "docx_textbox_style", "docx_textbox_effect"}:
        return "manual", "Manual review required for this DOCX feature."

    if check_type == "docx_cover_page":
        text = reader.full_text()
        controls = reader.content_control_texts()
        if prop == "template":
            return ("pass", "Cover page content control detected.") if reader.has_cover_page_control() else ("fail", "No cover page content control detected.")
        if prop == "title_present":
            return ("pass", "Document title text found.") if "Analysis of the Types of People Who Use Uber" in text else ("fail", "Document title text not found.")
        if prop == "title_text":
            return ("pass", "Expected title text found.") if "Analysis of the Types of People Who Use Uber" in text else ("fail", "Expected title text not found.")
        if prop == "author_name_present":
            placeholders = {"author", "type the author name", "name and surname"}
            candidate_controls = [c for c in controls if c.lower() not in placeholders and "analysis of the types" not in c.lower()]
            return ("pass", "Non-placeholder author text found in content controls.") if candidate_controls else ("fail", "No non-placeholder author text found.")
        if prop == "author_control_box_filled":
            placeholders = {"author", "type the author name", "name and surname"}
            candidate_controls = [c for c in controls if c.lower() not in placeholders and "analysis of the types" not in c.lower()]
            return ("pass", "Author control appears filled.") if candidate_controls else ("fail", "Author control does not appear filled.")

    if check_type == "docx_content_controls":
        aliases = [alias.lower() for alias in reader.content_control_aliases()]
        allowed_aliases = {"title", "author"}
        leftovers = [alias for alias in aliases if alias and alias not in allowed_aliases]
        if leftovers:
            return "fail", f"Extra content controls still present: {leftovers[:4]}."
        title_count = aliases.count("title")
        author_count = aliases.count("author")
        if title_count < 1 or author_count < 1:
            return "fail", f"Expected title and author controls to remain, found title={title_count}, author={author_count}."
        return "pass", "Only the required title and author content controls remain."

    if check_type == "docx_cover_page_image":
        if prop == "image_replaced":
            return ("pass", "Document contains images.") if reader.media_hashes else ("fail", "No embedded images found.")
        if prop == "image_source_or_match":
            return ("pass", "Embedded image matches 2Hailing.png.") if reader.contains_image_matching_file("2Hailing.png") else ("fail", "No embedded image matches 2Hailing.png exactly.")

    if check_type == "docx_textbox":
        texts = reader.textbox_texts()
        if prop == "textbox_inserted_under_heading":
            return ("pass", "Text box content detected.") if texts else ("fail", "No text box content detected.")
        if prop == "textbox_position":
            return ("pass", "Right-side text box position detected.") if reader.has_right_side_textbox() else ("fail", "Right-side text box position not detected.")
        if prop == "second_paragraph_moved_into_textbox":
            return ("pass", "Non-empty text box content found.") if any(text for text in texts) else ("fail", "No moved paragraph text detected in a text box.")

    if check_type == "docx_textbox_style":
        styles = [style.replace(" ", "").lower() for style in reader.textbox_paragraph_styles()]
        return ("pass", "Quote style found in text box.") if any("quote" in style for style in styles) else ("fail", f"Text box styles: {styles[:5]}.")

    if check_type == "docx_textbox_effect":
        return ("pass", "Shadow/effect XML found.") if reader.has_textbox_shadow() else ("fail", "No shadow/effect XML found.")

    if check_type == "docx_page_setup":
        sect = reader.section_properties()
        if sect is None:
            return "fail", "No section properties found."
        if prop == "paper_size":
            pg = sect.find("./w:pgSz", NS)
            if pg is None:
                return "fail", "No page size found."
            width = int(pg.attrib.get(qn("w"), "0"))
            height = int(pg.attrib.get(qn("h"), "0"))
            return ("pass", "A4 page size found.") if {width, height} == {11906, 16838} else ("fail", f"Page size is {width} x {height} twips.")
        margins = sect.find("./w:pgMar", NS)
        if margins is None:
            return "fail", "No page margins found."
        top = int(margins.attrib.get(qn("top"), "0"))
        bottom = int(margins.attrib.get(qn("bottom"), "0"))
        left = int(margins.attrib.get(qn("left"), "0"))
        right = int(margins.attrib.get(qn("right"), "0"))
        if prop == "margin_top_bottom_cm":
            return ("pass", "Top and bottom margins are 2 cm.") if abs(top - 1134) <= 30 and abs(bottom - 1134) <= 30 else ("fail", f"Top/bottom margins are {top}/{bottom} twips.")
        if prop == "margin_left_right_cm":
            return ("pass", "Left and right margins are 3 cm.") if abs(left - 1701) <= 30 and abs(right - 1701) <= 30 else ("fail", f"Left/right margins are {left}/{right} twips.")

    if check_type == "docx_page_border":
        sect = reader.section_properties()
        borders = sect.find("./w:pgBorders", NS) if sect is not None else None
        return ("pass", "Page border present.") if borders is not None and list(borders) else ("fail", "No page border found.")

    if check_type == "docx_text_spacing":
        paragraph = reader.find_paragraph(anchor)
        if paragraph is None:
            return "fail", "Target paragraph not found."
        text = reader.paragraph_text(paragraph)
        return ("pass", "No repeated spaces found.") if not re.search(r" {2,}", text) else ("fail", "Repeated spaces still present.")

    if check_type == "docx_text_content":
        text = reader.full_text()
        return ("pass", "Registered symbol found near Uber.") if re.search(r"Uber\s*[®]", text) else ("fail", "Registered symbol near Uber not found.")

    if check_type == "docx_comment":
        if prop == "comment_present":
            return ("pass", "Comment exists.") if reader.comments else ("fail", "No comments found.")
        return ("pass", "Comment text found.") if any(norm_text(c) for c in reader.comments) else ("fail", "No comment text found.")

    if check_type == "docx_find_replace":
        occurrences = list(re.finditer(r"\bdrivers\b", reader.full_text(), re.I))
        return ("pass", f"Found {len(occurrences)} drivers occurrences for formatting checks.") if occurrences else ("fail", "No drivers occurrences found.")

    if check_type in {"docx_paragraph_format", "docx_paragraph_structure"}:
        paragraph = reader.find_paragraph(anchor)
        if anchor.get("heading") and prop in {"bulleted_list", "picture_bullet"}:
            bullets = reader.bullet_paragraphs_after_heading(anchor.get("heading", ""))
            paragraph = bullets[0] if bullets else None
        if prop == "bullet_point_count":
            bullets = reader.bullet_paragraphs_after_heading(anchor.get("heading", ""))
            has_inline_breaks = any(paragraph.find(".//w:br", NS) is not None for paragraph in bullets)
            if len(bullets) == int(expected) and not has_inline_breaks:
                return "pass", "Found 3 separate bullet paragraphs."
            reason = f"Found {len(bullets)} bullet points."
            if has_inline_breaks:
                reason += " Inline line breaks detected inside bullet paragraphs."
            return "fail", reason
        if paragraph is None:
            return "fail", "Target paragraph not found."
        ppr = paragraph.find("./w:pPr", NS)
        if prop == "bulleted_list":
            return ("pass", "Paragraph has numbering/bullet properties.") if ppr is not None and ppr.find("./w:numPr", NS) is not None else ("fail", "Paragraph is not a detected list item.")
        if prop == "picture_bullet":
            return ("pass", "Picture bullet detected in numbering XML.") if reader.uses_picture_bullet(paragraph) else ("fail", "Picture bullet not detected.")

    if check_type == "docx_run_format":
        text_anchor = str(anchor.get("text", "")).lower()
        if anchor.get("heading"):
            paragraphs = reader.paragraphs_in_section(anchor["heading"])
            if not paragraphs:
                found = reader.find_paragraph(anchor)
                paragraphs = [found] if found is not None else []
        else:
            paragraphs = reader.paragraphs
        for paragraph in [p for p in paragraphs if p is not None]:
            for run in reader.paragraph_runs(paragraph):
                if text_anchor and text_anchor not in run["text"].lower():
                    continue
                props = run["props"]
                if props is None:
                    continue
                if prop == "superscript":
                    va = props.find("./w:vertAlign", NS)
                    if va is not None and va.attrib.get(qn("val")) == "superscript":
                        return "pass", "Superscript found."
                if prop == "italic" and props.find("./w:i", NS) is not None:
                    return "pass", "Italic found."
                if prop == "bold_correctly_applied" and props.find("./w:b", NS) is not None:
                    return "pass", "Bold formatting found in target paragraph."
                if prop == "underline_correctly_applied" and props.find("./w:u", NS) is not None:
                    return "pass", "Underline formatting found in target paragraph."
                if prop == "double_underline":
                    underline = props.find("./w:u", NS)
                    if underline is not None and underline.attrib.get(qn("val")) == "double":
                        return "pass", "Double underline found."
                if prop == "font_color":
                    color = props.find("./w:color", NS)
                    underline = props.find("./w:u", NS)
                    values = {
                        color.attrib.get(qn("val"), "").upper() if color is not None else "",
                        underline.attrib.get(qn("color"), "").upper() if underline is not None else "",
                    }
                    if values & {"FF0000", "RED", "C0504D", "C00000"}:
                        return "pass", "Red font found."
        return "fail", f"{prop} not found."

    if check_type == "docx_table_row" and prop == "row_deleted":
        table = reader.table_for_anchor(anchor)
        text = "".join(reader.paragraph_text(p) for p in table.findall(".//w:p", NS)) if table is not None else ""
        return ("pass", "Driver Rating row not found.") if "Driver Rating".lower() not in text.lower() else ("fail", "Driver Rating row still present.")

    if check_type == "docx_table_cell":
        table = reader.table_for_anchor(anchor)
        if table is None:
            return "fail", "Target table not found."
        rows = table.findall("./w:tr", NS)
        row_index = int(anchor.get("row", 1)) - 1
        col_index = int(anchor.get("column", 1)) - 1
        if row_index >= len(rows):
            return "fail", "Target row not found."
        cells = rows[row_index].findall("./w:tc", NS)
        if col_index >= len(cells):
            return "fail", "Target column not found."
        cell = cells[col_index]
        cell_text = norm_text(" ".join(reader.paragraph_text(p) for p in cell.findall(".//w:p", NS)))
        if anchor.get("cell_text") and anchor["cell_text"].lower() not in cell_text.lower():
            return "fail", f"Target cell text is {cell_text or '<blank>'}."
        tcpr = cell.find("./w:tcPr", NS)
        if tcpr is None:
            return "fail", "Cell properties not found."
        if prop == "text_direction":
            direction = tcpr.find("./w:textDirection", NS)
            value = direction.attrib.get(qn("val"), "") if direction is not None else ""
            return ("pass", f"Text direction is {value}.") if value in {"btLr", "tbRl", "tbRlV"} else ("fail", f"Text direction is {value or '<none>'}.")
        if prop == "vertical_alignment":
            valign = tcpr.find("./w:vAlign", NS)
            value = valign.attrib.get(qn("val"), "") if valign is not None else ""
            return ("pass", "Cell vertical alignment is center/middle.") if value == "center" else ("fail", f"Vertical alignment is {value or '<none>'}.")

    if check_type == "docx_table_row":
        table = reader.table_for_anchor(anchor)
        if table is None:
            return "fail", "Target table not found."
        rows = table.findall("./w:tr", NS)[2:]
        if not rows:
            return "fail", "Rows from row 3 onwards not found."
        heights = []
        for row in rows:
            height = row.find("./w:trPr/w:trHeight", NS)
            heights.append(int(height.attrib.get(qn("val"), "0")) if height is not None else 0)
        if prop == "row_height_applied":
            return ("pass", "Row heights found from row 3 onwards.") if all(h > 0 for h in heights) else ("fail", f"Row heights: {heights}.")
        if prop == "row_height_cm":
            expected_twips = 284
            return ("pass", "Row heights are 0.5 cm.") if all(abs(h - expected_twips) <= 10 for h in heights) else ("fail", f"Row heights: {heights}.")

    if check_type == "docx_table_border":
        table = reader.table_for_anchor(anchor)
        if table is None:
            return "fail", "Target table not found."
        borders = table.find("./w:tblPr/w:tblBorders", NS)
        outside = [borders.find(f"./w:{side}", NS) if borders is not None else None for side in ("top", "left", "bottom", "right")]
        cell_outside = reader.table_outer_borders(table)

        def side_edges(side: str) -> list[ET.Element]:
            idx = {"top": 0, "left": 1, "bottom": 2, "right": 3}[side]
            tbl_edge = outside[idx]
            edges = []
            if tbl_edge is not None:
                edges.append(tbl_edge)
            edges.extend(cell_outside.get(side, []))
            return edges

        if prop == "outside_border_applied":
            ok = all(side_edges(side) for side in ("top", "left", "bottom", "right"))
            return ("pass", "Outside borders found.") if ok else ("fail", "Not all outside borders found.")
        if prop == "outside_border_color":
            values = []
            ok = True
            for side in ("top", "left", "bottom", "right"):
                edges = side_edges(side)
                side_values = [edge.attrib.get(qn("color"), "") for edge in edges]
                values.append(side_values)
                if not edges or not any(looks_red(value) for value in side_values):
                    ok = False
            return ("pass", "Outside border colour is red.") if ok else ("fail", f"Outside border colours: {values}.")
        if prop == "outside_border_style":
            values = []
            valid_styles = {"double", "thinThickSmallGap"}
            ok = True
            for side in ("top", "left", "bottom", "right"):
                edges = side_edges(side)
                side_values = [edge.attrib.get(qn("val"), "") for edge in edges]
                values.append(side_values)
                if not edges or not any(value in valid_styles for value in side_values):
                    ok = False
            return ("pass", "Outside border style is double.") if ok else ("fail", f"Outside border styles: {values}.")
        if prop == "outside_border_width_pt":
            values = []
            valid_style_widths = {
                "double": {12},
                "thinThickSmallGap": {12},
            }
            ok = True
            for side in ("top", "left", "bottom", "right"):
                edges = side_edges(side)
                side_values = [(edge.attrib.get(qn("val"), ""), int(edge.attrib.get(qn("sz"), "0"))) for edge in edges]
                values.append(side_values)
                if not edges or not any(
                    style in valid_style_widths and width in valid_style_widths[style]
                    for style, width in side_values
                ):
                    ok = False
            return ("pass", "Outside border width is 1.5 pt.") if ok else ("fail", f"Outside border widths: {values}.")

    if check_type == "docx_text_count":
        count = len(re.findall(r"\btaxi\b", reader.full_text(), re.I))
        return ("pass", f"Found taxi {count} times.") if count > 0 else ("fail", "Word taxi not found.")

    if check_type == "docx_document_property":
        author = reader.author()
        expected_text = str(expected)
        if prop == "author_contains_value":
            accepted = {"5", "6", "7"} if expected_text == "6" else {expected_text}
            found = [value for value in accepted if re.search(rf"(?<!\d){re.escape(value)}(?!\d)", author)]
            if found:
                return "pass", f"Author contains accepted value {found[0]}."
            full_text = reader.full_text()
            found_in_text = [value for value in accepted if re.search(rf"(?<!\d){re.escape(value)}(?!\d)", full_text)]
            return ("pass", f"Accepted value {found_in_text[0]} found in document text.") if found_in_text else ("fail", f"Author is {author or '<blank>'}.")
        accepted = {"5", "6", "7"} if expected_text == "6" else {expected_text}
        return ("pass", f"Author is accepted value {author}.") if author in accepted else ("fail", f"Author is {author or '<blank>'}.")

    if check_type == "docx_paragraph_style":
        paragraph = reader.find_paragraph(anchor)
        if paragraph is None:
            return "fail", "Target paragraph not found."
        style = reader.paragraph_style(paragraph).replace(" ", "").lower()
        if prop == "style_applied":
            return ("pass", f"Style is {style}.") if "intensequote" in style else ("fail", f"Style is {style or '<none>'}.")
        return "pass", "Target paragraph found."

    if check_type in {"docx_heading_numbering", "docx_heading_style"}:
        if check_type == "docx_heading_numbering":
            return ("pass", "Heading 2 numbering uses upper letters.") if reader.heading2_numbering_uses_upper_letters() else ("fail", "Heading 2 upper-letter numbering not detected.")
        heading2 = [p for p in reader.paragraphs if "heading2" in reader.paragraph_style(p).replace(" ", "").lower()]
        return ("pass", f"Found {len(heading2)} Heading 2 paragraphs.") if heading2 else ("fail", "No Heading 2 paragraphs found.")

    return "manual", f"No automated DOCX handler for {check_type}:{prop}."


def color_is_yellow(color: Any) -> bool:
    rgb = getattr(color, "rgb", None)
    indexed = getattr(color, "indexed", None)
    return str(rgb).upper().endswith("FFFF00") or indexed == 5


def populated_row_cells(ws: Any, row_number: int) -> list[Any]:
    return [cell for cell in ws[row_number] if cell.value is not None]


def find_header_column(ws: Any, expected: str) -> int | None:
    expected_norm = norm_text(expected).lower()
    for cell in ws[1]:
        value = norm_text(cell.value).lower()
        if value == expected_norm or expected_norm in value or value in expected_norm:
            return cell.column
    return None


def category_sort_state(ws: Any) -> dict[str, Any]:
    primary_col = find_header_column(ws, "PURPOSE")
    secondary_col = find_header_column(ws, "Distance (KMS)")
    if primary_col is None or secondary_col is None:
        return {"headers_found": False, "message": "PURPOSE or Distance (KMS) header not found."}

    rows = []
    for row in range(2, ws.max_row + 1):
        primary = norm_text(ws.cell(row, primary_col).value)
        secondary = ws.cell(row, secondary_col).value
        if not primary:
            continue
        try:
            distance = float(secondary)
        except (TypeError, ValueError):
            distance = None
        rows.append((primary.lower(), distance))

    if not rows:
        return {"headers_found": True, "message": "No sortable category rows found."}

    primary_values = [primary for primary, _distance in rows]
    primary_sorted = all(left <= right for left, right in zip(primary_values, primary_values[1:]))

    seen: set[str] = set()
    current = ""
    grouped = True
    for primary in primary_values:
        if primary != current:
            if primary in seen:
                grouped = False
                break
            seen.add(primary)
            current = primary

    contiguous_groups: list[list[float | None]] = []
    current_group: list[float | None] = []
    current_primary = None
    for primary, distance in rows:
        if primary != current_primary:
            if current_group:
                contiguous_groups.append(current_group)
            current_group = [distance]
            current_primary = primary
        else:
            current_group.append(distance)
    if current_group:
        contiguous_groups.append(current_group)

    groups_with_multiple_rows = [group for group in contiguous_groups if len(group) > 1]
    secondary_desc = True
    for group in groups_with_multiple_rows:
        for left, right in zip(group, group[1:]):
            if left is not None and right is not None and left < right:
                secondary_desc = False
                break
        if not secondary_desc:
            break

    second_level_only = bool(groups_with_multiple_rows) and secondary_desc
    all_levels = grouped and primary_sorted and second_level_only

    return {
        "headers_found": True,
        "primary_grouped": grouped,
        "primary_sorted": primary_sorted,
        "second_level_only": second_level_only,
        "secondary_desc": secondary_desc,
        "all_levels": all_levels,
    }


def defined_names_for_cell(workbook: Any, sheet_name: str, cell_ref: str) -> set[str]:
    names = set()
    target = cell_ref.replace("$", "").upper()
    for defined_name in workbook.defined_names.values():
        try:
            destinations = list(defined_name.destinations)
        except Exception:
            continue
        for title, coordinate in destinations:
            if title == sheet_name and coordinate.replace("$", "").upper() == target:
                names.add(defined_name.name.upper())
    return names


def xlsx_check(workbook_cache: dict[tuple[Path, bool], Any], path: Path, check: dict[str, Any]) -> tuple[str, str]:
    if not path.exists():
        return "error", "XLSX file missing."
    try:
        cache_key = (path, False)
        if cache_key not in workbook_cache:
            workbook_cache[cache_key] = load_workbook(path, data_only=False)
        wb = workbook_cache[cache_key]
    except Exception as exc:
        return "error", str(exc)

    anchor = check.get("anchor") or {}
    prop = check["property"]
    expected = check.get("expected")
    requested_sheet_name = anchor.get("sheet")
    sheet_name = resolve_sheet_name(wb, requested_sheet_name)
    if requested_sheet_name and sheet_name is None:
        return "fail", f"Sheet {requested_sheet_name} not found."
    ws = wb[sheet_name] if sheet_name else wb.active

    if check["type"] == "xlsx_sheet":
        if prop == "sheet_name":
            return ("pass", "Categories sheet found.") if resolve_sheet_name(wb, "Categories") else ("fail", f"Sheets: {', '.join(wb.sheetnames)}")
        if prop == "tab_color":
            color = ws.sheet_properties.tabColor
            rgb = getattr(color, "rgb", "") if color else ""
            return ("pass", "Yellow tab colour detected.") if str(rgb).upper().endswith("FFFF00") else ("fail", f"Tab color is {rgb or '<none>'}.")

    if check["type"] == "xlsx_sort":
        state = category_sort_state(ws)
        if not state.get("headers_found"):
            return "fail", str(state["message"])
        if prop == "first_level_field":
            return ("pass", "PURPOSE appears to be the primary sort field.") if state.get("primary_grouped") else ("fail", "PURPOSE values are not grouped.")
        if prop == "first_level_order":
            return ("pass", "PURPOSE is sorted ascending.") if state.get("primary_sorted") else ("fail", "PURPOSE is not sorted ascending.")
        if prop == "second_level_field":
            return ("pass", "Distance order detected within PURPOSE groups.") if state.get("second_level_only") else ("fail", "Distance is not sorted within PURPOSE groups.")
        if prop == "second_level_order":
            return ("pass", "All sort levels are correct.") if state.get("all_levels") else ("fail", "The full two-level sort is not correct.")

    if check["type"] == "xlsx_cell_format" and prop == "number_format":
        if expected == "euro" and anchor.get("cell"):
            fmt = ws[anchor["cell"]].number_format
            euro = "\u20ac" in fmt or "EUR" in fmt.upper()
            return ("pass", "Euro number format found.") if euro else ("fail", f"Number format is {fmt}.")
        formats = {ws[f"E{row}"].number_format for row in range(2, min(ws.max_row, 50) + 1)}
        currency = any("R" in fmt or "$" in fmt or "€" in fmt or "[$" in fmt for fmt in formats)
        return ("pass", "Currency-like number format found in column E.") if currency else ("fail", f"Formats found: {sorted(formats)[:5]}")

    if check["type"] == "xlsx_font":
        if prop == "font_size_pt":
            sizes = {ws.cell(1, col).font.sz for col in range(1, ws.max_column + 1) if ws.cell(1, col).value is not None}
            return ("pass", "Row 1 font size is 16 pt.") if 16 in sizes or 16.0 in sizes else ("fail", f"Row 1 sizes: {sorted(str(s) for s in sizes)}")
        if prop == "bold":
            cells = populated_row_cells(ws, 2)
            return ("pass", "Bold found in row 2.") if cells and all(cell.font.bold for cell in cells) else ("fail", "Not all populated row 2 cells are bold.")

    if check["type"] == "xlsx_fill":
        if prop == "fill_color":
            cell = ws[anchor["cell"]]
            return ("pass", "Yellow fill found.") if color_is_yellow(cell.fill.fgColor) else ("fail", f"Fill color is {cell.fill.fgColor.rgb}.")
        if prop == "row_shading":
            cells = populated_row_cells(ws, 2)
            filled = any(cell.fill.fill_type and cell.fill.fill_type != "none" for cell in cells)
            return ("pass", "Row 2 shading found.") if filled else ("fail", "No row 2 shading detected.")

    if check["type"] == "xlsx_border":
        cells = populated_row_cells(ws, 2)
        has_border = any(any(getattr(cell.border, side).style for side in ("left", "right", "top", "bottom")) for cell in cells)
        return ("pass", "Borders found in row 2.") if has_border else ("fail", "No row 2 borders detected.")

    if check["type"] == "xlsx_column_format":
        width = ws.column_dimensions[anchor["column"]].width
        return ("pass", f"Column width is {width}.") if width and width > 10 else ("fail", f"Column width is {width}.")

    if check["type"] == "xlsx_cell_text":
        value = norm_text(ws[anchor["cell"]].value)
        return ("pass", f"Cell text is {value}.") if value and not re.search(r"\b[a-z]*teh[a-z]*\b", value, re.I) else ("fail", f"Cell text is {value or '<blank>'}.")

    if check["type"] == "xlsx_formula":
        formula = norm_formula(ws[anchor["cell"]].value)
        expected_formula = check.get("expected")
        if check.get("id") == "Q3.7a" and "COUNT" in formula:
            return "pass", "COUNT function accepted for Q3.7."
        if check.get("id") == "Q3.7b":
            if "COUNTIF(" in formula:
                return "pass", "COUNTIF accepted for Q3.7."
            if re.search(r"\$?[A-Z]{1,3}\$?2:\$?[A-Z]{1,3}\$?50", formula):
                return "pass", "Rows 2:50 accepted for Q3.7."
        if prop == "formula":
            return ("pass", f"Formula is {formula}.") if canonical_formula(formula) == canonical_formula(expected_formula) else ("fail", f"Formula is {formula or '<blank>'}.")
        if prop == "function":
            return ("pass", f"{expected} found.") if str(expected).upper() + "(" in formula else ("fail", f"Formula is {formula or '<blank>'}.")
        if prop == "operation":
            symbols = {"addition": "+", "subtraction": "-", "division": "/", "multiplication": "*"}
            symbol = symbols.get(str(expected))
            return ("pass", f"Operation {expected} found.") if symbol and symbol in formula else ("fail", f"Formula is {formula or '<blank>'}.")
        if prop in {"reference", "range", "absolute_reference"}:
            expected_formula = norm_formula(expected)
            if prop == "absolute_reference":
                return ("pass", f"{expected} found.") if expected_formula in formula else ("fail", f"Formula is {formula or '<blank>'}.")
            return ("pass", f"{expected} found.") if formula_without_absolute_markers(expected) in formula_without_absolute_markers(formula) else ("fail", f"Formula is {formula or '<blank>'}.")
        if prop == "fixed_rate_reference":
            names = defined_names_for_cell(wb, ws.title, anchor.get("rate_cell", "G1"))
            uses_name = any(re.search(rf"(?<![A-Z0-9_]){re.escape(name)}(?![A-Z0-9_])", formula) for name in names)
            if re.search(r"(?<![A-Z0-9_])\$?G\$?1(?![A-Z0-9_])", formula) or "12.6" in formula or uses_name:
                return "pass", "Fixed rate reference accepted."
            return "fail", f"Formula is {formula or '<blank>'}."
        if prop == "second_range":
            occurrences = formula.count(norm_formula(expected))
            return ("pass", f"{expected} appears twice.") if occurrences >= 2 else ("fail", f"Formula is {formula or '<blank>'}.")
        if prop == "operand":
            return ("pass", f"{expected} found.") if str(expected).upper() in formula else ("fail", f"Formula is {formula or '<blank>'}.")

    if check["type"] == "xlsx_value":
        try:
            cache_key = (path, True)
            if cache_key not in workbook_cache:
                workbook_cache[cache_key] = load_workbook(path, data_only=True)
            value_wb = workbook_cache[cache_key]
            value_ws = value_wb[sheet_name] if sheet_name and sheet_name in value_wb.sheetnames else value_wb.active
            value = value_ws[anchor["cell"]].value
        except Exception as exc:
            return "error", f"Could not read cached value: {exc}"
        if value is None:
            return "manual", "No cached calculated value found; workbook likely needs Excel recalculation."
        try:
            actual = float(value)
            expected_number = float(expected)
        except (TypeError, ValueError):
            return ("pass", f"Cached value is {value}.") if norm_text(value) == norm_text(expected) else ("fail", f"Cached value is {value}.")
        tolerance = 0.01 if abs(expected_number) >= 1 else 0.001
        return ("pass", f"Cached value is {actual}.") if abs(actual - expected_number) <= tolerance else ("fail", f"Cached value is {actual}; expected {expected_number}.")

    return "manual", f"No automated XLSX handler for {check['type']}:{prop}."


def resolve_file(learner_dir: Path, expected_name: str) -> Path:
    exact = learner_dir / expected_name
    if exact.exists():
        return exact
    expected_lower = expected_name.lower()
    for path in learner_dir.iterdir():
        if path.is_file() and path.name.lower() == expected_lower:
            return path
    if expected_name == "3Uber Ride Bookings.xlsx":
        for fallback in ("Categories.xlsx", "Bookings.xlsx"):
            candidate = learner_dir / fallback
            if candidate.exists():
                return candidate
    return exact


def find_learners(root: Path, groups: list[str]) -> list[tuple[str, Path]]:
    learners = []
    for group in groups:
        group_dir = root / group
        if not group_dir.exists():
            continue
        for learner_dir in sorted(p for p in group_dir.iterdir() if p.is_dir()):
            learners.append((group, learner_dir))
    return learners


def run_checks(root: Path, expectations_path: Path, groups: list[str]) -> list[CheckResult]:
    expectations = json.loads(expectations_path.read_text(encoding="utf-8"))
    checks = expectations["checks"]
    results = []
    for class_group, learner_dir in find_learners(root, groups):
        docx_cache: dict[Path, DocxReader] = {}
        workbook_cache: dict[tuple[Path, bool], Any] = {}
        for check in checks:
            expected_file = check["file"]
            path = resolve_file(learner_dir, expected_file)
            if expected_file.endswith(".docx"):
                if path not in docx_cache:
                    docx_cache[path] = DocxReader(path)
                status, message = docx_check(docx_cache[path], check)
            elif expected_file.endswith(".xlsx"):
                status, message = xlsx_check(workbook_cache, path, check)
            else:
                status, message = "manual", "Unsupported file type."
            results.append(
                CheckResult(
                    learner=learner_dir.name,
                    class_group=class_group,
                    check_id=check["id"],
                    file=expected_file,
                    status=status,
                    mark_awarded=1 if status == "pass" else 0,
                    mark_possible=int(check.get("mark", 1)),
                    message=message,
                )
            )
    return results


def write_detail_csv(results: list[CheckResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(CheckResult.__dataclass_fields__.keys()))
        writer.writeheader()
        for result in results:
            writer.writerow(result.__dict__)


def write_summary_csv(results: list[CheckResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    grouped = defaultdict(list)
    for result in results:
        grouped[(result.class_group, result.learner)].append(result)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["class_group", "learner", "auto_score", "possible_total", "passed", "failed", "manual", "errors"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for (class_group, learner), learner_results in sorted(grouped.items()):
            writer.writerow(
                {
                    "class_group": class_group,
                    "learner": learner,
                    "auto_score": sum(r.mark_awarded for r in learner_results),
                    "possible_total": sum(r.mark_possible for r in learner_results),
                    "passed": sum(1 for r in learner_results if r.status == "pass"),
                    "failed": sum(1 for r in learner_results if r.status == "fail"),
                    "manual": sum(1 for r in learner_results if r.status == "manual"),
                    "errors": sum(1 for r in learner_results if r.status == "error"),
                }
            )


def split_combined_check_rows(workbook: Any, check_cells: dict[str, list[str]]) -> dict[str, list[str]]:
    combined_cells: dict[str, list[str]] = defaultdict(list)
    for check_id, cell_refs in check_cells.items():
        for cell_ref in cell_refs:
            combined_cells[cell_ref].append(check_id)

    replacement_cells: dict[str, list[str]] = {}
    inserted_rows: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for sheet_name in ("Q1", "Q2", "Q3", "Q4"):
        ws = workbook[sheet_name]

        sheet_combined = [
            (coordinate_to_tuple(cell_ref.split("!", 1)[1])[0], cell_ref, check_ids)
            for cell_ref, check_ids in combined_cells.items()
            if cell_ref.startswith(f"{sheet_name}!") and len(check_ids) > 1
        ]
        row_shift = 0
        for original_row, _cell_ref, check_ids in sorted(sheet_combined):
            row = original_row + row_shift
            for offset, check_id in enumerate(check_ids):
                if offset:
                    ws.insert_rows(row + offset)
                    ws.row_dimensions[row + offset].height = ws.row_dimensions[row + offset - 1].height
                replacement_cells[check_id] = [f"{sheet_name}!C{row + offset}"]
            extra_rows = len(check_ids) - 1
            inserted_rows[sheet_name].append((original_row, extra_rows))
            row_shift += extra_rows

    shifted_cells: dict[str, list[str]] = {}
    for check_id, cell_refs in check_cells.items():
        if check_id in replacement_cells:
            shifted_cells[check_id] = replacement_cells[check_id]
            continue

        shifted_refs = []
        for cell_ref in cell_refs:
            sheet_name, coordinate = cell_ref.split("!", 1)
            row, column = coordinate_to_tuple(coordinate)
            shift = sum(extra_rows for insert_after, extra_rows in inserted_rows[sheet_name] if row > insert_after)
            shifted_refs.append(f"{sheet_name}!{get_column_letter(column)}{row + shift}")
        shifted_cells[check_id] = shifted_refs
    return shifted_cells


def write_marksheets(root: Path, expectations: dict[str, Any], results: list[CheckResult]) -> list[Path]:
    marksheet_dir = root / "Marksheets"
    marksheet_dir.mkdir(parents=True, exist_ok=True)

    template_path = root / "MArksheet gr 10.xlsx"

    grouped = defaultdict(list)
    for result in results:
        grouped[(result.class_group, result.learner)].append(result)

    workbook_paths: list[Path] = []

    check_cells: dict[str, list[str]] = {
        "Q1.1.1": ["Q1!C3"],
        "Q1.1.2a": ["Q1!C5"],
        "Q1.1.2b": ["Q1!C6"],
        "Q1.1.3": ["Q1!C8"],
        "Q1.2.1": ["Q1!C10"],
        "Q1.2.2": ["Q1!C12"],
        "Q1.2.3a": ["Q1!C14"],
        "Q1.2.3b": ["Q1!C15"],
        "Q1.3a": ["Q1!C17"],
        "Q1.3b": ["Q1!C18"],
        "Q1.3c": ["Q1!C19"],
        "Q1.3d": ["Q1!C20"],
        "Q1.3e": ["Q1!C21"],
        "Q1.3f": ["Q1!C22"],
        "Q1.3g": ["Q1!C23"],
        "Q1.4.1a": ["Q1!C25"],
        "Q1.4.1b": ["Q1!C26"],
        "Q1.4.2a": ["Q1!C28"],
        "Q1.4.2b": ["Q1!C29"],
        "Q1.4.3": ["Q1!C31"],
        "Q1.4.4a": ["Q1!C33"],
        "Q1.4.4b": ["Q1!C34"],
        "Q1.4.4c": ["Q1!C35"],
        "Q1.4.4d": ["Q1!C36"],
        "Q1.5a": ["Q1!C38"],
        "Q1.5b": ["Q1!C39"],
        "Q1.5c": ["Q1!C40"],
        "Q1.6a": ["Q1!C42"],
        "Q1.6b": ["Q1!C43"],
        "Q1.7a": ["Q1!C45"],
        "Q1.7b": ["Q1!C46"],
        "Q2.1.1": ["Q2!C3"],
        "Q2.1.2a": ["Q2!C5"],
        "Q2.1.2b": ["Q2!C6"],
        "Q2.1.3a": ["Q2!C8"],
        "Q2.1.3b": ["Q2!C9"],
        "Q2.1.4a": ["Q2!C11"],
        "Q2.1.4b": ["Q2!C12"],
        "Q2.1.5": ["Q2!C14"],
        "Q2.2a": ["Q2!C16"],
        "Q2.2b": ["Q2!C17"],
        "Q2.2c": ["Q2!C18"],
        "Q2.3": ["Q2!C20"],
        "Q2.4a": ["Q2!C22"],
        "Q2.4b": ["Q2!C23"],
        "Q2.4c": ["Q2!C24"],
        "Q2.4d": ["Q2!C25"],
        "Q2.4e": ["Q2!C26"],
        "Q3.1.1": ["Q3!C3"],
        "Q3.1.2": ["Q3!C5"],
        "Q3.1.3a": ["Q3!C7"],
        "Q3.1.3b": ["Q3!C8"],
        "Q3.1.3c": ["Q3!C9"],
        "Q3.1.3d": ["Q3!C10"],
        "Q3.2": ["Q3!C13"],
        "Q3.3": ["Q3!C15"],
        "Q3.4a": ["Q3!C18"],
        "Q3.4b": ["Q3!C19"],
        "Q3.4c": ["Q3!C20"],
        "Q3.5a": ["Q3!C24"],
        "Q3.5b": ["Q3!C25"],
        "Q3.5c": ["Q3!C26"],
        "Q3.5d": ["Q3!C27"],
        "Q3.6a": ["Q3!C30"],
        "Q3.6b": ["Q3!C31"],
        "Q3.7a": ["Q3!C35"],
        "Q3.7b": ["Q3!C36"],
        "Q3.8": ["Q3!C39"],
        "Q3.9": ["Q3!C41"],
        "Q3.10a": ["Q3!C45"],
        "Q3.10b": ["Q3!C46"],
        "Q3.10c": ["Q3!C47"],
        "Q3.10d": ["Q3!C48"],
        "Q3.11a": ["Q3!C52"],
        "Q3.11b": ["Q3!C53"],
        "Q3.11c": ["Q3!C54"],
        "Q3.11d": ["Q3!C55"],
        "Q4.1.1": ["Q4!C3"],
        "Q4.1.2a": ["Q4!C5"],
        "Q4.1.2b": ["Q4!C6"],
        "Q4.1.3a": ["Q4!C9"],
        "Q4.1.3b": ["Q4!C10"],
        "Q4.1.3c": ["Q4!C11"],
        "Q4.2a": ["Q4!C14"],
        "Q4.2b": ["Q4!C15"],
        "Q4.2c": ["Q4!C16"],
        "Q4.2d": ["Q4!C17"],
        "Q4.3a": ["Q4!C21"],
        "Q4.3b": ["Q4!C22"],
        "Q4.4a": ["Q4!C26"],
        "Q4.4b": ["Q4!C27"],
        "Q4.4c": ["Q4!C28"],
        "Q4.4d": ["Q4!C29"],
        "Q4.4e": ["Q4!C30"],
        "Q4.5a": ["Q4!C34"],
        "Q4.5b": ["Q4!C35"],
        "Q4.6": ["Q4!C38"],
        "Q4.7a": ["Q4!C41"],
        "Q4.7b": ["Q4!C42"],
        "Q4.8": ["Q4!C45"],
    }

    for (class_group, learner), learner_results in sorted(grouped.items()):
        filename = f"{learner}.xlsx"
        workbook_path = marksheet_dir / filename

        shutil.copyfile(template_path, workbook_path)

        workbook = load_workbook(workbook_path)
        summary_ws = workbook["Summary"]
        summary_ws["B3"] = filename
        learner_check_cells = split_combined_check_rows(workbook, check_cells)

        cell_results: dict[str, list[CheckResult]] = defaultdict(list)
        for result in learner_results:
            for cell_ref in learner_check_cells.get(result.check_id, []):
                cell_results[cell_ref].append(result)

        for cell_ref, mapped_results in cell_results.items():
            sheet_name, coordinate = cell_ref.split("!", 1)
            cell = workbook[sheet_name][coordinate]
            cell.value = int(any(result.status == "pass" for result in mapped_results))

        workbook.save(workbook_path)
        workbook_paths.append(workbook_path)

    return workbook_paths


def main() -> None:
    parser = argparse.ArgumentParser(description="Check F2/F3 learner files against structured expectations.")
    parser.add_argument("--root", default=".", help="Exam root folder containing F2 and F3.")
    parser.add_argument("--expectations", default="structured_expectations.json")
    parser.add_argument("--groups", nargs="+", default=["F2", "F3"])
    parser.add_argument("--out-dir", default="marking_output")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    out_dir = root / args.out_dir
    results = run_checks(root, root / args.expectations, args.groups)
    write_summary_csv(results, out_dir / "summary.csv")
    write_detail_csv(results, out_dir / "detail.csv")

    expectations = json.loads((root / args.expectations).read_text(encoding="utf-8"))
    workbook_paths = write_marksheets(root, expectations, results)

    learners = {(r.class_group, r.learner) for r in results}
    print(f"Checked {len(learners)} learner folders.")
    print(f"Wrote {out_dir / 'summary.csv'}")
    print(f"Wrote {out_dir / 'detail.csv'}")
    print(f"Wrote {len(workbook_paths)} learner mark sheets to {root / 'Marksheets'}")


if __name__ == "__main__":
    main()
