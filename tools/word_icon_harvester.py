from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import subprocess
import sys
import urllib.request
from pathlib import Path

from PIL import Image

try:
    import pythoncom
    import win32com.client
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"pywin32 is required for this script: {exc}")

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required for this script: {exc}")

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
VBS_HELPER = SCRIPT_DIR / "word_icon_save.vbs"
VBA_EXPORTER = SCRIPT_DIR / "word_icon_exporter.bas"
BERT_IMAGEMSO_LIST_URL = "https://bert-toolkit.com/imagemso-list.html"
BERT_IMAGEMSO_SPRITE_URL = "https://bert-toolkit.com/img/mso-composite-16.png"

IMAGE_MSO_ALIASES = {
    "AcceptChange": ["ReviewAcceptChange"],
    "AddIns": ["AddInManager", "AddInsDialog"],
    "AddTextToTableOfContents": ["TableOfContentsAddText", "TableOfContentsGallery"],
    "BordersGallery": ["BordersAll", "BorderOutside", "BorderBottom"],
    "CitationsManageSources": ["BibliographyManageSources", "CitationInsert"],
    "ColumnsGallery": ["ColumnsDialog"],
    "ComAddIns": ["ComAddInsDialog", "AddIns"],
    "CompareDocuments": ["CompareAndCombine", "ReviewCompareMenu", "ReviewCompareTwoVersions", "ReviewCompareDocuments", "ReviewCompare"],
    "ContentControlDatePicker": ["ContentControlDate", "CalendarInsert", "DatePickerInsert"],
    "ContentControlGroup": ["ObjectsGroup", "ObjectsGroupMenu"],
    "ContentControlProperties": ["PropertySheet", "ContentControlRichText"],
    "DeleteComment": ["ReviewDeleteComment"],
    "EndnoteInsert": ["FootnoteInsert"],
    "FontSizeDecrease": ["FontSizeDecreaseWord", "FontSizeDecrease1Point"],
    "FontSizeIncrease": ["FontSizeIncreaseWord", "FontSizeIncrease1Point"],
    "FootnotesShow": ["FootnotesEndnotesShow", "FootnoteInsert"],
    "HyphenationGallery": ["HyphenationMenu", "HyphenationOptions", "HyphenationOptionsDialog"],
    "LegacyTools": ["FormControlButton", "ControlToolboxOutlook"],
    "LineNumberingGallery": ["LineNumbersRestartEachPage", "Numbering"],
    "LineSpacing": ["ParagraphSpacingIncrease", "ParagraphSpacingDecrease"],
    "Macros": ["MacroPlay", "MacroDefault", "VisualBasic"],
    "MailMergeCheckForErrors": ["MailMergeAutoCheckForErrors"],
    "MailMergeEditRecipientList": ["MailMergeRecipientsEditList", "DataFormSource"],
    "MailMergeFinishAndMerge": ["MailMergeMergeToDocument", "MailMergeMergeToPrinter"],
    "MailMergeInsertMergeField": ["MailMergeFieldInsert", "MailMergeHelper"],
    "MailMergePreviewResults": ["MailMergeResultsPreview"],
    "MailMergeStart": ["MailMergeDocument"],
    "MultiLevelListGallery": ["OutlineNumbering", "Numbering"],
    "NewComment": ["ReviewNewComment"],
    "NextChange": ["ReviewNextChange"],
    "NextComment": ["ReviewNextComment"],
    "ObjectPositionGallery": ["AboveText", "BehindText"],
    "ObjectsAlignMenu": ["ObjectsAlignLeft", "ObjectsAlignCenterHorizontal"],
    "PageBordersDialog": ["BordersAndShadingDialog", "BorderOutside"],
    "PageInsert": ["PageBreakInsertWord", "FileNew"],
    "PreviousChange": ["ReviewPreviousChange"],
    "PreviousComment": ["ReviewPreviousComment"],
    "RecordMacro": ["MacroRecord"],
    "RejectChange": ["ReviewRejectChange"],
    "RestrictEditing": ["ReviewProtectDocument", "Lock"],
    "ReviewShowMarkup": ["ReviewShowOrHideComment", "ReviewDisplayForReview"],
    "ReviewingPane": ["ReviewingPaneVertical", "ReviewDisplayForReview"],
    "SetAsDefault": ["FileSaveAs", "ApplyStyles"],
    "SymbolInsert": ["SymbolInsertDialog", "SymbolInsert"],
    "TableInsertDialog": ["TableInsert"],
    "TextWrappingGallery": ["TextWrappingSquare", "TextWrappingTight"],
    "TrackChanges": ["ReviewTrackChanges", "ReviewDisplayForReview"],
    "ViewMacros": ["MacroPlay", "VisualBasic"],
    "ViewOnePage": ["ZoomOnePage"],
    "ViewPageWidth": ["ZoomPageWidth"],
    "ViewRuler": ["RulerShowHide", "ViewRulerPowerPoint"],
    "ViewTwoPages": ["MultiplePages"],
    "WindowArrangeAll": ["WindowsArrangeAll"],
    "WindowSynchronousScrolling": ["ViewSideBySide"],
    "WindowViewSideBySide": ["ViewSideBySide"],
    "XmlMappingPane": ["XmlSource", "XmlStructure"],
    "Zoom100Percent": ["Zoom100"],
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def looks_like_idmso(value):
    text = clean(value)
    if not text:
        return False
    if len(text) > 150:
        return False
    return bool(re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", text))


def extract_ids_from_workbook(filename: Path):
    if not filename.exists():
        raise FileNotFoundError(f"Workbook not found: {filename}")

    workbook = load_workbook(filename, read_only=True, data_only=True)
    candidates = set()

    interesting_headers = (
        "control name",
        "command name",
        "idmso",
        "id mso",
        "office id",
        "control id",
    )

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40), values_only=True))

        discovered_header_positions = set()

        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                header = clean(value).lower()
                if any(key == header or key in header for key in interesting_headers):
                    discovered_header_positions.add((row_index, col_index))

        if discovered_header_positions:
            for header_row, col_index in discovered_header_positions:
                for row in sheet.iter_rows(min_row=header_row + 1, min_col=col_index, max_col=col_index, values_only=True):
                    value = clean(row[0])
                    if looks_like_idmso(value):
                        candidates.add(value)
        else:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    text = clean(value)
                    if looks_like_idmso(text):
                        candidates.add(text)

    workbook.close()
    return sorted(candidates)


def get_label_from_word(id_mso: str):
    pythoncom.CoInitialize()
    word = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        return clean(word.CommandBars.GetLabelMso(id_mso))
    finally:
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def save_icon_via_vbs(id_mso: str, output_path: Path, size: int = 128):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not VBS_HELPER.exists():
        raise FileNotFoundError(f"VBScript helper not found: {VBS_HELPER}")

    proc = subprocess.run(
        [
            "cscript.exe",
            "//NoLogo",
            str(VBS_HELPER),
            id_mso,
            str(output_path),
            str(size),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    if proc.returncode != 0:
        error_text = (proc.stdout or "") + (proc.stderr or "")
        raise RuntimeError(f"VBS save failed for {id_mso}: {error_text.strip()}")

    return output_path.exists()


class WordVbaIconExporter:
    def __init__(self, size: int = 128):
        self.size = size
        self.word = None
        self.document = None

    def __enter__(self):
        if not VBA_EXPORTER.exists():
            raise FileNotFoundError(f"VBA exporter not found: {VBA_EXPORTER}")

        pythoncom.CoInitialize()
        self.word = win32com.client.DispatchEx("Word.Application")
        self.word.Visible = False
        self.word.DisplayAlerts = 0
        self.document = self.word.Documents.Add()
        try:
            self.document.VBProject.VBComponents.Import(str(VBA_EXPORTER))
        except Exception as exc:
            raise RuntimeError(
                "Could not import VBA exporter. Enable 'Trust access to the VBA project object model' "
                "in Word Trust Center, then rerun the harvester."
            ) from exc
        return self

    def __exit__(self, exc_type, exc, traceback):
        if self.document is not None:
            try:
                self.document.Close(False)
            except Exception:
                pass
        if self.word is not None:
            try:
                self.word.Quit(False)
            except Exception:
                pass
        pythoncom.CoUninitialize()

    def label(self, id_mso: str):
        return clean(self.word.CommandBars.GetLabelMso(id_mso))

    def save(self, id_mso: str, output_path: Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        ok = self.word.Run("ExportImageMso", id_mso, str(output_path), int(self.size))
        return bool(ok) and output_path.exists()


def save_png_from_bitmap(bmp_path: Path, png_path: Path):
    with Image.open(bmp_path) as img:
        if img.mode not in {"RGBA", "RGB", "L", "P", "LA"}:
            img = img.convert("RGBA")
        else:
            img = img.convert("RGBA")
        img.save(png_path, format="PNG")


def write_csv(path: Path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def extract_ids_from_layout(filename: Path):
    with open(filename, "r", encoding="utf-8") as handle:
        data = json.load(handle)

    ids = set()

    def walk(value):
        if isinstance(value, dict):
            id_mso = clean(value.get("idMso"))
            image = clean(value.get("image"))
            if id_mso and image:
                ids.add(id_mso)
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(data)
    return sorted(ids)


def harvest_from_public_gallery(ids, output_dir: Path, size: int = 128):
    print("Downloading public imageMso sprite fallback...")
    html = urllib.request.urlopen(BERT_IMAGEMSO_LIST_URL, timeout=30).read().decode("utf-8", errors="replace")
    sprite_bytes = urllib.request.urlopen(BERT_IMAGEMSO_SPRITE_URL, timeout=30).read()
    sprite = Image.open(io.BytesIO(sprite_bytes)).convert("RGBA")

    positions = {}
    pattern = re.compile(
        r"<a name=([^>\s]+)></a>\s*([^<]+)</td><td><div class=composite style=\"background-position:0 -?(\d+)px\"",
        re.IGNORECASE,
    )
    for anchor, label, offset in pattern.findall(html):
        positions[anchor] = (clean(label), int(offset))

    output_dir.mkdir(parents=True, exist_ok=True)
    rows = []

    for index, id_mso in enumerate(ids, start=1):
        print(f"[fallback {index}/{len(ids)}] {id_mso}")
        item = positions.get(id_mso)
        if item is None:
            print("  MISSING_FROM_PUBLIC_GALLERY")
            continue

        label, y = item
        icon = sprite.crop((0, y, 16, y + 16))
        if size != 16:
            icon = icon.resize((size, size), Image.Resampling.LANCZOS)

        png_path = output_dir / f"{id_mso}.png"
        icon.save(png_path, format="PNG")
        rows.append({
            "idMso": id_mso,
            "Label": label,
            "BitmapPath": "",
            "PngPath": str(png_path),
            "Source": "BERT Office 2016 imageMso sprite",
        })
        print(f"  OK: saved {png_path.name}")

    return rows


def harvest_ids_via_vba(ids, output_dir: Path, size: int = 128):
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = []

    with WordVbaIconExporter(size=size) as exporter:
        for index, id_mso in enumerate(ids, start=1):
            print(f"[vba {index}/{len(ids)}] {id_mso}")
            label = ""
            try:
                label = exporter.label(id_mso)
            except Exception as exc:
                print(f"  LABEL_UNAVAILABLE: {exc}")

            candidates = [id_mso] + IMAGE_MSO_ALIASES.get(id_mso, [])

            bmp_path = output_dir / f"{id_mso}.bmp"
            png_path = output_dir / f"{id_mso}.png"

            for image_mso in candidates:
                try:
                    if exporter.save(image_mso, bmp_path):
                        if not label:
                            try:
                                label = exporter.label(image_mso)
                            except Exception:
                                label = id_mso
                        save_png_from_bitmap(bmp_path, png_path)
                        rows.append({
                            "idMso": id_mso,
                            "Label": label,
                            "BitmapPath": str(bmp_path),
                            "PngPath": str(png_path),
                            "Source": f"Local Word VBA GetImageMso:{image_mso}",
                        })
                        if image_mso == id_mso:
                            print(f"  OK: saved {png_path.name}")
                        else:
                            print(f"  OK: saved {png_path.name} via {image_mso}")
                        break
                except Exception as exc:
                    print(f"  SAVE_FAILED {image_mso}: {exc}")
            else:
                print("  NO FILE CREATED")

    return rows


def harvest_sample_ids(ids, output_dir: Path, size: int = 128):
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = []

    for index, id_mso in enumerate(ids, start=1):
        print(f"[{index}/{len(ids)}] {id_mso}")
        try:
            label = get_label_from_word(id_mso)
        except Exception as exc:
            print(f"  INVALID: {exc}")
            continue

        if not label:
            print("  INVALID: no label returned")
            continue

        bmp_path = output_dir / f"{id_mso}.bmp"
        png_path = output_dir / f"{id_mso}.png"

        try:
            save_icon_via_vbs(id_mso, bmp_path, size=size)
            if bmp_path.exists():
                    save_png_from_bitmap(bmp_path, png_path)
                    rows.append({
                        "idMso": id_mso,
                        "Label": label,
                        "BitmapPath": str(bmp_path),
                        "PngPath": str(png_path),
                        "Source": "Local Word VBS GetImageMso",
                    })
                    print(f"  OK: saved {png_path.name}")
            else:
                print("  NO FILE CREATED")
        except Exception as exc:  # pragma: no cover
            print(f"  SAVE_FAILED: {exc}")

    return rows


def can_export_local_icons(size: int):
    try:
        bmp_path = Path(os.environ.get("TEMP", str(SCRIPT_DIR))) / "word_icon_harvester_preflight_vba.bmp"
        if bmp_path.exists():
            bmp_path.unlink()
        with WordVbaIconExporter(size=size) as exporter:
            exporter.label("Bold")
            exists = exporter.save("Bold", bmp_path)
        if exists:
            bmp_path.unlink()
        return exists
    except Exception as exc:
        print(f"Local Word image export unavailable: {exc}")
        return False


def main():
    parser = argparse.ArgumentParser(description="Word 2021 icon harvester")
    parser.add_argument("--source", type=str, default="", help="Path to Microsoft wordcontrols.xlsx")
    parser.add_argument("--ids", type=str, default="", help="Comma-separated IDs to test (e.g. Bold,Paste,Underline)")
    parser.add_argument("--size", type=int, default=128, help="Icon size to request from Word")
    parser.add_argument("--layout", type=str, default=str(PROJECT_ROOT / "Practical" / "Layout" / "word2021_ribbon_layout_full.json"), help="Path to local ribbon layout JSON")
    parser.add_argument("--output", type=str, default=str(PROJECT_ROOT / "Practical" / "Images" / "Word"), help="Directory for harvested images")
    parser.add_argument("--fallback-public-gallery", action="store_true", help="Use public Office 2016 imageMso sprite for icons that local Office cannot export")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.ids:
        ids = [p.strip() for p in args.ids.split(",") if p.strip()]
    elif args.source:
        ids = extract_ids_from_workbook(Path(args.source))
    elif args.layout:
        ids = extract_ids_from_layout(Path(args.layout))
    else:
        ids = ["Bold", "Paste", "Underline", "ParagraphMarks", "NavigationPaneShowHide"]

    print("Starting Word icon harvest...")
    rows = []
    if can_export_local_icons(args.size):
        rows = harvest_ids_via_vba(ids, output_dir, size=args.size)
    else:
        print("Skipping local export and continuing with configured fallback.")

    if args.fallback_public_gallery:
        harvested = {row["idMso"] for row in rows}
        missing_ids = [id_mso for id_mso in ids if id_mso not in harvested]
        rows.extend(harvest_from_public_gallery(missing_ids, output_dir, size=args.size))

    write_csv(
        output_dir / "word_icons.csv",
        rows,
        ["idMso", "Label", "BitmapPath", "PngPath", "Source"],
    )

    print()
    print(f"Saved {len(rows)} valid icons to {output_dir}")
    print("Most important file:")
    print(output_dir / "word_icons.csv")


if __name__ == "__main__":
    main()
