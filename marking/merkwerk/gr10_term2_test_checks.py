#!/usr/bin/env python3
"""
Grade 10 CAT practical marker.

Checks every automatable item in AGAIN Memo.xlsx and writes marks into the
cells that the memo totals actually use.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

try:
    import win32com.client
except ImportError:
    win32com = None


BASE_DIR = Path(r"E:\RTT melkies\Merk werk 2026\MErk werk\GR10 TERM 2 TEST")
MEMO_FILE = BASE_DIR / "AGAIN Memo.xlsx"
MARKSHEETS_DIR = BASE_DIR / "marksheets"
FP_G2_DIR = BASE_DIR / "FP G2 Gr10"
FP_G3_DIR = BASE_DIR / "FP G3 Gr10"

CM_TO_POINTS = 28.3464567


def norm_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\r", " ").replace("\x07", " ")).strip()


def is_red_color(value) -> bool:
    if value in (None, -16777216, 0):
        return False
    try:
        return int(value) in (255, 192, 16711680)
    except Exception:
        return False


def excel_color_is_red(color) -> bool:
    if not color:
        return False
    rgb = getattr(color, "rgb", None)
    indexed = getattr(color, "indexed", None)
    if rgb and str(rgb).upper().endswith("FF0000"):
        return True
    return indexed == 10


def cell_in_merged(ws, coordinate: str) -> bool:
    return any(coordinate in merged for merged in ws.merged_cells.ranges)


def writable_coordinate(ws, coordinate: str) -> str:
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).coordinate
    return coordinate


def set_mark(marks: dict[int, int], row: int, passed: bool) -> None:
    marks[row] = 1 if passed else 0


class WordReader:
    def __init__(self):
        self.app = None
        if win32com:
            self.app = win32com.client.DispatchEx("Word.Application")
            try:
                self.app.Visible = False
            except Exception:
                pass
            try:
                self.app.DisplayAlerts = 0
            except Exception:
                pass

    def open(self, path: Path):
        if not self.app or not path.exists():
            return None
        try:
            return self.app.Documents.Open(str(path.resolve()), ReadOnly=True, AddToRecentFiles=False)
        except Exception:
            return None

    def close_doc(self, doc) -> None:
        try:
            if doc:
                doc.Close(False)
        except Exception:
            pass

    def close(self) -> None:
        try:
            if self.app:
                self.app.Quit()
        except Exception:
            pass


class AutoMarker:
    def __init__(self, learner_name: str, learner_path: Path, word: WordReader):
        self.learner_name = learner_name
        self.learner_path = learner_path
        self.word = word
        self.marks = {"Sheet1": {}, "Sheet2": {}, "Sheet3": {}, "Sheet4": {}}

    def mark(self) -> None:
        self.marks["Sheet1"] = self.check_newsletter(self.learner_path / "1TourNews.docx")
        self.marks["Sheet2"] = self.check_venue(self.learner_path / "2Venue.docx")
        self.marks["Sheet3"] = self.check_wildfires(self.find_wildfires_workbook())
        self.marks["Sheet4"] = self.check_learners(self.learner_path / "4Learners.docx")
        self.update_marksheet()

    def find_wildfires_workbook(self) -> Path:
        exact = self.learner_path / "3GlobalWildfires.xlsx"
        candidates = []
        for path in self.learner_path.glob("*.xls*"):
            name = path.name.lower()
            if path.name.startswith("~$"):
                continue
            if any(token in name for token in ("global", "wildfire", "fire")):
                candidates.append(path)

        def workbook_priority(path: Path) -> tuple[int, float, str]:
            name = path.stem.lower().replace(" ", "")
            if name.startswith("globalfires"):
                rank = 0
            elif name.startswith("globalwildfires"):
                rank = 1
            elif path.name.lower() == "3globalwildfires.xlsx":
                rank = 2
            else:
                rank = 3
            return (rank, -path.stat().st_mtime, path.name.lower())

        for path in sorted(candidates, key=workbook_priority):
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
                try:
                    sheet_names = " ".join(wb.sheetnames).lower()
                    first = wb[wb.sheetnames[0]]
                    title = " ".join(str(first[cell].value or "") for cell in ("A1", "B1", "C1", "D1", "E1")).lower()
                    if "global" in sheet_names or "fire" in sheet_names or "global" in title or "fire" in title:
                        return path
                finally:
                    wb.close()
            except Exception:
                continue
        return exact

    def check_newsletter(self, path: Path) -> dict[int, int]:
        marks: dict[int, int] = {row: 0 for row in (3, 5, 6, 8, 9, 12, 13, 14, 15, 17, 22, 24, 26, 28, 29)}
        doc = self.word.open(path)
        if not doc:
            return marks

        try:
            text = norm_text(doc.Content.Text)
            lower = text.lower()
            heading_range = None
            found = doc.Content.Find
            found.ClearFormatting()
            found.Text = "Blackwood High"
            if found.Execute():
                heading_range = found.Parent

            set_mark(marks, 3, bool(heading_range and "calibri" in str(heading_range.Font.Name).lower()))
            set_mark(marks, 5, bool(heading_range and round(float(heading_range.Font.Size), 1) == 26.0))
            set_mark(marks, 6, bool(heading_range and is_red_color(heading_range.Font.Color)))

            school_at = lower.find("school trip")
            dear_at = lower.find("dear parents")
            set_mark(marks, 8, school_at >= 0 and dear_at > school_at)

            image_width_ok = self.top_page_image_width_is_6cm(doc)
            set_mark(marks, 9, image_width_ok)
            set_mark(marks, 12, "8" in text)
            set_mark(marks, 13, True)
            set_mark(marks, 14, self.images_have_no_border(doc))
            set_mark(marks, 15, self.paragraph_below_heading_is_justified(doc, "Wrapping up"))
            set_mark(marks, 17, self.only_em_dashes_in_target_text(text))
            set_mark(marks, 22, self.text_has_heading_or_title_style(doc, "Appendix: Places of Interest"))
            set_mark(marks, 24, self.red_sentence_case_under_isimangaliso(doc))

            footer_text = " ".join(
                norm_text(sec.Footers(index).Range.Text)
                for sec in doc.Sections
                for index in range(1, 4)
            )
            has_footer = bool(footer_text)
            has_page_fields = self.has_page_number_fields(doc)
            set_mark(marks, 26, has_page_fields or re.search(r"page\s+\d+\s+of\s+\d+", footer_text, re.I))
            set_mark(marks, 28, has_footer)
            set_mark(marks, 29, "page" in footer_text.lower() and "of" in footer_text.lower())
        finally:
            self.word.close_doc(doc)

        return marks

    def check_venue(self, path: Path) -> dict[int, int]:
        marks: dict[int, int] = {row: 0 for row in (3, 5, 7, 8, 9, 11, 12, 13, 14, 17, 18, 19, 20)}
        doc = self.word.open(path)
        if not doc:
            return marks

        try:
            section = doc.Sections(1)
            setup = section.PageSetup
            set_mark(marks, 3, self.page_background_is_white(doc))
            set_mark(marks, 5, int(setup.Orientation) == 0)
            set_mark(marks, 7, abs(float(setup.PageWidth) - 595.3) <= 8 and abs(float(setup.PageHeight) - 841.9) <= 8)
            set_mark(marks, 8, abs(float(setup.RightMargin) - (2.5 * CM_TO_POINTS)) <= 3)

            venue_paragraphs = self.find_venue_paragraphs(doc)
            set_mark(marks, 9, any(abs(float(p.Format.LeftIndent)) > 1 or abs(float(p.Format.FirstLineIndent)) > 1 for p in venue_paragraphs))
            set_mark(marks, 11, any(abs(float(p.Format.LeftIndent) - (2 * CM_TO_POINTS)) <= 6 or abs(float(p.Format.FirstLineIndent) - (2 * CM_TO_POINTS)) <= 6 for p in venue_paragraphs))
            set_mark(marks, 12, bool(venue_paragraphs) and all(int(p.Format.LineSpacingRule) == 5 for p in venue_paragraphs))
            set_mark(marks, 13, bool(venue_paragraphs) and all(abs(float(p.Format.LineSpacing) - 14.4) <= 1.5 for p in venue_paragraphs))

            experience = self.find_text_range(doc, "experience")
            enjoy = self.find_text_range(doc, "Enjoy")
            set_mark(marks, 14, bool(experience and experience.Font.Superscript))
            set_mark(marks, 17, bool(enjoy and enjoy.Font.SmallCaps))
            set_mark(marks, 18, "\f" not in doc.Content.Text)

            red_ranges = self.red_text_ranges(doc)
            set_mark(marks, 19, bool(red_ranges and all(r.Font.Bold for r in red_ranges)))
            set_mark(marks, 20, bool(red_ranges and all(r.Font.Italic for r in red_ranges)))
        finally:
            self.word.close_doc(doc)

        return marks

    def check_learners(self, path: Path) -> dict[int, int]:
        marks: dict[int, int] = {row: 0 for row in (3, 5, 7, 8)}
        doc = self.word.open(path)
        if doc:
            try:
                text = norm_text(doc.Content.Text).lower()
                set_mark(marks, 3, "inlone" not in text and "inline" in text)
                set_mark(marks, 5, "themas" not in text and "themes" in text and "applie" not in text and "apply" in text)
            finally:
                self.word.close_doc(doc)

        pdf_path = self.find_learners_pdf()
        set_mark(marks, 7, pdf_path is not None)
        set_mark(marks, 8, pdf_path is not None)
        return marks

    def find_learners_pdf(self) -> Path | None:
        pdfs = [p for p in self.learner_path.glob("*.pdf") if not p.name.startswith("~$")]
        if not pdfs:
            return None

        preferred = []
        for pdf in pdfs:
            name = re.sub(r"[^a-z0-9]+", "", pdf.stem.lower())
            if "4learnernew" in name or "4learnersnew" in name:
                preferred.append(pdf)
        if preferred:
            return sorted(preferred)[0]

        learner_pdfs = []
        for pdf in pdfs:
            name = re.sub(r"[^a-z0-9]+", "", pdf.stem.lower())
            if "4learner" in name or "4learners" in name:
                learner_pdfs.append(pdf)
        if learner_pdfs:
            return sorted(learner_pdfs)[0]
        return sorted(pdfs)[0]

    def check_wildfires(self, path: Path) -> dict[int, int]:
        marks: dict[int, int] = {row: 0 for row in (2, 3, 4, 6, 8, 10, 11, 13, 14, 16, 17, 19, 20, 22, 23, 24, 25)}
        if not path.exists():
            return marks

        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb.active
        try:
            set_mark(marks, 2, ws.row_dimensions[1].height is not None and abs(ws.row_dimensions[1].height - 40) <= 0.5)
            set_mark(marks, 3, self.has_merged(ws, "A1:E1") and ws["A1"].alignment.horizontal == "center")
            set_mark(marks, 4, all(ws[f"{col}2"].font.bold for col in "CDE"))
            set_mark(marks, 6, all(ws[f"{col}2"].font.size == 14 for col in "CDE") and self.has_merged(ws, "C2:E2"))
            set_mark(marks, 8, self.c2e2_matches_a1_visual(ws))
            set_mark(marks, 10, all(self.is_currency(ws[f"D{row}"]) for row in range(3, 8)))

            d9 = self.formula(ws["D9"].value)
            set_mark(marks, 11, d9.startswith("=MAX("))
            set_mark(marks, 13, "C3:C7" in d9)
            d10 = self.formula(ws["D10"].value)
            set_mark(marks, 14, d10.startswith("=AVERAGE("))
            set_mark(marks, 16, "D3:D7" in d10)
            d11 = self.formula(ws["D11"].value)
            set_mark(marks, 17, d11.startswith("=SUM("))
            set_mark(marks, 19, "C3:C7" in d11)
            d12 = self.formula(ws["D12"].value, ws["D12"])
            set_mark(marks, 20, "D3" in d12)
            set_mark(marks, 22, "*" in d12)
            set_mark(marks, 23, "25%" in d12 or "25/100" in d12 or "0.25" in d12)
            set_mark(marks, 24, ws.title == "GlobalFires")
            set_mark(marks, 25, excel_color_is_red(ws.sheet_properties.tabColor) or self.workbook_tab_is_red(path, ws.title))
        finally:
            wb.close()
        return marks

    def update_marksheet(self) -> None:
        marksheet = MARKSHEETS_DIR / f"{self.learner_name} - Marksheet.xlsx"
        if not marksheet.exists():
            print(f"  missing marksheet: {marksheet.name}")
            return

        try:
            wb = openpyxl.load_workbook(marksheet)
        except PermissionError:
            print(f"  locked marksheet skipped: {marksheet.name}")
            return
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        try:
            for sheet_name, sheet_marks in self.marks.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for row, value in sheet_marks.items():
                    coordinate = writable_coordinate(ws, f"C{row}")
                    ws[coordinate].value = value
                    ws[coordinate].fill = yellow if value == 0 else PatternFill(fill_type=None)
            try:
                wb.save(marksheet)
            except PermissionError:
                if not self.update_open_excel_marksheet(marksheet):
                    print(f"  locked marksheet skipped: {marksheet.name}")
        finally:
            wb.close()

    def update_open_excel_marksheet(self, marksheet: Path) -> bool:
        if not win32com:
            return False
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            try:
                excel = win32com.client.Dispatch("Excel.Application")
            except Exception:
                return False

        workbook = None
        target = str(marksheet.resolve()).lower()
        try:
            for i in range(1, excel.Workbooks.Count + 1):
                candidate = excel.Workbooks(i)
                if str(candidate.FullName).lower() == target:
                    workbook = candidate
                    break
            if workbook is None:
                workbook = excel.Workbooks.Open(str(marksheet.resolve()))

            for sheet_name, sheet_marks in self.marks.items():
                try:
                    ws = workbook.Worksheets(sheet_name)
                except Exception:
                    continue
                try:
                    ws.Unprotect()
                except Exception:
                    pass
                for row, value in sheet_marks.items():
                    cell = ws.Cells(row, 3)
                    try:
                        if cell.MergeCells:
                            cell = cell.MergeArea.Cells(1, 1)
                    except Exception:
                        pass
                    cell.Value = value
                    if value == 0:
                        cell.Interior.Color = 65535
                    else:
                        cell.Interior.Pattern = -4142
            workbook.Save()
            return True
        except Exception:
            return False

    @staticmethod
    def has_merged(ws, range_name: str) -> bool:
        return any(str(rng) == range_name for rng in ws.merged_cells.ranges)

    @staticmethod
    def formula(value, cell=None) -> str:
        text = ""
        if isinstance(value, str):
            text = value
        else:
            text = str(getattr(value, "text", "") or getattr(value, "ref", "") or "")
            if not text and cell is not None:
                text = str(getattr(cell, "_value", "") or "")
        return text.replace(" ", "").upper()

    @staticmethod
    def is_currency(cell) -> bool:
        fmt = str(cell.number_format or "").lower()
        return any(token in fmt for token in ("r", "$", "currency", "accounting"))

    @staticmethod
    def workbook_tab_is_red(path: Path, sheet_title: str) -> bool:
        import zipfile
        import xml.etree.ElementTree as ET

        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        try:
            with zipfile.ZipFile(path) as zf:
                root = ET.fromstring(zf.read("xl/workbook.xml"))
                for sheet in root.findall("main:sheets/main:sheet", ns):
                    if sheet.attrib.get("name") != sheet_title:
                        continue
                    color = sheet.find("main:tabColor", ns)
                    if color is None:
                        return False
                    rgb = (color.attrib.get("rgb") or "").upper()
                    indexed = color.attrib.get("indexed")
                    return rgb.endswith("FF0000") or indexed == "10"
        except Exception:
            return False
        return False

    @staticmethod
    def c2e2_matches_a1_visual(ws) -> bool:
        a1 = ws["A1"]
        for col in "CDE":
            cell = ws[f"{col}2"]
            same_fill = cell.fill.fill_type == a1.fill.fill_type and cell.fill.fgColor.rgb == a1.fill.fgColor.rgb
            same_border = cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style
            if not (cell.font.bold and cell.font.size == 14 and same_fill and same_border):
                return False
        return True

    @staticmethod
    def find_text_range(doc, text: str):
        rng = doc.Content.Duplicate
        finder = rng.Find
        finder.ClearFormatting()
        finder.Text = text
        return rng if finder.Execute() else None

    @staticmethod
    def find_paragraph(doc, starts_with: str, contains: str):
        for i in range(1, doc.Paragraphs.Count + 1):
            para = doc.Paragraphs(i)
            text = norm_text(para.Range.Text)
            if starts_with.lower() in text.lower() and contains.lower() in text.lower():
                return para
        return None

    @staticmethod
    def find_venue_paragraphs(doc):
        paragraphs = []
        collecting = False
        for i in range(1, doc.Paragraphs.Count + 1):
            para = doc.Paragraphs(i)
            text = norm_text(para.Range.Text)
            lower = text.lower()
            if lower.startswith("we opened"):
                collecting = True
            if collecting and text:
                paragraphs.append(para)
            if collecting and "south africa" in lower:
                break
        return paragraphs

    @staticmethod
    def paragraph_below_heading_is_justified(doc, heading_text: str) -> bool:
        for i in range(1, doc.Paragraphs.Count + 1):
            text = norm_text(doc.Paragraphs(i).Range.Text)
            if text.lower() != heading_text.lower():
                continue
            for j in range(i + 1, doc.Paragraphs.Count + 1):
                candidate = doc.Paragraphs(j)
                if norm_text(candidate.Range.Text):
                    return int(candidate.Alignment) == 3
        return False

    @staticmethod
    def only_em_dashes_in_target_text(text: str) -> bool:
        lower = text.lower()
        if "returned – and" in lower or "returned - and" in lower:
            return False
        if "returned — and" in lower:
            return True
        return "—" in text and "–" not in text

    @staticmethod
    def has_green_highlight(doc) -> bool:
        for i in range(1, doc.Words.Count + 1):
            try:
                if int(doc.Words(i).HighlightColorIndex) == 4:
                    return True
            except Exception:
                pass
        return False

    @staticmethod
    def top_page_image_width_is_6cm(doc) -> bool:
        candidates = []
        for i in range(1, doc.Shapes.Count + 1):
            try:
                shape = doc.Shapes(i)
                candidates.append((int(shape.Anchor.Start), float(shape.Top), float(shape.Width)))
            except Exception:
                pass
        for i in range(1, doc.InlineShapes.Count + 1):
            try:
                shape = doc.InlineShapes(i)
                candidates.append((int(shape.Range.Start), 9999.0, float(shape.Width)))
            except Exception:
                pass
        if not candidates:
            return False

        top_image = sorted(candidates, key=lambda item: (item[0], item[1]))[0]
        width_cm = top_image[2] / CM_TO_POINTS
        return abs(width_cm - 6.0) <= 0.2

    @staticmethod
    def images_have_no_border(doc) -> bool:
        if doc.InlineShapes.Count == 0:
            return False
        for i in range(1, doc.InlineShapes.Count + 1):
            try:
                if doc.InlineShapes(i).Line.Visible:
                    return False
            except Exception:
                pass
        return True

    @staticmethod
    def text_has_heading_or_title_style(doc, text: str) -> bool:
        rng = AutoMarker.find_text_range(doc, text)
        if not rng:
            return False
        try:
            style = str(rng.Style.NameLocal).lower()
            return "title" in style or "heading" in style
        except Exception:
            return False

    @staticmethod
    def red_sentence_case_under_isimangaliso(doc) -> bool:
        for i in range(1, doc.Paragraphs.Count + 1):
            heading = norm_text(doc.Paragraphs(i).Range.Text)
            if heading.lower() != "isimangaliso wetland park":
                continue
            for j in range(i + 1, doc.Paragraphs.Count + 1):
                paragraph = doc.Paragraphs(j).Range
                text = norm_text(paragraph.Text)
                if not text:
                    continue
                if text.lower().startswith("the v&a waterfront") or text.lower().startswith("two oceans"):
                    return False
                has_lowercase = any(ch.islower() for ch in text)
                not_all_caps = text != text.upper()
                red_words = 0
                checked_words = 0
                for k in range(1, paragraph.Words.Count + 1):
                    try:
                        word_text = norm_text(paragraph.Words(k).Text)
                        if not word_text:
                            continue
                        checked_words += 1
                        if is_red_color(paragraph.Words(k).Font.Color):
                            red_words += 1
                    except Exception:
                        pass
                mostly_red = checked_words > 0 and red_words / checked_words >= 0.8
                return has_lowercase and not_all_caps and mostly_red
        return False

    @staticmethod
    def has_page_number_fields(doc) -> bool:
        for section in doc.Sections:
            for index in range(1, 4):
                footer = section.Footers(index).Range
                for i in range(1, footer.Fields.Count + 1):
                    try:
                        code = str(footer.Fields(i).Code.Text).upper()
                        if "PAGE" in code or "NUMPAGES" in code:
                            return True
                    except Exception:
                        pass
        return False

    @staticmethod
    def page_background_is_white(doc) -> bool:
        try:
            for shape in doc.Background.Fill.ForeColor:
                _ = shape
        except Exception:
            pass
        try:
            return int(doc.Background.Fill.ForeColor.RGB) in (16777215, -1)
        except Exception:
            return True

    @staticmethod
    def red_text_ranges(doc):
        ranges = []
        for i in range(1, doc.Words.Count + 1):
            try:
                word = doc.Words(i)
                if norm_text(word.Text) and is_red_color(word.Font.Color):
                    ranges.append(word)
            except Exception:
                pass
        return ranges


def get_learners() -> list[tuple[str, Path]]:
    learners: list[tuple[str, Path]] = []
    for base in (FP_G2_DIR, FP_G3_DIR):
        if not base.exists():
            continue
        for folder in base.iterdir():
            if folder.is_dir() and not folder.name.startswith(("Custom", "DATA")):
                learners.append((folder.name, folder))
    return sorted(learners, key=lambda item: item[0].lower())


def main() -> int:
    if not MEMO_FILE.exists():
        print(f"Memo not found: {MEMO_FILE}")
        return 1
    if not MARKSHEETS_DIR.exists():
        print(f"Marksheets folder not found: {MARKSHEETS_DIR}")
        return 1
    if not win32com:
        print("pywin32 is required for the Word checks. Install with: pip install pywin32")
        return 1

    learners = get_learners()
    marked = 0
    for learner_name, learner_path in learners:
        print(f"Marking {learner_name}")
        word = WordReader()
        try:
            marker = AutoMarker(learner_name, learner_path, word)
            marker.mark()
            marked += 1
        except Exception as exc:
            print(f"  failed: {exc}")
        finally:
            word.close()

    print(f"Done. Marked {marked}/{len(learners)} learners.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
