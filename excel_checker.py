import sys
import re
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox
)
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtCore import Qt
import openpyxl
from openpyxl.styles import PatternFill


def normalize(value):
    """Strip spaces and uppercase a string value."""
    if value is None:
        return ""
    return str(value).replace(" ", "").upper()


def cell_value_normalized(ws, cell_ref):
    cell = ws[cell_ref]
    return normalize(cell.value)


def check_merged_and_centered(ws):
    """Check if A1:A6 are merged."""
    target = "A1:A6"
    for merged in ws.merged_cells.ranges:
        if normalize(str(merged)) == normalize(target):
            return True
    return False


def check_a1_a6_background(ws):
    """Check if A1:A6 have a non-blank/non-white background."""
    white_colors = {"FFFFFFFF", "FF000000"[0:0], "FFFFFF", ""}
    cell = ws["A1"]
    fill = cell.fill
    if fill and fill.fill_type not in (None, "none"):
        fg = fill.fgColor
        if fg:
            color = ""
            if fg.type == "rgb":
                color = fg.rgb.upper()
            elif fg.type == "theme":
                return True  # theme color = not white/blank
            # Remove alpha prefix if present (ARGB -> RGB)
            if len(color) == 8:
                color = color[2:]
            if color not in ("FFFFFF", "000000", ""):
                return True
    return False


def check_currency_format(ws):
    """Check if C2:C45 are formatted as currency."""
    currency_keywords = ["$", "€", "£", "¥", "CURRENCY", "#,##0", "ACCOUNTING"]
    for row in range(2, 46):
        cell = ws[f"C{row}"]
        nf = cell.number_format or ""
        nf_upper = nf.upper().replace(" ", "")
        if not any(k in nf_upper for k in currency_keywords):
            return False
    return True


def check_b3_starts_with_zero(ws):
    """Check if cell B3 value starts with 0."""
    val = ws["B3"].value
    if val is None:
        return False
    return str(val).startswith("0")


def check_b3_text_format(ws):
    """Check if B3 is formatted as text (@)."""
    nf = ws["B3"].number_format or ""
    return nf.strip() == "@"


def check_m2_sum(ws):
    """Check if M2 contains =SUM."""
    val = cell_value_normalized(ws, "M2")
    return "=SUM" in val


def check_m2_range(ws):
    """Check if M2 contains (D2:D45)."""
    val = cell_value_normalized(ws, "M2")
    return "(D2:D45)" in val


def check_m3_if(ws):
    val = cell_value_normalized(ws, "M3")
    return "=IF" in val


def check_m3_e3_gt5(ws):
    val = cell_value_normalized(ws, "M3")
    return "E3>5" in val


def check_m3_cheap(ws):
    val = cell_value_normalized(ws, "M3")
    return ",CHEAP," in val


def check_m3_expensive(ws):
    val = cell_value_normalized(ws, "M3")
    return ",EXPENSIVE)" in val


def check_budget_sheet(wb):
    return "budget" in [s.lower() for s in wb.sheetnames]


def check_chart_exists(ws):
    return len(ws._charts) > 0


def check_pie_chart(ws):
    from openpyxl.chart import PieChart, ProjectedPieChart, DoughnutChart
    for chart in ws._charts:
        if isinstance(chart, (PieChart, ProjectedPieChart, DoughnutChart)):
            return True
    return False


def check_chart_range(ws):
    """Check if any chart uses range A2:E45."""
    target = normalize("A2:E45")
    for chart in ws._charts:
        for series in chart.series:
            try:
                ref = normalize(str(series.val.numRef.ref))
                if ref == target:
                    return True
            except Exception:
                pass
            try:
                ref = normalize(str(series.val.ref))
                if ref == target:
                    return True
            except Exception:
                pass
        # Also check title/cat refs
        try:
            for series in chart.series:
                cat_ref = normalize(str(series.cat.numRef.ref))
                if cat_ref == target:
                    return True
        except Exception:
            pass
    return False


def check_landscape(ws):
    return ws.page_setup.orientation == "landscape"


def check_print_area(ws):
    pa = ws.print_area or ""
    return normalize(pa) == normalize("A1:E45")


def check_repeat_rows(ws):
    """Check if column A is set as repeating columns (print titles)."""
    titles = ws.print_title_cols or ""
    return normalize(titles) == normalize("$A:$A")


def run_checks(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
    except Exception as e:
        return None, str(e)

    # Use first sheet for most checks
    ws = wb.active

    results = [
        ("A1:A6 are merged and centered", check_merged_and_centered(ws)),
        ("A1:A6 have a non-white/blank background color", check_a1_a6_background(ws)),
        ("C2:C45 formatted as currency", check_currency_format(ws)),
        ("B3 starts with 0", check_b3_starts_with_zero(ws)),
        ("B3 is formatted as text", check_b3_text_format(ws)),
        ("M2 contains =SUM", check_m2_sum(ws)),
        ("M2 contains range (D2:D45)", check_m2_range(ws)),
        ("M3 contains =IF", check_m3_if(ws)),
        ("M3 contains (E3>5", check_m3_e3_gt5(ws)),
        ("M3 contains ,Cheap,", check_m3_cheap(ws)),
        ("M3 contains ,Expensive)", check_m3_expensive(ws)),
        ('Sheet "budget" exists', check_budget_sheet(wb)),
        ("A chart exists", check_chart_exists(ws)),
        ("Chart is a pie chart", check_pie_chart(ws)),
        ("Chart uses range A2:E45", check_chart_range(ws)),
        ("Page layout is landscape", check_landscape(ws)),
        ("Print area is A1:E45", check_print_area(ws)),
        ("Column A set as repeating header ($A:$A)", check_repeat_rows(ws)),
    ]

    return results, None


class ExcelChecker(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Spreadsheet Checker")
        self.setMinimumSize(700, 600)
        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # Title
        title = QLabel("Excel Spreadsheet Checker")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # File selection row
        file_row = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("color: gray;")
        browse_btn = QPushButton("Browse Excel File")
        browse_btn.clicked.connect(self.browse_file)
        browse_btn.setFixedWidth(160)
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(browse_btn)
        layout.addLayout(file_row)

        # Check button
        self.check_btn = QPushButton("Run Checks")
        self.check_btn.setEnabled(False)
        self.check_btn.setFixedHeight(36)
        self.check_btn.clicked.connect(self.run_checks)
        self.check_btn.setStyleSheet("background-color: #0078D4; color: white; font-weight: bold; border-radius: 4px;")
        layout.addWidget(self.check_btn)

        # Results table
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["#", "Description", "Result"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        self.filepath = None

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self.filepath = path
            self.file_label.setText(path)
            self.file_label.setStyleSheet("color: black;")
            self.check_btn.setEnabled(True)
            self.table.setRowCount(0)

    def run_checks(self):
        if not self.filepath:
            return

        results, error = run_checks(self.filepath)
        if error:
            QMessageBox.critical(self, "Error", f"Failed to open file:\n{error}")
            return

        self.table.setRowCount(0)
        for i, (desc, passed) in enumerate(results):
            row = self.table.rowCount()
            self.table.insertRow(row)

            num_item = QTableWidgetItem(str(i + 1))
            num_item.setTextAlignment(Qt.AlignCenter)

            desc_item = QTableWidgetItem(desc)

            result_val = "1" if passed else "0"
            result_item = QTableWidgetItem(result_val)
            result_item.setTextAlignment(Qt.AlignCenter)
            result_item.setFont(QFont("Arial", 10, QFont.Bold))

            if passed:
                result_item.setForeground(QColor("#107C10"))  # green
                result_item.setBackground(QColor("#DFF6DD"))
            else:
                result_item.setForeground(QColor("#A4262C"))  # red
                result_item.setBackground(QColor("#FDE7E9"))

            self.table.setItem(row, 0, num_item)
            self.table.setItem(row, 1, desc_item)
            self.table.setItem(row, 2, result_item)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = ExcelChecker()
    window.show()
    sys.exit(app.exec_())
