import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class CheckResult:
    status: str
    awarded: int | float | str
    reason: str


def normalize_formula(value: str) -> str:
    value = (value or "").replace("&amp;", "&")
    value = re.sub(r"\s+", "", value)
    return value.upper()


class Q4Workbook:
    def __init__(self, path: Path):
        self.path = path
        self.exists = path.exists()
        self.errors: List[str] = []

        self.sheet_title = ""
        self.freeze_panes = None
        self.table_exists = False
        self.table_ref = ""
        self.table_style = ""
        self.table_name = ""
        self.total_purchases_number_format = ""

        self.m3_formula = ""
        self.m4_formula = ""
        self.m3_value = None
        self.m4_value = None
        self.row3_formulas: Dict[str, str] = {}
        self.row4_formulas: Dict[str, str] = {}
        self.row3_values: Dict[str, str] = {}
        self.row4_values: Dict[str, str] = {}
        self.helper_column_formulas: Dict[str, List[str]] = {}
        self.expected_customer_id = ""
        self.expected_customer_id_row4 = ""
        self.expected_longest_city_len = 0

        if self.exists:
            try:
                self._load()
            except Exception as exc:  # pragma: no cover
                self.errors.append(f"Q4 parse error: {exc}")
        else:
            self.errors.append("Q4 file not found")

    def _load(self) -> None:
        wb = load_workbook(self.path, data_only=False)
        ws = wb[wb.sheetnames[0]]
        self.sheet_title = ws.title
        self.freeze_panes = ws.freeze_panes
        self.m3_value = ws["M3"].value
        self.m4_value = ws["M4"].value
        self.total_purchases_number_format = str(ws["G2"].number_format or "")
        for col in range(13, 21):
            coord3 = ws.cell(3, col).coordinate
            coord4 = ws.cell(4, col).coordinate
            value3 = ws.cell(3, col).value
            value4 = ws.cell(4, col).value
            self.row3_values[coord3] = "" if value3 is None else str(value3)
            self.row4_values[coord4] = "" if value4 is None else str(value4)
            if isinstance(value3, str) and value3.startswith("="):
                self.row3_formulas[coord3] = value3
            if isinstance(value4, str) and value4.startswith("="):
                self.row4_formulas[coord4] = value4
        for col in range(15, 21):
            col_letter = ws.cell(1, col).column_letter
            formulas: List[str] = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, col).value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append(value)
            if formulas:
                self.helper_column_formulas[col_letter] = formulas

        if ws.tables:
            t = next(iter(ws.tables.values()))
            self.table_exists = True
            self.table_ref = t.ref
            self.table_style = t.tableStyleInfo.name if t.tableStyleInfo else ""
            self.table_name = t.name or t.displayName or ""

        gender = str(ws["B3"].value or "")
        country = str(ws["C3"].value or "")
        age = ws["A3"].value
        self.expected_customer_id = f"{gender[:1]}-{country}{age}" if gender and country and age is not None else ""
        gender4 = str(ws["B4"].value or "")
        country4 = str(ws["C4"].value or "")
        age4 = ws["A4"].value
        self.expected_customer_id_row4 = f"{gender4[:1]}-{country4}{age4}" if gender4 and country4 and age4 is not None else ""

        city_values = [str(ws[f"D{row}"].value or "") for row in range(2, ws.max_row + 1) if ws[f"D{row}"].value is not None]
        self.expected_longest_city_len = max((len(v) for v in city_values), default=0)

        self._load_raw_formulas()

    def _load_raw_formulas(self) -> None:
        with zipfile.ZipFile(self.path) as zf:
            xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="ignore")
        self.m3_formula = self._extract_formula(xml, "M3")
        self.m4_formula = self._extract_formula(xml, "M4")

    def _extract_formula(self, xml: str, cell_ref: str) -> str:
        match = re.search(rf'<c[^>]*r="{cell_ref}"[^>]*>(.*?)</c>', xml)
        if not match:
            return ""
        formula_match = re.search(r"<f[^>]*>(.*?)</f>", match.group(0))
        return formula_match.group(1) if formula_match else ""


def evaluate_q4_check(doc: Q4Workbook, check: Dict) -> CheckResult:
    if not doc.exists:
        return CheckResult("manual", "", "Q4 file missing")
    if doc.errors:
        return CheckResult("manual", "", "; ".join(doc.errors))

    desc = check["description"]
    mark = check["mark"]
    m3 = normalize_formula(doc.m3_formula)
    m4 = normalize_formula(doc.m4_formula)
    row3_formulas = {cell: normalize_formula(value) for cell, value in doc.row3_formulas.items()}
    row4_formulas = {cell: normalize_formula(value) for cell, value in doc.row4_formulas.items()}
    expected_id = str(doc.expected_customer_id or "").upper()
    expected_id_row4 = str(doc.expected_customer_id_row4 or "").upper()

    def pass_fail(ok: bool, ok_reason: str, fail_reason: str) -> CheckResult:
        return CheckResult("pass" if ok else "fail", mark if ok else 0, ok_reason if ok else fail_reason)

    def number_format_ok() -> bool:
        fmt = doc.total_purchases_number_format
        return fmt in {"0.0", "#,##0.0"}

    def formula_texts(*groups: Dict[str, str]) -> List[str]:
        texts = []
        for group in groups:
            texts.extend(group.values())
        return texts

    def any_formula_has(*tokens: str, groups: Optional[List[Dict[str, str]]] = None) -> bool:
        scan_groups = groups or [row3_formulas, row4_formulas]
        for formula in formula_texts(*scan_groups):
            if all(token in formula for token in tokens):
                return True
        return False

    def any_row_value(*tokens: str) -> bool:
        values = [value.upper() for value in list(doc.row3_values.values()) + list(doc.row4_values.values())]
        return any(all(token in value for token in tokens) for value in values)

    def m3_references_helper_cells() -> bool:
        return any(ref in m3 for ref in ("O2", "O3", "O4", "P2", "P3", "P4", "Q2", "Q3", "Q4", "R2", "R3", "R4", "S2", "S3", "S4", "T2", "T3", "T4"))

    def m3_has_concat() -> bool:
        return (
            "CONCATENATE(" in m3
            or "_XLFN.CONCAT(" in m3
            or "TEXTJOIN(" in m3
            or "&" in m3
            or (m3_references_helper_cells() and ("CONCATENATE(" in m3 or "_XLFN.CONCAT(" in m3 or "&" in m3))
        )

    def m3_has_left() -> bool:
        return "LEFT(" in m3 or "MID(" in m3 or any_formula_has("LEFT(", groups=[row3_formulas])

    def m3_has_gender_ref() -> bool:
        return any(
            token in m3
            for token in ("[GENDER]", "[@GENDER]", "[[#THISROW],[GENDER]]", "B3", "B4")
        ) or any_formula_has("LEFT(", "B3", groups=[row3_formulas]) or any_formula_has("LEFT(", "B4", groups=[row4_formulas]) or any_formula_has("LEFT(", "[GENDER]", groups=[row3_formulas, row4_formulas])

    def m3_has_length_one() -> bool:
        return ",1)" in m3 or ", 1)" in m3 or any_formula_has(",1)", groups=[row3_formulas]) or any_formula_has(", 1)", groups=[row3_formulas])

    def m3_has_dash() -> bool:
        return (
            '"-"' in m3
            or "'-'" in m3
            or "-&" in m3
            or '&"-"&' in m3
            or any_row_value("-")
        )

    def m3_has_country() -> bool:
        return (
            "COUNTRY" in m3
            or "C3" in m3
            or "C4" in m3
            or "P3" in m3
            or "Q3" in m3
            or "Q4" in m3
            or "R3" in m3
            or any_formula_has("[COUNTRY]", groups=[row3_formulas, row4_formulas])
            or any_formula_has("C3", groups=[row3_formulas, row4_formulas])
            or any_formula_has("C4", groups=[row3_formulas, row4_formulas])
        )

    def m3_has_age() -> bool:
        return (
            "AGE" in m3
            or "A3" in m3
            or "A4" in m3
            or "36" in m3
            or "R3" in m3
            or "R4" in m3
            or any_formula_has("[AGE]", groups=[row3_formulas, row4_formulas])
            or any_formula_has("A3", groups=[row3_formulas, row4_formulas])
            or any_formula_has("A4", groups=[row3_formulas, row4_formulas])
        )

    def m3_result_ok() -> bool:
        value = str(doc.m3_value or "").upper()
        return value in {expected_id, expected_id_row4}

    def m3_uses_valid_row_refs() -> bool:
        return any(ref in m3 for ref in ("A3", "B3", "C3", "A4", "B4", "C4", "[[#THISROW],"))

    def m3_formula_ok() -> bool:
        accepted = {
            'LEFT(B3,1)&"-"&C3&A3',
            'LEFT(B4,1)&"-"&C4&A4',
            'CONCATENATE(LEFT(B3,1),"-",C3,A3)',
            'CONCATENATE(LEFT(B4,1),"-",C4,A4)',
        }
        if m3 in accepted:
            return True
        if all(token in m3 for token in ("LEFT(", '"-"')) and m3_has_country() and m3_has_age() and m3_uses_valid_row_refs():
            return True
        if m3_has_concat() and m3_has_left() and m3_has_gender_ref() and m3_has_length_one() and m3_has_dash() and m3_has_country() and m3_has_age():
            return True
        return False

    def m4_has_max() -> bool:
        return "MAX(" in m4 or "LARGE(" in m4

    def m4_has_len() -> bool:
        if "LEN(" in m4:
            return True
        return any("LEN(" in normalize_formula(formula) for formulas in doc.helper_column_formulas.values() for formula in formulas)

    def m4_has_city_ref() -> bool:
        if "CITY" in m4 or "D2:D61" in m4 or "D2:D60" in m4 or "D:D" in m4 or "D19" in m4 or "D56" in m4:
            return True
        for formulas in doc.helper_column_formulas.values():
            for formula in formulas:
                normalized = normalize_formula(formula)
                if "LEN(D" in normalized or "LEN(TABLE1[CITY])" in normalized or "LEN(TABLE2[CITY])" in normalized or "LEN(TABLE3[CITY])" in normalized:
                    return True
        return False

    def m4_result_ok() -> bool:
        return str(doc.m4_value or "") == str(doc.expected_longest_city_len)

    def m4_formula_ok() -> bool:
        if m4_result_ok() and m4_has_max() and (m4_has_len() or "O2:O61" in m4 or "O:O" in m4):
            return True
        accepted = {
            "MAX(LEN(D2:D61))",
            "MAX(LEN(D2:D60))",
            "MAX(LEN(TABLE1[CITY]))",
            "MAX(LEN(TABLE2[CITY]))",
            "MAX(LEN(TABLE3[CITY]))",
            "MAX(O2:O61)",
            "LARGE(O2:O61,1)",
        }
        return m4 in accepted

    def m4_max_len_ok() -> bool:
        accepted = {
            "MAX(LEN(D2:D61))",
            "MAX(LEN(D2:D60))",
            "MAX(LEN(TABLE1[CITY]))",
            "MAX(LEN(TABLE2[CITY]))",
            "MAX(LEN(TABLE3[CITY]))",
        }
        return m4 in accepted or m4_has_max() or m4_has_len()

    mapping = {
        "dataset into a table": lambda: pass_fail(
            doc.table_exists,
            "Worksheet contains a table",
            "Worksheet table not detected",
        ),
        "converted into a table": lambda: pass_fail(
            doc.table_exists,
            "Worksheet contains a table",
            "Worksheet table not detected",
        ),
        "with the ‘Dark Teal Medium 2’ table style": lambda: pass_fail(
            doc.table_style == "TableStyleMedium2",
            "Table style matches Dark Teal Medium 2",
            f"Table style is {doc.table_style!r}",
        ),
        "freeze panes": lambda: pass_fail(
            bool(doc.freeze_panes),
            "Freeze panes applied",
            "Freeze panes not detected",
        ),
        "panes are frozen at Freeze cell A2": lambda: pass_fail(
            str(doc.freeze_panes) == "A2",
            "Freeze panes set at A2",
            f"Freeze panes set at {doc.freeze_panes!r}",
        ),
        "cell formatting": lambda: pass_fail(
            number_format_ok(),
            "Total_Purchases column uses one-decimal number format",
            f"Total_Purchases format is {doc.total_purchases_number_format!r}",
        ),
        "values in Total_Purchases to the Number format": lambda: pass_fail(
            number_format_ok(),
            "Total_Purchases column uses number format",
            f"Total_Purchases format is {doc.total_purchases_number_format!r}",
        ),
        "display with one decimal place": lambda: pass_fail(
            number_format_ok(),
            "Total_Purchases column displays one decimal place",
            f"Total_Purchases format is {doc.total_purchases_number_format!r}",
        ),
        "function in cell M": lambda: pass_fail(
            bool(m3),
            "Formula detected in M3",
            "Formula not detected in M3",
        ),
        '=CONCATENATE(LEFT([Gender],1),"-",[country],[age])': lambda: pass_fail(
            m3_formula_ok() or m3_result_ok(),
            "M3 generates the expected customer ID",
            f"M3 formula/value are {doc.m3_formula!r} / {doc.m3_value!r}",
        ),
        "Function CONCATENATE": lambda: pass_fail(
            m3_has_concat(),
            "M3 uses concatenation",
            f"M3 formula is {doc.m3_formula!r}",
        ),
        "The first letter of the gender:": lambda: pass_fail(
            m3_formula_ok() or m3_result_ok(),
            "M3 uses the first letter of Gender",
            f"M3 formula/value are {doc.m3_formula!r} / {doc.m3_value!r}",
        ),
        "LEFT (Table1[@Gender],1": lambda: pass_fail(
            m3_has_left() or m3_result_ok(),
            "M3 extracts the first gender letter",
            f"M3 formula is {doc.m3_formula!r}",
        ),
        "a dash (-),": lambda: pass_fail(
            m3_formula_ok() or m3_result_ok(),
            "M3 includes a dash",
            f"M3 formula/value are {doc.m3_formula!r} / {doc.m3_value!r}",
        ),
        "the country: Table1[@Country]": lambda: pass_fail(
            m3_formula_ok() or m3_result_ok(),
            "M3 includes country",
            f"M3 formula/value are {doc.m3_formula!r} / {doc.m3_value!r}",
        ),
        "the customer’s age Table1[@Age].": lambda: pass_fail(
            m3_formula_ok() or m3_result_ok(),
            "M3 includes age",
            f"M3 formula/value are {doc.m3_formula!r} / {doc.m3_value!r}",
        ),
        "function in cell M4": lambda: pass_fail(
            bool(m4),
            "Formula detected in M4",
            "Formula not detected in M4",
        ),
        "=MAX(LEN([City))": lambda: pass_fail(
            m4_has_max(),
            "M4 uses MAX or LARGE to return the longest city-name length",
            f"M4 formula/value are {doc.m4_formula!r} / {doc.m4_value!r}",
        ),
        "correct function/operation: MAX": lambda: pass_fail(
            m4_has_max(),
            "M4 uses MAX or LARGE",
            f"M4 formula is {doc.m4_formula!r}",
        ),
        "correct function/operation: LEN": lambda: pass_fail(
            m4_has_len(),
            "LEN is used directly or in helper cells",
            f"M4/helper formulas are {doc.m4_formula!r} / {doc.helper_column_formulas!r}",
        ),
        "correct cell/cell range: Table1[City]": lambda: pass_fail(
            m4_has_city_ref(),
            "Formula references City directly or through helper cells",
            f"M4 formula is {doc.m4_formula!r}",
        ),
    }

    if desc.startswith("The first letter of the gender"):
        return pass_fail(
            m3_has_left(),
            "LEFT is used for the first letter of Gender",
            f"M3/helper formulas are {doc.m3_formula!r} / {doc.row3_formulas!r}",
        )
    if "award mark for the length" in desc or desc.endswith(",1"):
        return pass_fail(
            m3_has_length_one(),
            "LEFT uses a length argument of 1",
            f"M3/helper formulas are {doc.m3_formula!r} / {doc.row3_formulas!r}",
        )
    if desc.startswith("LEFT (Table1[@Gender]") or desc.startswith("LEFT (Table2[@Gender]") or "followed by parameter" in desc:
        return pass_fail(
            m3_has_gender_ref(),
            "Gender field/cell is referenced for the LEFT function",
            f"M3/helper formulas are {doc.m3_formula!r} / {doc.row3_formulas!r}",
        )
    if desc == "Function CONCATENATE":
        return pass_fail(
            m3_has_concat(),
            "M3 uses concatenation",
            f"M3 formula is {doc.m3_formula!r}",
        )
    if desc == "a dash (-),":
        return pass_fail(
            m3_has_dash(),
            "Dash separator is present",
            f"M3/helper values are {doc.m3_formula!r} / {doc.row3_values!r}",
        )
    if desc == "the country: Table1[@Country]":
        return pass_fail(
            m3_has_country(),
            "Country is included",
            f"M3/helper formulas are {doc.m3_formula!r} / {doc.row3_formulas!r}",
        )
    if desc == "the customer’s age Table1[@Age].":
        return pass_fail(
            m3_has_age(),
            "Age is included",
            f"M3/helper formulas are {doc.m3_formula!r} / {doc.row3_formulas!r}",
        )

    if desc in mapping:
        return mapping[desc]()
    return CheckResult("manual", "", f"Q4 actual checker not implemented for {desc}")
